#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Importa la lista de precios de Abelson (Excel) al endpoint de scraping de OBYRA.

N8N descarga el Excel de Google Drive cada 48h y llama a este script pasandole la
ruta del archivo:

    python scripts/scraping_abelson.py "/ruta/listaPrecios.xlsx"

Lee la solapa 'bd' (ART_CODIGO, ART_DESCRI, ART_PREVT1), arma lotes de 500 y los
POSTea a /presupuestos/precio-scraping con el token. El endpoint hace upsert por
(proveedor, material, unidad, zona), asi que correr esto dos veces NO duplica ni
infla el promedio: es idempotente.

Unidades: la lista NO trae columna de unidad y el precio es por unidad de VENTA
(un cano de 6.40m se cotiza entero). Detectamos solo envases-contenedor claros
(bolsa, balde, tambor, bidon, rollo, caja) y el resto va como 'un'. NUNCA se asigna
kg/m3/ml a la fuerza desde un token tipo "x 50 KG": eso describe el contenido, no la
unidad de venta, y cargarlo como $/kg es exactamente el bug que puso el adhesivo a
$146.834/kg. Una unidad 'un' que el APU no matchea queda inerte (inofensiva); una
'kg' mal puesta envenena la base. El server ademas tiene su propio guard.

Env vars:
    SCRAPING_TOKEN     (requerido)  Bearer token del endpoint.
    OBYRA_SCRAPING_URL (opcional)   default https://app.obyra.com.ar/presupuestos/precio-scraping
    SLACK_WEBHOOK_URL  (opcional)   si un lote falla tras 3 reintentos, avisa aca.
    EXCEL_PATH         (opcional)   ruta del Excel si no se pasa como argumento.
    PROVEEDOR          (opcional)   default 'Abelson'.
    ZONA               (opcional)   default 'Buenos Aires'.

Exit code 0 si todos los lotes entraron; 1 si alguno fallo definitivamente o si hay
un error fatal (sin token, sin archivo, solapa 'bd' faltante).
"""
import json
import os
import re
import sys
import time
import urllib.error
import urllib.request

SHEET = 'bd'
BATCH = 500
LOTE_MAX_SERVER = 2000          # tope del endpoint; BATCH debe ser <= esto
TIMEOUT = 180
REINTENTOS = 3

URL = os.getenv('OBYRA_SCRAPING_URL',
                'https://app.obyra.com.ar/presupuestos/precio-scraping')
TOKEN = os.getenv('SCRAPING_TOKEN', '')
SLACK = os.getenv('SLACK_WEBHOOK_URL', '')
PROVEEDOR = os.getenv('PROVEEDOR', 'Abelson')
ZONA = os.getenv('ZONA', 'Buenos Aires')

# Envases-contenedor que SI son una unidad de venta valida (no describen contenido).
_CONTENEDORES = [
    ('bolsa', ('BOLSA',)),
    ('balde', ('BALDE',)),
    ('tambor', ('TAMBOR',)),
    ('bidon', ('BIDON', 'BIDÓN')),
    ('rollo', ('ROLLO',)),
    ('caja', ('CAJA', 'CJA ')),
    ('juego', ('JUEGO', 'JGO')),
]
# "VALE POR PALLET ..." y similares no son productos reales -> se saltean.
_RE_VALE = re.compile(r'^\s*VALE\b|\bVALE POR PALLET\b')


def log(msg):
    print(msg, flush=True)


def detectar_unidad(desc_upper):
    """Unidad de venta a partir de la descripcion. Conservador: solo contenedores
    claros, el resto 'un'. Nunca infiere kg/m3/ml (vector de envenenamiento)."""
    for unidad, tokens in _CONTENEDORES:
        if any(t in desc_upper for t in tokens):
            return unidad
    return 'un'


def normalizar_desc(desc):
    """Normalizacion liviana del lado cliente (el server hace la canonica igual):
    trim, colapsar espacios, m2/m3. Se manda la descripcion; el server la vuelve a
    normalizar para la clave de material."""
    s = str(desc).strip()
    s = s.replace('²', '2').replace('³', '3')  # m2 / m3
    s = re.sub(r'\s+', ' ', s)
    return s


def leer_items(path):
    """Devuelve (items, stats). Cada item es el JSON que espera el endpoint."""
    try:
        import openpyxl
    except ImportError:
        log('[FATAL] falta openpyxl (pip install openpyxl)')
        raise

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"la solapa '{SHEET}' no existe. Solapas: {wb.sheetnames}")
    ws = wb[SHEET]

    filas = ws.iter_rows(values_only=True)
    header = next(filas, None)
    # Ubicar columnas por nombre (robusto ante reordenamiento), con fallback 0/1/2.
    idx = {'cod': 0, 'desc': 1, 'precio': 2}
    if header:
        low = [str(h).strip().upper() if h is not None else '' for h in header]
        for i, h in enumerate(low):
            if h == 'ART_CODIGO':
                idx['cod'] = i
            elif h == 'ART_DESCRI':
                idx['desc'] = i
            elif h == 'ART_PREVT1':
                idx['precio'] = i

    items = []
    stats = {'filas': 0, 'sin_desc': 0, 'vale_pallet': 0, 'precio_cero': 0,
             'precio_malo': 0, 'validos': 0, 'unidades': {}}
    for row in filas:
        stats['filas'] += 1
        desc = row[idx['desc']] if len(row) > idx['desc'] else None
        precio_raw = row[idx['precio']] if len(row) > idx['precio'] else None
        if desc is None or not str(desc).strip():
            stats['sin_desc'] += 1
            continue
        desc_upper = str(desc).upper()
        if _RE_VALE.search(desc_upper):
            stats['vale_pallet'] += 1
            continue
        try:
            precio = float(precio_raw) if precio_raw not in (None, '') else 0.0
        except (TypeError, ValueError):
            stats['precio_malo'] += 1
            continue
        if precio <= 0:
            stats['precio_cero'] += 1
            continue

        unidad = detectar_unidad(desc_upper)
        stats['unidades'][unidad] = stats['unidades'].get(unidad, 0) + 1
        stats['validos'] += 1
        items.append({
            'material': normalizar_desc(desc),
            'precio_unitario': round(precio, 2),
            'unidad': unidad,
            'proveedor': PROVEEDOR,
            'zona': ZONA,
            'fuente': 'scraping',   # el server lo fuerza igual; explicito por claridad
        })
    wb.close()
    return items, stats


def post_lote(items):
    """POST de un lote con reintentos. Devuelve (ok, resultado|error_str)."""
    payload = json.dumps({'items': items}).encode('utf-8')
    ultimo_error = None
    for intento in range(1, REINTENTOS + 1):
        try:
            req = urllib.request.Request(
                URL, data=payload, method='POST',
                headers={'Content-Type': 'application/json',
                         'Authorization': 'Bearer ' + TOKEN})
            with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
                return True, json.loads(resp.read().decode('utf-8'))
        except urllib.error.HTTPError as e:
            detalle = e.read().decode('utf-8', 'replace')[:300]
            ultimo_error = f'HTTP {e.code}: {detalle}'
            # 401/400/413: no se arregla reintentando (token/validacion) -> cortar.
            if e.code not in (429, 500, 502, 503, 504):
                break
        except Exception as e:  # noqa: BLE001 (timeout, conexion, DNS, JSON)
            ultimo_error = f'{type(e).__name__}: {e}'
        if intento < REINTENTOS:
            time.sleep(2 * intento)
    return False, ultimo_error


def slack(texto):
    if not SLACK:
        return
    try:
        req = urllib.request.Request(
            SLACK, data=json.dumps({'text': texto}).encode('utf-8'),
            headers={'Content-Type': 'application/json'}, method='POST')
        urllib.request.urlopen(req, timeout=30).read()
    except Exception as e:  # noqa: BLE001
        log(f'  [slack fallo] {type(e).__name__}: {e}')


def main():
    if not TOKEN:
        log('[FATAL] falta la env var SCRAPING_TOKEN')
        return 1
    path = (sys.argv[1] if len(sys.argv) > 1 else os.getenv('EXCEL_PATH', '')).strip()
    if not path:
        log('[FATAL] falta la ruta del Excel (argumento 1 o env EXCEL_PATH)')
        return 1
    if not os.path.isfile(path):
        log(f'[FATAL] no existe el archivo: {path}')
        return 1

    log(f'== Scraping {PROVEEDOR} -> {URL}')
    log(f'   archivo: {path}')
    try:
        items, stats = leer_items(path)
    except Exception as e:  # noqa: BLE001
        log(f'[FATAL] no se pudo leer el Excel: {type(e).__name__}: {e}')
        slack(f':x: {PROVEEDOR}: no se pudo leer el Excel — {type(e).__name__}: {e}')
        return 1

    log(f'   filas leidas: {stats["filas"]} | validos: {stats["validos"]} | '
        f'sin_desc: {stats["sin_desc"]} | vale/pallet: {stats["vale_pallet"]} | '
        f'precio<=0: {stats["precio_cero"]} | precio_malo: {stats["precio_malo"]}')
    log(f'   unidades detectadas: {stats["unidades"]}')

    if not items:
        log('[WARN] no hay items validos para enviar')
        slack(f':warning: {PROVEEDOR}: 0 items validos en la lista '
              f'({stats["filas"]} filas leidas).')
        return 0

    total = {'lotes': 0, 'lotes_ok': 0, 'lotes_fallidos': 0, 'recibidos': 0,
             'guardados': 0, 'lista_proveedor': 0, 'ignorados': 0,
             'envase_sospechoso': 0, 'curados_preservados': 0}
    ejemplos_error = []
    n_lotes = (len(items) + BATCH - 1) // BATCH

    for i in range(0, len(items), BATCH):
        lote = items[i:i + BATCH]
        nro = i // BATCH + 1
        total['lotes'] += 1
        ok, res = post_lote(lote)
        if not ok:
            total['lotes_fallidos'] += 1
            log(f'   lote {nro}/{n_lotes}: FALLO -> {res}')
            slack(f':x: {PROVEEDOR}: lote {nro}/{n_lotes} fallo tras '
                  f'{REINTENTOS} intentos — {res}')
            continue
        total['lotes_ok'] += 1
        for k in ('recibidos', 'guardados', 'lista_proveedor', 'ignorados',
                  'envase_sospechoso', 'curados_preservados'):
            total[k] += int(res.get(k, 0) or 0)
        for err in (res.get('errores') or []):
            if len(ejemplos_error) < 15:
                ejemplos_error.append(err)
        log(f'   lote {nro}/{n_lotes}: ok | recibidos {res.get("recibidos")} '
            f'guardados {res.get("guardados")} lista {res.get("lista_proveedor")} '
            f'ignorados {res.get("ignorados")} envase {res.get("envase_sospechoso")}')

    log('== Resumen')
    for k in ('lotes', 'lotes_ok', 'lotes_fallidos', 'recibidos', 'guardados',
              'lista_proveedor', 'ignorados', 'envase_sospechoso',
              'curados_preservados'):
        log(f'   {k}: {total[k]}')
    if ejemplos_error:
        log('   ejemplos de items ignorados por el server:')
        for e in ejemplos_error[:10]:
            log(f'     - {e.get("material")}: {e.get("motivo")}')

    exito = total['lotes_fallidos'] == 0
    icono = ':white_check_mark:' if exito else ':x:'
    resumen = (
        f'{icono} {PROVEEDOR}: {total["guardados"]} guardados '
        f'({total["lista_proveedor"]} a lista de proveedor) de '
        f'{stats["validos"]} validos. Ignorados {total["ignorados"]}, '
        f'envase sospechoso {total["envase_sospechoso"]}. '
        f'Lotes {total["lotes_ok"]}/{total["lotes"]} ok.')
    if total['lotes_fallidos']:
        resumen += f' :warning: {total["lotes_fallidos"]} lote(s) fallaron.'
    slack(resumen)
    log('== ' + resumen)

    return 0 if exito else 1


if __name__ == '__main__':
    sys.exit(main())
