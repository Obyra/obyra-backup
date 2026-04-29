"""Importacion del directorio global de proveedores curado por OBYRA.

Lee `OBYRA_directorio_proveedores_v2.xlsx` (hoja `Todos para importar`) y
upserta cada fila como `ProveedorOC` con `scope='global'` y
`organizacion_id IS NULL`. Idempotencia por `external_key`:

    external_key = slug(razon_social) + '-' + slug(provincia)

Re-correr el script ACTUALIZA los registros existentes, NO duplica.

Uso:
    python -m scripts.import_proveedores_globales [--archivo PATH] [--hoja NOMBRE] [--dry-run]

Por defecto:
    archivo = OBYRA_directorio_proveedores_v2.xlsx (en el cwd)
    hoja    = 'Todos para importar'
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import unicodedata
from typing import Optional

# Cargar variables de entorno desde .env si existe
try:
    from dotenv import load_dotenv  # type: ignore
    load_dotenv()
except Exception:
    pass


# ----------------------------------------------------------------------
# Helpers de normalizacion
# ----------------------------------------------------------------------

def slugify(text: Optional[str]) -> str:
    """Normaliza un texto a slug ASCII compacto. Vacio si text es None/blanco."""
    if text is None:
        return ''
    s = str(text).strip().lower()
    if not s:
        return ''
    # Quitar acentos
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    # Reemplazar todo lo no alfanumerico por '-'
    s = re.sub(r'[^a-z0-9]+', '-', s)
    s = re.sub(r'-+', '-', s).strip('-')
    return s


def limpiar_consultar(valor: Optional[str]) -> Optional[str]:
    """Si el texto es 'Consultar...' o variantes, devuelve None.

    El Excel marca campos sin dato con 'Consultar...', 'Consultar', '-' etc.
    """
    if valor is None:
        return None
    s = str(valor).strip()
    if not s:
        return None
    low = s.lower()
    if low.startswith('consultar') or low in ('-', 'n/a', 'na', 's/d', 'sin datos'):
        return None
    return s


def normalizar_email(valor: Optional[str]) -> Optional[str]:
    s = limpiar_consultar(valor)
    if not s:
        return None
    s = s.strip().lower()
    if '@' not in s:
        return None
    return s[:200]


def truncar(valor: Optional[str], maxlen: int) -> Optional[str]:
    if valor is None:
        return None
    s = str(valor).strip()
    if not s:
        return None
    return s[:maxlen]


# ----------------------------------------------------------------------
# Carga
# ----------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(description='Importa el directorio global de proveedores OBYRA.')
    parser.add_argument('--archivo', default='OBYRA_directorio_proveedores_v2.xlsx',
                        help='Ruta al archivo Excel.')
    parser.add_argument('--hoja', default='Todos para importar',
                        help='Nombre de la hoja a leer.')
    parser.add_argument('--dry-run', action='store_true',
                        help='No commitea cambios, solo reporta lo que haria.')
    parser.add_argument('--verbose', action='store_true',
                        help='Imprime una linea por cada proveedor procesado.')
    return parser.parse_args()


def importar_directorio_global(archivo_o_stream, hoja='Todos para importar', dry_run=False, verbose=False, log=print):
    """Importa el directorio global de proveedores desde un Excel.

    `archivo_o_stream` puede ser:
      - Una ruta absoluta a un .xlsx (str/Path).
      - Un file-like object (BytesIO) — útil para uploads HTTP.
      - Un FileStorage de Flask (tiene `.stream` o `.read()`).

    Usa la sesion `db` activa (de Flask). REQUIERE estar dentro de un
    `app.app_context()`. NO crea ni cierra la app.

    Retorna dict: {creados, actualizados, skipped, zonas_creadas, errores}
    donde `errores` es lista de (fila, mensaje).
    """
    import openpyxl
    from extensions import db
    from models.proveedores_oc import ProveedorOC, Zona

    # Aceptar FileStorage (flask) o file-like
    if hasattr(archivo_o_stream, 'stream'):
        archivo_in = archivo_o_stream.stream
    else:
        archivo_in = archivo_o_stream

    log(f'[IMPORT] Leyendo Excel - hoja {hoja!r}')
    wb = openpyxl.load_workbook(archivo_in, data_only=True, read_only=True)
    if hoja not in wb.sheetnames:
        raise ValueError(f'La hoja {hoja!r} no existe. Hojas: {wb.sheetnames}')
    ws = wb[hoja]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)

    def encontrar_col(nombres):
        for i, h in enumerate(header):
            if h is None:
                continue
            norm = slugify(h)
            for n in nombres:
                if norm == slugify(n):
                    return i
        return None

    col_empresa      = encontrar_col(['Empresa', 'Razon Social', 'Razon social'])
    col_categoria    = encontrar_col(['Categoria'])
    col_subcategoria = encontrar_col(['Subcategoria'])
    col_tier         = encontrar_col(['Tier'])
    col_provincia    = encontrar_col(['Provincia'])
    col_zona         = encontrar_col(['Ubicacion Zona', 'Zona'])
    col_zona_det     = encontrar_col(['Ubicacion Detalle', 'Zona Detalle'])
    col_cobertura    = encontrar_col(['Cobertura'])
    col_web          = encontrar_col(['Web', 'Sitio Web'])
    col_telefono     = encontrar_col(['Telefono'])
    col_email        = encontrar_col(['Email', 'Mail'])
    col_direccion    = encontrar_col(['Direccion'])
    col_alianza      = encontrar_col(['Tipo Alianza', 'Alianza'])
    col_notas        = encontrar_col(['Notas', 'Observaciones'])

    if col_empresa is None or col_provincia is None:
        raise ValueError('Faltan columnas obligatorias Empresa / Provincia.')

    creados = 0
    actualizados = 0
    skipped = 0
    zonas_creadas = 0
    errores = []

    zonas_cache = {z.slug: z for z in Zona.query.all()}

    for i, row in enumerate(rows, start=2):
        try:
            empresa = limpiar_consultar(row[col_empresa] if col_empresa is not None else None)
            if not empresa:
                skipped += 1
                continue

            provincia = limpiar_consultar(row[col_provincia]) if col_provincia is not None else None

            key_base = slugify(empresa)
            key_prov = slugify(provincia) if provincia else 'sin-provincia'
            external_key = f'{key_base}-{key_prov}'[:160]
            if not key_base:
                skipped += 1
                continue

            zona_id = None
            if col_zona is not None:
                zona_nombre = limpiar_consultar(row[col_zona])
                if zona_nombre:
                    zona_slug = slugify(zona_nombre)
                    zona = zonas_cache.get(zona_slug)
                    if not zona:
                        zona = Zona(
                            nombre=zona_nombre[:120],
                            slug=zona_slug[:140],
                            provincia=truncar(provincia, 100),
                        )
                        db.session.add(zona)
                        db.session.flush()
                        zonas_cache[zona_slug] = zona
                        zonas_creadas += 1
                    zona_id = zona.id

            campos = dict(
                razon_social=truncar(empresa, 200),
                categoria=truncar(limpiar_consultar(row[col_categoria]) if col_categoria is not None else None, 120),
                subcategoria=truncar(limpiar_consultar(row[col_subcategoria]) if col_subcategoria is not None else None, 160),
                tier=truncar(limpiar_consultar(row[col_tier]) if col_tier is not None else None, 20),
                provincia=truncar(provincia, 100),
                zona_id=zona_id,
                ubicacion_detalle=truncar(limpiar_consultar(row[col_zona_det]) if col_zona_det is not None else None, 255),
                cobertura=truncar(limpiar_consultar(row[col_cobertura]) if col_cobertura is not None else None, 255),
                web=truncar(limpiar_consultar(row[col_web]) if col_web is not None else None, 300),
                telefono=truncar(limpiar_consultar(row[col_telefono]) if col_telefono is not None else None, 50),
                email=normalizar_email(row[col_email]) if col_email is not None else None,
                direccion=truncar(limpiar_consultar(row[col_direccion]) if col_direccion is not None else None, 300),
                tipo_alianza=truncar(limpiar_consultar(row[col_alianza]) if col_alianza is not None else None, 80),
                notas=limpiar_consultar(row[col_notas]) if col_notas is not None else None,
            )

            prov = ProveedorOC.query.filter_by(scope='global', external_key=external_key).first()
            if prov:
                for k, v in campos.items():
                    setattr(prov, k, v)
                actualizados += 1
                if verbose:
                    log(f'  [UPDATE] {external_key}')
            else:
                prov = ProveedorOC(
                    scope='global',
                    organizacion_id=None,
                    external_key=external_key,
                    activo=True,
                    tipo='materiales',
                    **campos,
                )
                db.session.add(prov)
                creados += 1
                if verbose:
                    log(f'  [CREATE] {external_key}')

        except Exception as e:
            errores.append((i, str(e)))
            log(f'  [ERROR fila {i}] {e}')

    if dry_run:
        log('[DRY-RUN] Rollback - no se commitean cambios.')
        db.session.rollback()
    else:
        db.session.commit()
        log('[OK] Commit realizado.')

    return {
        'creados': creados,
        'actualizados': actualizados,
        'skipped': skipped,
        'zonas_creadas': zonas_creadas,
        'errores': errores,
    }


def main():
    args = parse_args()

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print('ERROR: instala openpyxl: pip install openpyxl', file=sys.stderr)
        sys.exit(2)

    sys.path.insert(0, os.getcwd())
    from app import app  # noqa: E402

    archivo = args.archivo
    if not os.path.isabs(archivo):
        archivo = os.path.join(os.getcwd(), archivo)

    if not os.path.exists(archivo):
        print(f'ERROR: no existe el archivo {archivo}', file=sys.stderr)
        sys.exit(1)

    with app.app_context():
        try:
            resumen = importar_directorio_global(
                archivo, hoja=args.hoja,
                dry_run=args.dry_run, verbose=args.verbose,
            )
        except Exception as e:
            print(f'ERROR: {e}', file=sys.stderr)
            sys.exit(1)

    print()
    print('=' * 60)
    print(' RESUMEN IMPORT DIRECTORIO GLOBAL')
    print('=' * 60)
    print(f"  Proveedores creados      : {resumen['creados']}")
    print(f"  Proveedores actualizados : {resumen['actualizados']}")
    print(f"  Filas omitidas (vacias)  : {resumen['skipped']}")
    print(f"  Zonas nuevas creadas     : {resumen['zonas_creadas']}")
    print(f"  Errores                  : {len(resumen['errores'])}")
    if resumen['errores']:
        print('  Detalle de errores:')
        for fila, msg in resumen['errores'][:10]:
            print(f'    fila {fila}: {msg}')
        if len(resumen['errores']) > 10:
            print(f"    ... ({len(resumen['errores']) - 10} mas)")


if __name__ == '__main__':
    main()
