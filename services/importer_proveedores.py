"""Importer de proveedores desde Excel del usuario.

Permite que cualquier admin/PM cargue su propia base de proveedores via Excel
con columnas fijas. Si el usuario marca 'Compartir con Directorio OBYRA',
ademas de crear el ProveedorOC scope='tenant' (privado de su org), se crea
un duplicado con scope='global' + estado_compartido='pendiente' que el super
admin debe aprobar.

Columnas esperadas (header en fila 1, case-insensitive):
    razon_social   *obligatorio
    cuit
    categoria
    subcategoria
    zona
    contacto       (nombre del contacto)
    email
    telefono
    whatsapp
    notas

Reglas de unicidad para UPSERT:
  - Privado (tenant): (organizacion_id, cuit) si hay CUIT,
                      sino (organizacion_id, descripcion_normalizada).
  - Global compartido: (cuit) si hay CUIT,
                       sino (razon_social_normalizada).
"""

from __future__ import annotations

import os
import re
import unicodedata
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple


CABECERA_TEMPLATE = [
    'razon_social', 'cuit', 'categoria', 'subcategoria', 'zona',
    'contacto', 'email', 'telefono', 'whatsapp', 'notas',
]


def _normalizar(texto: Any) -> str:
    """lowercase + sin tildes + trim."""
    if texto is None:
        return ''
    s = str(texto).strip().lower()
    s = ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )
    return ' '.join(s.split())


def _normalizar_cuit(cuit: Any) -> Optional[str]:
    """Quita guiones y espacios. Devuelve None si no parece CUIT (no tiene digitos)."""
    if cuit is None:
        return None
    s = re.sub(r'[^0-9]', '', str(cuit))
    return s if len(s) >= 8 else None


def _normalizar_razon(razon: Any) -> str:
    """Razon social normalizada para dedup. Quita 'S.A.', 'S.R.L.', puntos, comas."""
    base = _normalizar(razon)
    base = re.sub(r'\b(s\.?\s?a\.?|s\.?\s?r\.?\s?l\.?|s\.?a\.?s\.?|sa|srl|sas)\b', '', base)
    base = re.sub(r'[.,]', '', base)
    return ' '.join(base.split())


def _g(row: Dict[str, Any], key: str, default: Any = None) -> Any:
    """Lee una columna del row de Excel, tolera variantes de mayusculas/espacios."""
    if not row:
        return default
    if key in row:
        v = row[key]
        return v if v not in (None, '') else default
    # buscar case-insensitive
    klow = key.lower().strip()
    for k, v in row.items():
        if str(k or '').lower().strip() == klow:
            return v if v not in (None, '') else default
    return default


def _parse_xlsx(xlsx_path: str) -> List[Dict[str, Any]]:
    """Parsea el Excel y devuelve lista de dicts con keys estandar."""
    import openpyxl

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f'No existe el archivo: {xlsx_path}')

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # Soportar hoja 'Proveedores' o la primera disponible.
    if 'Proveedores' in wb.sheetnames:
        ws = wb['Proveedores']
    else:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header_raw = [(str(c).strip().lower() if c else '') for c in rows[0]]
    col_idx = {h: i for i, h in enumerate(header_raw)}

    def _read(row, key, default=None):
        i = col_idx.get(key.lower())
        if i is None or i >= len(row):
            return default
        v = row[i]
        return v if v not in (None, '') else default

    salida = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        razon = _read(row, 'razon_social')
        if not razon:
            continue
        salida.append({
            'razon_social': str(razon).strip()[:200],
            'cuit': _normalizar_cuit(_read(row, 'cuit')),
            'categoria': (str(_read(row, 'categoria') or '').strip() or None) and str(_read(row, 'categoria')).strip()[:120],
            'subcategoria': (str(_read(row, 'subcategoria') or '').strip() or None) and str(_read(row, 'subcategoria')).strip()[:160],
            'zona': (str(_read(row, 'zona') or '').strip() or None) and str(_read(row, 'zona')).strip()[:120],
            'contacto': (str(_read(row, 'contacto') or '').strip() or None) and str(_read(row, 'contacto')).strip()[:200],
            'email': (str(_read(row, 'email') or '').strip() or None) and str(_read(row, 'email')).strip()[:200],
            'telefono': (str(_read(row, 'telefono') or '').strip() or None) and str(_read(row, 'telefono')).strip()[:50],
            'whatsapp': (str(_read(row, 'whatsapp') or '').strip() or None) and str(_read(row, 'whatsapp')).strip()[:50],
            'notas': (str(_read(row, 'notas') or '').strip() or None) and str(_read(row, 'notas')).strip(),
        })
    return salida


def importar_excel_proveedores(
    *,
    db,
    xlsx_path: str,
    organizacion_id: int,
    user_id: Optional[int] = None,
    compartir_con_directorio: bool = False,
) -> Dict[str, Any]:
    """Procesa el Excel y carga proveedores en proveedores_oc.

    Args:
      db: instancia SQLAlchemy.
      xlsx_path: path al archivo subido.
      organizacion_id: organizacion del usuario que sube.
      user_id: usuario que dispara el import (para auditoria).
      compartir_con_directorio: si True, ademas del privado se crea uno global
        con estado_compartido='pendiente' para revision del super admin.

    Returns:
      dict con contadores: filas_total, privados_creados, privados_actualizados,
      compartidos_pendientes, invalidos, errores[]
    """
    from models.proveedores_oc import ProveedorOC

    filas = _parse_xlsx(xlsx_path)

    contadores = {
        'filas_total': len(filas),
        'privados_creados': 0,
        'privados_actualizados': 0,
        'compartidos_pendientes': 0,
        'compartidos_ya_existian': 0,
        'invalidos': 0,
        'errores': [],
    }

    if not filas:
        return contadores

    for fila in filas:
        razon = fila['razon_social']
        cuit = fila.get('cuit')
        razon_norm = _normalizar_razon(razon)

        try:
            # 1) UPSERT en privado (scope='tenant', organizacion_id=org actual)
            existente = None
            if cuit:
                existente = ProveedorOC.query.filter_by(
                    organizacion_id=organizacion_id,
                    scope='tenant',
                    cuit=cuit,
                ).first()
            if not existente:
                # buscar por razon_social normalizada dentro de la org
                candidatos = ProveedorOC.query.filter_by(
                    organizacion_id=organizacion_id,
                    scope='tenant',
                ).all()
                for c in candidatos:
                    if _normalizar_razon(c.razon_social) == razon_norm:
                        existente = c
                        break

            if existente:
                _aplicar_campos(existente, fila, user_id)
                contadores['privados_actualizados'] += 1
                proveedor_privado = existente
            else:
                proveedor_privado = ProveedorOC(
                    organizacion_id=organizacion_id,
                    scope='tenant',
                    razon_social=razon[:200],
                    cuit=cuit,
                    created_by_id=user_id,
                )
                _aplicar_campos(proveedor_privado, fila, user_id)
                db.session.add(proveedor_privado)
                contadores['privados_creados'] += 1

            # 2) Si el usuario marco compartir, crear duplicado global pendiente
            if compartir_con_directorio:
                # Verificar si ya hay un global con mismo CUIT o razon_norm.
                # Si hay un APROBADO, no hace falta volver a crear pendiente.
                # Si hay un PENDIENTE de la misma org, tampoco duplicamos.
                global_existente = None
                if cuit:
                    global_existente = ProveedorOC.query.filter_by(
                        scope='global', cuit=cuit,
                    ).filter(
                        ProveedorOC.estado_compartido.in_(('pendiente', 'aprobado'))
                    ).first()
                if not global_existente:
                    candidatos_g = ProveedorOC.query.filter_by(
                        scope='global',
                    ).filter(
                        ProveedorOC.estado_compartido.in_(('pendiente', 'aprobado'))
                    ).all()
                    for c in candidatos_g:
                        if _normalizar_razon(c.razon_social) == razon_norm:
                            global_existente = c
                            break

                if global_existente:
                    contadores['compartidos_ya_existian'] += 1
                else:
                    proveedor_global = ProveedorOC(
                        organizacion_id=None,  # global
                        scope='global',
                        razon_social=razon[:200],
                        cuit=cuit,
                        created_by_id=user_id,
                        compartido_por_org_id=organizacion_id,
                        estado_compartido='pendiente',
                    )
                    _aplicar_campos(proveedor_global, fila, user_id)
                    db.session.add(proveedor_global)
                    contadores['compartidos_pendientes'] += 1
        except Exception as e:
            contadores['invalidos'] += 1
            contadores['errores'].append({
                'razon_social': razon,
                'error': f'{type(e).__name__}: {str(e)[:200]}',
            })

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        contadores['errores'].append({
            'razon_social': '(commit)',
            'error': f'Error commiteando: {type(e).__name__}: {str(e)[:200]}',
        })

    return contadores


def _aplicar_campos(proveedor, fila: Dict[str, Any], user_id: Optional[int]):
    """Aplica los campos del Excel al ProveedorOC. Solo sobrescribe si el Excel
    trae el dato (no pisa con None)."""
    if fila.get('cuit'):
        proveedor.cuit = fila['cuit']
    if fila.get('categoria'):
        proveedor.categoria = fila['categoria']
    if fila.get('subcategoria'):
        proveedor.subcategoria = fila['subcategoria']
    if fila.get('zona'):
        # zona del Excel es texto libre, lo guardamos en ubicacion_detalle
        # (zona_id requiere FK al modelo Zona que es independiente).
        proveedor.ubicacion_detalle = fila['zona']
    if fila.get('contacto'):
        proveedor.contacto_nombre = fila['contacto']
    if fila.get('email'):
        proveedor.email = fila['email']
    if fila.get('telefono'):
        proveedor.telefono = fila['telefono']
        if not proveedor.contacto_telefono:
            proveedor.contacto_telefono = fila['telefono']
    if fila.get('whatsapp'):
        proveedor.contacto_whatsapp = fila['whatsapp']
    if fila.get('notas'):
        proveedor.notas = fila['notas']
    proveedor.activo = True
    proveedor.updated_at = datetime.utcnow()


def generar_template_xlsx(destino_path: str) -> str:
    """Genera el Excel template descargable con columnas fijas y 1 fila ejemplo.
    Returns: path absoluto al archivo creado.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proveedores'

    # Header con formato
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2D3F66', end_color='2D3F66', fill_type='solid')
    for col_idx, col_name in enumerate(CABECERA_TEMPLATE, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # Fila ejemplo
    ejemplo = {
        'razon_social': 'Cementos Avellaneda S.A.',
        'cuit': '30-50001234-5',
        'categoria': 'Cemento',
        'subcategoria': 'Cemento gris portland',
        'zona': 'CABA / GBA',
        'contacto': 'Juan Pérez',
        'email': 'ventas@cementosavellaneda.com.ar',
        'telefono': '+54 11 4000 0000',
        'whatsapp': '+54 9 11 4000 0000',
        'notas': 'Pedidos mínimos: 100 bolsas. Plazo entrega 48hs CABA.',
    }
    for col_idx, col_name in enumerate(CABECERA_TEMPLATE, start=1):
        ws.cell(row=2, column=col_idx, value=ejemplo.get(col_name, ''))

    # Ancho columnas
    anchos = {
        'razon_social': 30, 'cuit': 16, 'categoria': 18, 'subcategoria': 24,
        'zona': 18, 'contacto': 22, 'email': 30, 'telefono': 18,
        'whatsapp': 18, 'notas': 40,
    }
    for col_idx, col_name in enumerate(CABECERA_TEMPLATE, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = anchos.get(col_name, 18)

    wb.save(destino_path)
    return destino_path
