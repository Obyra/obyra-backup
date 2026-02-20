"""
Excel Budget Parser for Construction Pliegos
=============================================
Parses .xlsx files from constructor budgets into structured data for OBYRA import.

Supports:
- Single-vendor files (one set of price columns)
- Multi-vendor comparison files (repeated column sets per vendor)
- 3 planillas: Organización, Hormigón Armado, Albañilería
"""

import re
import unicodedata
from decimal import Decimal, InvalidOperation
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Optional, Tuple
from io import BytesIO

import openpyxl


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class ParsedItem:
    codigo: str
    descripcion: str
    unidad: str
    cantidad: float
    precio_unitario: float
    total: float
    rubro_nombre: str          # Maps to etapa_nombre
    planilla: str              # organizacion / hormigon_armado / albanileria
    nivel: int                 # 1=Rubro header, 2=Item, 3=Sub-item
    es_subtotal: bool = False
    valor_especial: Optional[str] = None
    fila_excel: int = 0

    def to_dict(self):
        return asdict(self)


@dataclass
class DetectedVendor:
    nombre: str
    col_cantidad: int
    col_precio_unit: int
    col_parcial: int
    col_subtotal: Optional[int] = None

    def to_dict(self):
        return {'nombre': self.nombre, 'col_cantidad': self.col_cantidad}


@dataclass
class ParseResult:
    vendors: List[DetectedVendor]
    items_by_vendor: Dict[str, List[ParsedItem]]
    sheet_names: List[str]
    file_type: str             # single_vendor / multi_vendor
    total_items: int = 0
    skipped_items: int = 0
    warnings: List[str] = field(default_factory=list)

    def to_dict(self):
        result = {
            'file_type': self.file_type,
            'sheet_names': self.sheet_names,
            'total_items': self.total_items,
            'skipped_items': self.skipped_items,
            'warnings': self.warnings,
            'vendors': [v.to_dict() for v in self.vendors],
            'items_by_vendor': {},
        }
        for vname, items in self.items_by_vendor.items():
            etapas_dict = {}
            for item in items:
                etapa = item.rubro_nombre
                if etapa not in etapas_dict:
                    etapas_dict[etapa] = {'nombre': etapa, 'items': [], 'subtotal': 0}
                etapas_dict[etapa]['items'].append(item.to_dict())
                etapas_dict[etapa]['subtotal'] += item.total
            result['items_by_vendor'][vname] = {
                'etapas': list(etapas_dict.values()),
                'total_items': len(items),
                'total_monto': sum(i.total for i in items),
                'skipped': sum(1 for i in items if i.valor_especial),
            }
        return result


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SPECIAL_VALUES_SKIP = {
    'incluido', 'incluído', 'no aplica', 'no se cotiza',
    'no cotiza', 'n/a', 'nc', 'sin cotizar', 'excluido',
    'se elimina', 'anulado',
}

SPECIAL_VALUES_ZERO = {
    'a def', 'a definir', 'pendiente', 'preguntar',
}

REFERENCE_RE = re.compile(
    r'(?:incluido en|incluído en|inc\.?\s*en)\s*(?:item|ítem)?\s*([\d,\.]+)',
    re.IGNORECASE,
)

UNIT_MAP = {
    'gl': 'gl', 'global': 'gl', 'gbl': 'gl', 'pa': 'gl',
    'm2': 'm2', 'm²': 'm2',
    'm3': 'm3', 'm³': 'm3',
    'ml': 'ml', 'm.l.': 'ml', 'mts': 'ml', 'm lin': 'ml',
    'un': 'unidades', 'u': 'unidades', 'unid': 'unidades', 'unidad': 'unidades',
    'mes': 'mes', 'meses': 'mes',
    'kg': 'kg', 'kgs': 'kg',
    'tn': 'tn', 'ton': 'tn',
    'hrs': 'hrs', 'hs': 'hrs', 'hora': 'hrs', 'horas': 'hrs',
    'dia': 'dia', 'dias': 'dia',
}

# Mapeo de rubros del pliego a etapas OBYRA
RUBRO_ETAPA_MAP = {
    # Organización
    'organizacion': 'Preliminares y Obrador',
    'organizacion de obra': 'Preliminares y Obrador',
    'tareas preliminares': 'Preliminares y Obrador',
    'items complementarios': 'Instalaciones Complementarias',
    # Hormigón Armado
    'depresion de napa': 'Depresión de Napa / Bombeo',
    'demoliciones': 'Demoliciones',
    'movimiento de suelos y rellenos': 'Movimiento de Suelos',
    'movimiento de suelos': 'Movimiento de Suelos',
    'apuntalamientos': 'Apuntalamientos',
    'estructuras de h.a.': 'Estructura',
    'estructuras de ha': 'Estructura',
    'estructuras de h.a. fundaciones': 'Estructura',
    'estructura': 'Estructura',
    'fundaciones': 'Fundaciones',
    # Albañilería
    'mamposteria': 'Mampostería',
    'construccion en seco': 'Construcción en Seco',
    'aislaciones / impermeabilizaciones': 'Impermeabilizaciones y Aislaciones',
    'aislaciones': 'Impermeabilizaciones y Aislaciones',
    'impermeabilizaciones': 'Impermeabilizaciones y Aislaciones',
    'conductos de ventilacion': 'Ventilaciones y Conductos',
    'revoques interiores': 'Revoque Grueso',
    'revoques exteriores': 'Revoque Fino',
    'cielorrasos': 'Cielorrasos',
    'contrapisos y carpetas': 'Contrapisos y Carpetas',
    'contrapisos': 'Contrapisos y Carpetas',
    'pisos': 'Pisos',
    'zocalos': 'Pisos',
    'revestimientos': 'Revestimientos',
    'yeseria': 'Yesería y Enlucidos',
    'provisiones y colocaciones': 'Provisiones y Colocaciones',
    'pintura': 'Pintura',
    'carpinteria': 'Carpintería',
    'instalaciones electricas': 'Instalaciones Eléctricas',
    'instalaciones sanitarias': 'Instalaciones Sanitarias',
    'instalaciones de gas': 'Instalaciones de Gas',
    'limpieza final': 'Limpieza Final',
    'limpieza': 'Limpieza Final',
}

# Planilla-name detection keywords
PLANILLA_KEYWORDS = {
    'organizacion': 'organizacion',
    'orga': 'organizacion',
    'tareas prelimina': 'organizacion',
    'presentacion tp': 'organizacion',
    'h.a': 'hormigon_armado',
    'ha': 'hormigon_armado',
    'hormigon': 'hormigon_armado',
    'presentacion ho': 'hormigon_armado',
    'albanileria': 'albanileria',
    'albañileria': 'albanileria',
    'alb': 'albanileria',
    'presup-alb': 'albanileria',
    'total': 'resumen',
    'resumen': 'resumen',
    'propuesta': 'completo',
    'pre ': 'completo',
}

VENDOR_NAMES_KNOWN = [
    'cortes y sistemas', 'cys', 'c y sistemas',
    'overcon',
    'tosud',
    'azzollini',
    'del tejar',
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _slugify(text: str) -> str:
    if not text:
        return ''
    nfkd = unicodedata.normalize('NFKD', text)
    ascii_text = nfkd.encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'[^a-z0-9]+', ' ', ascii_text.lower()).strip()


def _normalize_unit(raw: str) -> str:
    if not raw:
        return 'gl'
    key = raw.strip().lower().replace('.', '').replace(' ', '')
    return UNIT_MAP.get(key, UNIT_MAP.get(raw.strip().lower(), raw.strip().lower()))


def _safe_decimal(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned or cleaned == '-':
            return None
        # Check special values first
        lower = cleaned.lower()
        for sv in SPECIAL_VALUES_SKIP | SPECIAL_VALUES_ZERO:
            if sv in lower:
                return None
        # Remove currency symbols and thousands separators
        cleaned = cleaned.replace('$', '').replace(' ', '').strip()
        # Handle Argentine format: 1.234.567,89
        if ',' in cleaned and '.' in cleaned:
            if cleaned.rindex(',') > cleaned.rindex('.'):
                cleaned = cleaned.replace('.', '').replace(',', '.')
            else:
                cleaned = cleaned.replace(',', '')
        elif ',' in cleaned and '.' not in cleaned:
            cleaned = cleaned.replace(',', '.')
        try:
            return float(cleaned)
        except (ValueError, InvalidOperation):
            return None
    return None


def _is_special(value) -> Tuple[bool, Optional[str], bool]:
    """Returns (is_special, special_text, should_skip).
    should_skip=True means don't create item. False means create with $0."""
    if value is None:
        return False, None, False
    text = str(value).strip().lower()
    if not text or text == '-':
        return False, None, False
    # Check skip values
    for sv in SPECIAL_VALUES_SKIP:
        if sv in text:
            return True, sv.upper(), True
    # Check reference patterns
    if REFERENCE_RE.search(str(value)):
        return True, 'INCLUIDO', True
    # Check zero-price values
    for sv in SPECIAL_VALUES_ZERO:
        if sv in text:
            return True, sv.upper(), False
    return False, None, False


def _detect_item_level(code_value) -> int:
    """1=Rubro, 2=Item (1.01), 3=Sub-item (5,01,01)."""
    if code_value is None:
        return 0
    s = str(code_value).strip()
    if not s:
        return 0
    # Sub-item: contains commas like 5,01,01
    if ',' in s and re.match(r'\d+,\d+,\d+', s):
        return 3
    # Item: decimal like 1.01
    if '.' in s and re.match(r'\d+\.\d+', s):
        return 2
    # Rubro: integer
    try:
        int(float(s))
        if float(s) == int(float(s)):
            return 1
    except (ValueError, TypeError):
        pass
    return 0


def _map_rubro_to_etapa(rubro_name: str, planilla: str) -> str:
    if not rubro_name:
        return planilla.replace('_', ' ').title()
    slug = _slugify(rubro_name)
    # Direct lookup
    for key, etapa in RUBRO_ETAPA_MAP.items():
        if key in slug or slug in key:
            return etapa
    # For HA sub-rubros like "Losa s/PB", "Losa s/5to subsuelo" -> Estructura
    if any(kw in slug for kw in ['losa', 'columna', 'viga', 'tabique', 'escalera', 'rampa']):
        return 'Estructura'
    # Return cleaned original name
    return rubro_name.strip()


def _classify_planilla(sheet_name: str) -> str:
    lower = sheet_name.lower().strip()
    for keyword, planilla in PLANILLA_KEYWORDS.items():
        if keyword in lower:
            return planilla
    return 'desconocido'


def _classify_tipo(descripcion: str) -> str:
    lower = descripcion.lower() if descripcion else ''
    mo_keywords = ['mano de obra', 'oficial', 'ayudante', 'jornal', 'capataz',
                   'supervision', 'supervisión', 'personal', 'sereno', 'vigilancia']
    equipo_keywords = ['alquiler', 'equipo', 'grua', 'grúa', 'volquete', 'guinche',
                       'montacargas', 'generador', 'electrogeno', 'electrógeno',
                       'bomba', 'herramienta']
    for kw in mo_keywords:
        if kw in lower:
            return 'mano_obra'
    for kw in equipo_keywords:
        if kw in lower:
            return 'equipo'
    return 'material'


# ---------------------------------------------------------------------------
# Header / vendor detection
# ---------------------------------------------------------------------------

def _find_header_row(ws, max_row=15) -> Optional[int]:
    """Find the row containing column headers (Rubros, Designacion, Unidad, etc.)."""
    for row_idx in range(1, min(max_row + 1, ws.max_row + 1)):
        for col_idx in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and isinstance(val, str):
                lower = val.strip().lower()
                if any(kw in lower for kw in ['designacion', 'descripcion', 'trabajos']):
                    return row_idx
    return None


def _find_columns(ws, header_row: int) -> Dict[str, int]:
    """Identify key column positions from the header row."""
    cols = {}
    for col_idx in range(1, min(40, ws.max_column + 1)):
        val = ws.cell(row=header_row, column=col_idx).value
        if not val or not isinstance(val, str):
            continue
        lower = val.strip().lower()
        if any(kw in lower for kw in ['rubros', 'rubro', 'item']) and 'code' not in cols:
            cols['code'] = col_idx
        elif any(kw in lower for kw in ['designacion', 'descripcion']) and 'desc' not in cols:
            cols['desc'] = col_idx
        elif lower in ('unidad', 'un', 'un.') and 'unit' not in cols:
            cols['unit'] = col_idx
    return cols


def _detect_vendors(ws, max_row=12) -> Tuple[str, List[DetectedVendor]]:
    """Detect vendor column groups by scanning header area."""
    vendors = []
    # Look for repeated patterns of Cant/Cantidad + Precio columns
    price_cols = []  # (col_idx, type) where type is 'cant', 'pu', 'parcial', 'subtotal'

    for row_idx in range(1, min(max_row + 1, ws.max_row + 1)):
        for col_idx in range(1, min(40, ws.max_column + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if not val or not isinstance(val, str):
                continue
            lower = val.strip().lower()
            if lower in ('cant', 'cant.', 'cantidad'):
                price_cols.append((col_idx, 'cant', row_idx))
            elif 'unitario' in lower or ('precio' in lower and 'parcial' not in lower and 'total' not in lower):
                price_cols.append((col_idx, 'pu', row_idx))
            elif 'parcial' in lower:
                price_cols.append((col_idx, 'parcial', row_idx))
            elif 'subtotal' in lower:
                price_cols.append((col_idx, 'subtotal', row_idx))

    # Group by proximity (columns within 4 of each other belong to same vendor)
    cant_cols = sorted([p for p in price_cols if p[1] == 'cant'], key=lambda x: x[0])

    if len(cant_cols) <= 1:
        # Single vendor - find its columns
        vendor = _build_single_vendor(price_cols, ws, max_row)
        if vendor:
            return 'single_vendor', [vendor]
        # Fallback: assume standard layout E=cant, F=pu, G=parcial, H=subtotal
        return 'single_vendor', [DetectedVendor(
            nombre='Constructora',
            col_cantidad=5, col_precio_unit=6, col_parcial=7, col_subtotal=8
        )]

    # Multiple vendors
    for i, (cant_col, _, _) in enumerate(cant_cols):
        # Find closest PU and Parcial columns after this Cant
        pu_col = None
        parcial_col = None
        subtotal_col = None
        for p_col, p_type, _ in price_cols:
            if p_col > cant_col and p_col <= cant_col + 6:
                if p_type == 'pu' and pu_col is None:
                    pu_col = p_col
                elif p_type == 'parcial' and parcial_col is None:
                    parcial_col = p_col
                elif p_type == 'subtotal' and subtotal_col is None:
                    subtotal_col = p_col

        if pu_col is None:
            pu_col = cant_col + 1
        if parcial_col is None:
            parcial_col = cant_col + 2

        # Try to find vendor name above the Cant column
        vendor_name = _find_vendor_name(ws, cant_col, max_row) or f'Vendor {i + 1}'
        vendors.append(DetectedVendor(
            nombre=vendor_name,
            col_cantidad=cant_col,
            col_precio_unit=pu_col,
            col_parcial=parcial_col,
            col_subtotal=subtotal_col,
        ))

    file_type = 'multi_vendor' if len(vendors) > 1 else 'single_vendor'
    return file_type, vendors


def _build_single_vendor(price_cols, ws, max_row) -> Optional[DetectedVendor]:
    cant_cols = [p for p in price_cols if p[1] == 'cant']
    pu_cols = [p for p in price_cols if p[1] == 'pu']
    parcial_cols = [p for p in price_cols if p[1] == 'parcial']
    subtotal_cols = [p for p in price_cols if p[1] == 'subtotal']

    if cant_cols:
        cant = cant_cols[0][0]
    else:
        cant = 5  # Default col E

    pu = pu_cols[0][0] if pu_cols else cant + 1
    parcial = parcial_cols[0][0] if parcial_cols else cant + 2
    subtotal = subtotal_cols[0][0] if subtotal_cols else cant + 3

    name = _find_vendor_name(ws, cant, max_row) or 'Constructora'
    return DetectedVendor(nombre=name, col_cantidad=cant,
                          col_precio_unit=pu, col_parcial=parcial, col_subtotal=subtotal)


def _find_vendor_name(ws, col_idx, max_row=10) -> Optional[str]:
    """Search for a vendor name in rows above the header near a column."""
    for row_idx in range(1, min(max_row + 1, ws.max_row + 1)):
        for c in range(max(1, col_idx - 1), min(col_idx + 3, ws.max_column + 1)):
            val = ws.cell(row=row_idx, column=c).value
            if val and isinstance(val, str):
                lower = val.strip().lower()
                for known in VENDOR_NAMES_KNOWN:
                    if known in lower:
                        return val.strip()
                # Check for "EMPRESA COTIZANTE: XXX"
                if 'empresa' in lower and 'cotizante' in lower:
                    parts = val.split(':')
                    if len(parts) > 1:
                        return parts[1].strip()
    return None


# ---------------------------------------------------------------------------
# Sheet parsing
# ---------------------------------------------------------------------------

def _parse_sheet(ws, vendor: DetectedVendor, planilla: str) -> List[ParsedItem]:
    """Parse a single sheet for one vendor."""
    items = []
    header_row = _find_header_row(ws)
    if not header_row:
        return items

    base_cols = _find_columns(ws, header_row)
    code_col = base_cols.get('code', 2)
    desc_col = base_cols.get('desc', 3)
    unit_col = base_cols.get('unit', 4)

    current_rubro = None
    current_rubro_5_sub = None  # Track sub-rubros for HA rubro 5 (per floor)
    data_start = header_row + 1

    for row_idx in range(data_start, ws.max_row + 1):
        code_val = ws.cell(row=row_idx, column=code_col).value
        desc_val = ws.cell(row=row_idx, column=desc_col).value

        if not desc_val and not code_val:
            continue

        desc_str = str(desc_val).strip() if desc_val else ''

        # Skip subtotal/total rows
        if any(kw in desc_str.lower() for kw in ['subtotal', 'total del', 'total general',
                                                   'total planilla', 'total +', 'sin iva']):
            continue

        # Detect level
        level = _detect_item_level(code_val)

        if level == 1:
            # Rubro header
            current_rubro = desc_str
            current_rubro_5_sub = None
            continue

        if level == 2:
            code_str = str(code_val).strip()
            # Check if this is a sub-rubro of rubro 5 (e.g., 5.01 = Fundaciones, 5.02 = Losa s/5SS)
            if code_str.startswith('5.') and planilla == 'hormigon_armado':
                current_rubro_5_sub = desc_str
                # Don't skip - some sub-rubros ARE items with prices
                # Fall through to check if it has price data

        # Read vendor-specific columns
        raw_cant = ws.cell(row=row_idx, column=vendor.col_cantidad).value
        raw_pu = ws.cell(row=row_idx, column=vendor.col_precio_unit).value
        raw_parcial = ws.cell(row=row_idx, column=vendor.col_parcial).value

        # Check for special values in any price cell
        for raw_val in [raw_cant, raw_pu, raw_parcial]:
            is_sp, sp_text, should_skip = _is_special(raw_val)
            if is_sp and should_skip:
                break
        else:
            is_sp, sp_text, should_skip = False, None, False

        if should_skip:
            items.append(ParsedItem(
                codigo=str(code_val) if code_val else '',
                descripcion=desc_str,
                unidad=_normalize_unit(str(ws.cell(row=row_idx, column=unit_col).value or '')),
                cantidad=0, precio_unitario=0, total=0,
                rubro_nombre=_get_etapa_name(current_rubro, current_rubro_5_sub, planilla),
                planilla=planilla,
                nivel=level,
                valor_especial=sp_text,
                fila_excel=row_idx,
            ))
            continue

        # Parse numeric values
        cantidad = _safe_decimal(raw_cant)
        precio_unitario = _safe_decimal(raw_pu)
        total_parcial = _safe_decimal(raw_parcial)

        # Skip rows with no useful data
        if cantidad is None and precio_unitario is None and total_parcial is None:
            if level == 2 and planilla == 'hormigon_armado':
                # This is likely a sub-rubro header (e.g., "5.02 Losa s/5to subsuelo")
                continue
            if not desc_str:
                continue
            # Row with description but no prices - could be "a def"
            is_sp_desc, sp_text_desc, _ = _is_special(desc_str)
            if is_sp_desc:
                continue
            # Skip description-only rows
            continue

        # Calculate missing values
        cant_val = cantidad or 0
        pu_val = precio_unitario or 0
        total_val = total_parcial or 0

        if total_val == 0 and cant_val and pu_val:
            total_val = cant_val * pu_val
        if pu_val == 0 and cant_val and total_val:
            pu_val = total_val / cant_val if cant_val != 0 else 0

        # Check for zero-price special values
        is_sp_zero, sp_text_zero, _ = _is_special(raw_pu)
        if not is_sp_zero:
            is_sp_zero, sp_text_zero, _ = _is_special(
                ws.cell(row=row_idx, column=unit_col).value)

        items.append(ParsedItem(
            codigo=str(code_val) if code_val else '',
            descripcion=desc_str,
            unidad=_normalize_unit(str(ws.cell(row=row_idx, column=unit_col).value or '')),
            cantidad=cant_val,
            precio_unitario=pu_val,
            total=total_val,
            rubro_nombre=_get_etapa_name(current_rubro, current_rubro_5_sub, planilla),
            planilla=planilla,
            nivel=level,
            valor_especial=sp_text_zero if is_sp_zero else None,
            fila_excel=row_idx,
        ))

    return items


def _get_etapa_name(rubro: Optional[str], sub_rubro_5: Optional[str], planilla: str) -> str:
    """Get the etapa name for an item, considering HA rubro 5 sub-rubros."""
    if rubro and _slugify(rubro).startswith('estructura'):
        # Items under Estructura rubro 5 - map to Estructura
        return _map_rubro_to_etapa(rubro, planilla)
    if sub_rubro_5:
        # Sub-rubro of 5 (per-floor items) - all map to Estructura
        return 'Estructura'
    if rubro:
        return _map_rubro_to_etapa(rubro, planilla)
    return planilla.replace('_', ' ').title()


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_excel_file(file_stream, filename: str) -> ParseResult:
    """Parse an Excel budget file and return structured data."""
    warnings = []

    try:
        wb = openpyxl.load_workbook(file_stream, read_only=False, data_only=True)
    except Exception as e:
        return ParseResult(
            vendors=[], items_by_vendor={}, sheet_names=[],
            file_type='error', warnings=[f'Error al abrir el archivo: {str(e)}']
        )

    sheet_names = wb.sheetnames

    # Detect vendors from first non-summary sheet
    file_type = 'single_vendor'
    all_vendors = []

    for sname in sheet_names:
        planilla = _classify_planilla(sname)
        if planilla in ('resumen', 'desconocido'):
            continue
        ws = wb[sname]
        ft, vendors = _detect_vendors(ws)
        if ft == 'multi_vendor':
            file_type = 'multi_vendor'
        for v in vendors:
            if not any(av.nombre == v.nombre for av in all_vendors):
                all_vendors.append(v)
        break  # Use first data sheet for vendor detection

    if not all_vendors:
        # Fallback
        all_vendors = [DetectedVendor(
            nombre='Constructora', col_cantidad=5,
            col_precio_unit=6, col_parcial=7, col_subtotal=8
        )]

    # Parse each sheet for each vendor
    items_by_vendor: Dict[str, List[ParsedItem]] = {v.nombre: [] for v in all_vendors}
    total_items = 0
    skipped = 0

    for sname in sheet_names:
        planilla = _classify_planilla(sname)
        if planilla in ('resumen', 'desconocido'):
            continue

        ws = wb[sname]

        # For multi-vendor files, re-detect vendors per sheet
        if file_type == 'multi_vendor':
            _, sheet_vendors = _detect_vendors(ws)
        else:
            sheet_vendors = all_vendors

        for vendor in sheet_vendors:
            # Make sure vendor is in our results
            if vendor.nombre not in items_by_vendor:
                items_by_vendor[vendor.nombre] = []
                if not any(av.nombre == vendor.nombre for av in all_vendors):
                    all_vendors.append(vendor)

            parsed = _parse_sheet(ws, vendor, planilla)

            for item in parsed:
                if item.valor_especial and item.total == 0 and item.cantidad == 0:
                    skipped += 1
                else:
                    total_items += 1

            items_by_vendor[vendor.nombre].extend(parsed)

    wb.close()

    # Filter out items with skip-type special values (keep zero-price ones)
    for vname in items_by_vendor:
        filtered = []
        for item in items_by_vendor[vname]:
            if item.valor_especial and item.valor_especial.upper() in (
                'INCLUIDO', 'INCLUÍDO', 'NO APLICA', 'NO SE COTIZA',
                'NO COTIZA', 'N/A', 'NC', 'EXCLUIDO', 'SE ELIMINA', 'ANULADO'
            ):
                continue
            filtered.append(item)
        items_by_vendor[vname] = filtered

    return ParseResult(
        vendors=all_vendors,
        items_by_vendor=items_by_vendor,
        sheet_names=sheet_names,
        file_type=file_type,
        total_items=total_items,
        skipped_items=skipped,
        warnings=warnings,
    )
