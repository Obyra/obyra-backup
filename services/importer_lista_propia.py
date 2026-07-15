"""Importer mínimo de lista propia OBYRA para demo viernes.

Soporta 2 archivos Excel curados por OBYRA:
1. OBYRA_base_precios_recursos_v1.xlsx → hoja '02_Materiales_CABA_GBA'
   (~25 filas, materiales típicos AMBA con precio promedio + zona).
2. Costo mano obra-abr2026.xlsx → hoja 'Import_OBYRA'
   (~16 filas, MO ya formateada con tipo_recurso, recurso_normalizado,
    unidad, precio, moneda, zona).

Ambos van a `provider_price_list` con `proveedor_id=NULL` (lista propia
OBYRA), `fuente='lista_propia'`, `modalidad='compra'`. Se traza con
`import_batch` para poder deshacer.

NO se importan: hoja 03 (Abelson), 04 (SINIS), 05 (Leiten), 01 (Catalogo
derivado). Esas son post-demo.
"""
from __future__ import annotations

import hashlib
import os
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple


def _normalizar_unidad_str(s: str) -> str:
    """Normaliza string de unidad: lowercase, sin acentos, sin espacios extra."""
    if not s:
        return ''
    import unicodedata
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _inferir_unidad_de_descripcion(descripcion: str) -> str:
    """Cuando la columna `unidad` está vacía, infiere desde la descripción.

    Heurística por patrones comunes en Excel de OBYRA.
    """
    if not descripcion:
        return 'un'
    d = str(descripcion).lower()

    # Patrones explícitos de paréntesis: "(m³)", "(m3)", "(50 kg)"
    if re.search(r'\(.*?m[³3].*?\)', d) or re.search(r'\bm[³3]\b', d):
        return 'm3'
    if re.search(r'\(.*?m[²2].*?\)', d) or re.search(r'\bm[²2]\b', d):
        return 'm2'
    if re.search(r'\bml\b|metro\s*lineal', d):
        return 'ml'
    if re.search(r'\bkg\b|kilo\b|por kg', d):
        return 'kg'
    if re.search(r'bolsa', d):
        return 'bolsa'
    if re.search(r'\bjornal\b', d):
        return 'jornal'
    if re.search(r'\bhora\b', d):
        return 'hora'
    if re.search(r'\blitro\b|\bl\b', d):
        return 'l'
    if re.search(r'\b(unidad|c/u|cada uno|ud)\b|\bun\b', d):
        return 'un'
    if re.search(r'\btn\b|\btonelada\b', d):
        return 'tn'
    if re.search(r'\bgl\b|\bglobal\b', d):
        return 'gl'
    return 'un'


def _decimal_safe(v) -> Optional[Decimal]:
    """Convierte a Decimal de forma segura. None si falla o es <= 0."""
    if v is None:
        return None
    try:
        if isinstance(v, str):
            # Limpiar formato AR: "$162.000" → 162000, "162.000" → 162000
            s = v.replace('$', '').replace(' ', '').strip()
            # Si tiene coma decimal: "1.234,56" → "1234.56"
            if ',' in s and s.count(',') == 1:
                s = s.replace('.', '').replace(',', '.')
            elif s.count('.') > 1:
                s = s.replace('.', '')
            v = s
        d = Decimal(str(v))
        return d if d > 0 else None
    except (InvalidOperation, ValueError, TypeError):
        return None


def _parse_hoja_02_materiales(xlsx_path: str) -> List[Dict[str, Any]]:
    """Parsea hoja '02_Materiales_CABA_GBA' del OBYRA_base_v1.

    Estructura esperada (header en fila 1):
      fuente | proveedor | categoria | tipo_recurso | recurso | descripcion |
      unidad | zona | precio_min | precio_max | precio_unitario | moneda | ...

    Filas con `recurso` o `descripcion` vacíos se saltan. Si `precio_unitario`
    está vacío, calcula promedio entre precio_min y precio_max.
    """
    import openpyxl
    if not os.path.exists(xlsx_path):
        return []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if '02_Materiales_CABA_GBA' not in wb.sheetnames:
        return []
    ws = wb['02_Materiales_CABA_GBA']
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    # Header → mapa columna
    header = [(_normalizar_unidad_str(str(c)) if c else '') for c in rows[0]]
    col_idx = {h: i for i, h in enumerate(header)}

    def _g(row, key, default=None):
        i = col_idx.get(key)
        if i is None or i >= len(row):
            return default
        return row[i]

    salida: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        recurso = _g(row, 'recurso')
        descripcion = _g(row, 'descripcion') or recurso
        if not recurso and not descripcion:
            continue
        unidad_raw = _g(row, 'unidad')
        unidad = _normalizar_unidad_str(unidad_raw) if unidad_raw else ''
        if not unidad:
            unidad = _inferir_unidad_de_descripcion(str(recurso or descripcion or ''))
        zona = (_g(row, 'zona') or '').strip() if _g(row, 'zona') else None
        moneda = (_g(row, 'moneda') or 'ARS').strip().upper() if _g(row, 'moneda') else 'ARS'
        precio_unit = _decimal_safe(_g(row, 'precio_unitario'))
        if not precio_unit:
            pmin = _decimal_safe(_g(row, 'precio_min'))
            pmax = _decimal_safe(_g(row, 'precio_max'))
            if pmin and pmax:
                precio_unit = (pmin + pmax) / Decimal('2')
            elif pmin:
                precio_unit = pmin
            elif pmax:
                precio_unit = pmax
        if not precio_unit:
            continue
        proveedor = (_g(row, 'proveedor') or '').strip() if _g(row, 'proveedor') else None
        # Si trae proveedor lo ignoramos por ahora (lista propia = proveedor=NULL).
        # En futuro podría matchearse con ProveedorOC.

        salida.append({
            'descripcion': str(descripcion).strip()[:300],
            'recurso': str(recurso).strip()[:300] if recurso else None,
            'unidad': unidad[:20] or 'un',
            'zona': zona[:40] if zona else None,
            'moneda': moneda[:3] if moneda else 'ARS',
            'precio_unitario': precio_unit,
            'tipo_recurso': (_g(row, 'tipo_recurso') or 'material').strip().lower(),
            'categoria': (_g(row, 'categoria') or '').strip()[:120] if _g(row, 'categoria') else None,
            'fuente_excel': (_g(row, 'fuente') or 'OBYRA_base_v1.02_Materiales_CABA_GBA')[:80],
        })
    return salida


def _parse_hoja_import_obyra_mano_obra(xlsx_path: str) -> List[Dict[str, Any]]:
    """Parsea hoja 'Import_OBYRA' del Costo mano obra-abr2026.xlsx.

    Estructura esperada (header en fila 1):
      tipo_recurso | categoria | recurso_normalizado | recurso_original |
      unidad | precio_unitario | moneda | zona | ...
    """
    import openpyxl
    if not os.path.exists(xlsx_path):
        return []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if 'Import_OBYRA' not in wb.sheetnames:
        return []
    ws = wb['Import_OBYRA']
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header = [(_normalizar_unidad_str(str(c)) if c else '') for c in rows[0]]
    col_idx = {h: i for i, h in enumerate(header)}

    def _g(row, key, default=None):
        i = col_idx.get(key)
        if i is None or i >= len(row):
            return default
        return row[i]

    salida: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        recurso = _g(row, 'recurso_normalizado') or _g(row, 'recurso_original')
        if not recurso:
            continue
        precio = _decimal_safe(_g(row, 'precio_unitario'))
        if not precio:
            continue
        unidad_raw = _g(row, 'unidad') or 'hora'
        unidad = _normalizar_unidad_str(str(unidad_raw))
        zona_raw = _g(row, 'zona')
        zona = None
        if zona_raw:
            zr = str(zona_raw).strip()
            # 'CABA / Zona A' → tomar primer segmento o 'CABA'
            if 'caba' in zr.lower():
                zona = 'CABA'
            elif 'gba' in zr.lower():
                zona = 'GBA'
            else:
                zona = zr[:40]
        moneda = (_g(row, 'moneda') or 'ARS').strip().upper() if _g(row, 'moneda') else 'ARS'
        salida.append({
            'descripcion': str(recurso).strip()[:300],
            'recurso': str(recurso).strip()[:300],
            'unidad': unidad[:20] or 'hora',
            'zona': zona,
            'moneda': moneda[:3] if moneda else 'ARS',
            'precio_unitario': precio,
            'tipo_recurso': 'mano_obra',
            'categoria': (_g(row, 'categoria') or 'Construcción')[:120],
            'fuente_excel': 'costo_mano_obra_abr2026.Import_OBYRA',
        })
    return salida


def _calcular_checksum(path: str) -> str:
    if not os.path.exists(path):
        return ''
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()


def importar_archivo(
    *,
    db,
    xlsx_path: str,
    organizacion_id: int,
    user_id: Optional[int] = None,
    perfil: str = 'lista_propia_obyra',
    sobrescribir: bool = True,
) -> Dict[str, Any]:
    """Orquestador principal. Importa un Excel completo a provider_price_list
    + precio_observado, con trazabilidad via import_batch.

    Args:
      sobrescribir: si True, hace UPSERT (actualiza precio si ya existe la fila
                    para misma org+proveedor+desc_norm+unidad+zona+modalidad).

    Returns:
      dict con resumen: filas leidas, insertadas, actualizadas, invalidas,
      batch_id, etc.
    """
    from models.import_batch import ImportBatch
    from models.provider_price_list import ProviderPriceList, normalizar_descripcion_precio
    from models.precio_observado import PrecioObservado

    if not os.path.exists(xlsx_path):
        return {'ok': False, 'error': f'archivo no encontrado: {xlsx_path}'}

    filename = os.path.basename(xlsx_path)
    checksum = _calcular_checksum(xlsx_path)

    # Crear batch
    batch = ImportBatch(
        organizacion_id=organizacion_id,
        perfil=perfil,
        filename=filename[:255],
        checksum_sha256=checksum,
        user_id=user_id,
        estado='en_curso',
        started_at=datetime.utcnow(),
    )
    db.session.add(batch)
    db.session.flush()  # tener batch.id

    # Parsear ambas hojas (si existen). Una sola fuente puede tener una u otra.
    filas_materiales = _parse_hoja_02_materiales(xlsx_path)
    filas_mo = _parse_hoja_import_obyra_mano_obra(xlsx_path)
    todas = filas_materiales + filas_mo

    inserted = 0
    updated = 0
    invalid = 0
    hoy = date.today()
    vigencia_default = date(hoy.year, hoy.month, hoy.day).replace(year=hoy.year + 1)

    for fila in todas:
        try:
            desc = fila['descripcion']
            desc_norm = normalizar_descripcion_precio(desc)
            unidad = fila['unidad']
            zona = fila.get('zona')
            moneda = fila['moneda'] or 'ARS'
            precio = fila['precio_unitario']
            tipo_recurso = fila.get('tipo_recurso') or 'material'

            if not desc_norm or not precio or precio <= 0:
                invalid += 1
                continue

            # Buscar existente con misma key (org, proveedor=NULL, desc_norm,
            # unidad, zona, modalidad='compra').
            existente = (ProviderPriceList.query
                         .filter(ProviderPriceList.organizacion_id == organizacion_id,
                                 ProviderPriceList.proveedor_id.is_(None),
                                 ProviderPriceList.descripcion_normalizada == desc_norm,
                                 ProviderPriceList.unidad == unidad)
                         .filter(
                            (ProviderPriceList.zona == zona) if zona is not None
                            else ProviderPriceList.zona.is_(None)
                         )
                         .first())

            if existente:
                if sobrescribir:
                    existente.precio_unitario = precio
                    existente.moneda = moneda
                    existente.fecha_actualizacion = hoy
                    existente.vigencia_hasta = vigencia_default
                    existente.fuente = perfil[:30]
                    existente.import_batch_id = batch.id
                    existente.modalidad = 'compra'
                    existente.notas = (
                        f'Lista propia OBYRA. Categoria: {fila.get("categoria") or "-"} · '
                        f'Origen Excel: {fila.get("fuente_excel") or "-"}'
                    )[:1000]
                    updated += 1
                # Si no sobrescribir, ignoramos.
                continue

            row = ProviderPriceList(
                organizacion_id=organizacion_id,
                proveedor_id=None,  # lista propia
                descripcion=desc[:300],
                descripcion_normalizada=desc_norm,
                unidad=unidad,
                precio_unitario=precio,
                moneda=moneda,
                fecha_actualizacion=hoy,
                vigencia_hasta=vigencia_default,
                fuente=perfil[:30],
                zona=zona,
                modalidad='compra',
                import_batch_id=batch.id,
                created_by_user_id=user_id,
                notas=(
                    f'Lista propia OBYRA. Categoria: {fila.get("categoria") or "-"} · '
                    f'Origen Excel: {fila.get("fuente_excel") or "-"}'
                )[:1000],
            )
            db.session.add(row)
            inserted += 1

            # Observación para memoria histórica (best-effort, no rompe import)
            try:
                obs = PrecioObservado(
                    organizacion_id=organizacion_id,
                    origen_tipo='lista_propia',
                    descripcion=desc[:300],
                    descripcion_normalizada=desc_norm,
                    unidad=unidad,
                    rubro_nombre=(fila.get('categoria') or '')[:100] or None,
                    tipo_recurso=tipo_recurso[:20] if tipo_recurso else 'material',
                    precio_unitario=precio,
                    moneda=moneda,
                    zona=zona,
                    modalidad='compra',
                    fecha_observado=hoy,
                    import_batch_id=batch.id,
                    notas=f'Importado desde {filename}',
                    valido=True,
                )
                db.session.add(obs)
            except Exception:
                pass
        except Exception as e:
            invalid += 1
            continue

    # Cerrar batch
    batch.total_input = len(todas)
    batch.total_inserted = inserted
    batch.total_updated = updated
    batch.total_invalid = invalid
    batch.estado = 'completado'
    batch.completed_at = datetime.utcnow()
    batch.metadata_json = {
        'hoja_02_filas': len(filas_materiales),
        'hoja_import_obyra_filas': len(filas_mo),
    }

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return {'ok': False, 'error': f'Error commiteando: {type(e).__name__}: {e}'}

    return {
        'ok': True,
        'batch_id': batch.id,
        'filename': filename,
        'total_input': len(todas),
        'inserted': inserted,
        'updated': updated,
        'invalid': invalid,
        'hoja_02_filas': len(filas_materiales),
        'hoja_import_obyra_filas': len(filas_mo),
    }


def deshacer_batch(*, db, batch_id: int, motivo: str = '',
                    organizacion_id: Optional[int] = None) -> Dict[str, Any]:
    """Borra todas las filas creadas/actualizadas por un import_batch."""
    from models.import_batch import ImportBatch
    from models.provider_price_list import ProviderPriceList
    from models.precio_observado import PrecioObservado

    batch = ImportBatch.query.get(batch_id)
    if not batch:
        return {'ok': False, 'error': 'batch no encontrado'}
    if organizacion_id and batch.organizacion_id != organizacion_id:
        return {'ok': False, 'error': 'sin permisos'}

    rows_ppl = ProviderPriceList.query.filter_by(import_batch_id=batch_id).delete(synchronize_session=False)
    rows_po = PrecioObservado.query.filter_by(import_batch_id=batch_id).delete(synchronize_session=False)
    batch.deshecho_at = datetime.utcnow()
    batch.estado = 'deshecho'
    batch.undo_motivo = motivo[:500] if motivo else None

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return {'ok': False, 'error': f'Error deshaciendo: {e}'}
    return {
        'ok': True,
        'batch_id': batch_id,
        'rows_ppl_borradas': rows_ppl,
        'rows_po_borradas': rows_po,
    }


# ============================================================================
# Fase 1 IA presupuestos: carga del CATALOGO COMPLETO (hoja 01, ~6.345 filas)
# ============================================================================

def _parse_hoja_01_catalogo(xlsx_path: str) -> "List[Dict[str, Any]]":
    """Parsea la hoja '01_Catalogo_OBYRA' del OBYRA_base_precios_recursos_v1.

    Es el catalogo normalizado (~6.345 filas). Header en fila 1:
      fuente | proveedor | categoria | tipo_recurso | recurso | descripcion |
      unidad | precio_min | precio_max | precio_unitario

    La hoja 01 NO trae zona (zona=None). Si `unidad` esta vacia se infiere de
    la descripcion. Si `precio_unitario` esta vacio se usa el promedio de
    precio_min/precio_max.
    """
    import openpyxl
    if not os.path.exists(xlsx_path):
        return []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if '01_Catalogo_OBYRA' not in wb.sheetnames:
        wb.close()
        return []
    ws = wb['01_Catalogo_OBYRA']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []

    header = [(_normalizar_unidad_str(str(c)) if c else '') for c in rows[0]]
    col_idx = {h: i for i, h in enumerate(header)}

    def _g(row, key, default=None):
        i = col_idx.get(key)
        if i is None or i >= len(row):
            return default
        return row[i]

    salida = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        recurso = _g(row, 'recurso')
        descripcion = _g(row, 'descripcion') or recurso
        if not recurso and not descripcion:
            continue
        unidad_raw = _g(row, 'unidad')
        unidad = _normalizar_unidad_str(unidad_raw) if unidad_raw else ''
        if not unidad:
            unidad = _inferir_unidad_de_descripcion(str(recurso or descripcion or ''))
        precio_unit = _decimal_safe(_g(row, 'precio_unitario'))
        if not precio_unit:
            pmin = _decimal_safe(_g(row, 'precio_min'))
            pmax = _decimal_safe(_g(row, 'precio_max'))
            if pmin and pmax:
                precio_unit = (pmin + pmax) / Decimal('2')
            elif pmin:
                precio_unit = pmin
            elif pmax:
                precio_unit = pmax
        if not precio_unit:
            continue
        salida.append({
            'descripcion': str(descripcion).strip()[:300],
            'recurso': str(recurso).strip()[:300] if recurso else None,
            'unidad': unidad[:20] or 'un',
            'zona': None,  # hoja 01 no trae zona
            'moneda': 'ARS',
            'precio_unitario': precio_unit,
            'tipo_recurso': (str(_g(row, 'tipo_recurso') or 'material')).strip().lower()[:20],
            'categoria': (str(_g(row, 'categoria') or '')).strip()[:120] or None,
            'fuente_excel': (str(_g(row, 'fuente') or 'OBYRA_base_v1.01_Catalogo'))[:80],
        })
    return salida


def importar_catalogo_base(
    *,
    db,
    xlsx_path: str,
    organizacion_id: int,
    user_id: "Optional[int]" = None,
    perfil: str = 'lista_propia_obyra',
    global_base: bool = False,
) -> "Dict[str, Any]":
    """Importa el catalogo COMPLETO (hoja 01 ~6.345 + hoja 02 zonas + MO) a
    provider_price_list de forma IDEMPOTENTE y eficiente.

    Idempotente: clave de unicidad (org, proveedor=NULL, desc_norm, unidad,
    zona). Correr 2 veces ACTUALIZA precio, no duplica. Se pre-cargan las claves
    existentes en memoria para hacer el UPSERT sin N+1.
    """
    from models.import_batch import ImportBatch
    from models.provider_price_list import ProviderPriceList, normalizar_descripcion_precio

    if not os.path.exists(xlsx_path):
        return {'ok': False, 'error': f'archivo no encontrado: {xlsx_path}'}

    # global_base=True -> los precios se cargan como BASE GLOBAL (org NULL),
    # accesibles por todas las orgs via el fallback de precio_recurso_service.
    # El ImportBatch sigue bajo `organizacion_id` (solo trazabilidad).
    org_precios = None if global_base else organizacion_id

    filename = os.path.basename(xlsx_path)
    checksum = _calcular_checksum(xlsx_path)
    # Idempotencia a nivel batch: ImportBatch tiene UNIQUE(org, checksum). Si ya
    # se importo este mismo archivo para esta org, se REUTILIZA el batch.
    batch = (ImportBatch.query
             .filter_by(organizacion_id=organizacion_id, checksum_sha256=checksum)
             .first())
    if batch is not None:
        batch.perfil = perfil
        batch.estado = 'en_curso'
        batch.started_at = datetime.utcnow()
        batch.user_id = user_id or batch.user_id
    else:
        batch = ImportBatch(
            organizacion_id=organizacion_id, perfil=perfil, filename=filename[:255],
            checksum_sha256=checksum, user_id=user_id,
            estado='en_curso', started_at=datetime.utcnow(),
        )
        db.session.add(batch)
    db.session.flush()

    filas = (
        _parse_hoja_01_catalogo(xlsx_path)
        + _parse_hoja_02_materiales(xlsx_path)
        + _parse_hoja_import_obyra_mano_obra(xlsx_path)
    )

    hoy = date.today()
    try:
        vigencia = date(hoy.year + 1, hoy.month, hoy.day)
    except ValueError:  # 29-feb
        vigencia = date(hoy.year + 1, hoy.month, 28)

    # Pre-cargar claves existentes (proveedor NULL) del scope destino -> UPSERT O(1).
    existentes = {}
    _org_filter = (ProviderPriceList.organizacion_id.is_(None) if org_precios is None
                   else ProviderPriceList.organizacion_id == org_precios)
    for r in (ProviderPriceList.query
              .filter(_org_filter,
                      ProviderPriceList.proveedor_id.is_(None))
              .all()):
        existentes[(r.descripcion_normalizada, r.unidad, r.zona)] = r

    inserted = updated = invalid = 0
    vistas = set()
    for fila in filas:
        desc = fila['descripcion']
        desc_norm = normalizar_descripcion_precio(desc)
        unidad = fila['unidad']
        zona = fila.get('zona')
        precio = fila['precio_unitario']
        moneda = (fila.get('moneda') or 'ARS')[:3]
        if not desc_norm or not precio or precio <= 0:
            invalid += 1
            continue
        key = (desc_norm, unidad, zona)
        notas = (f"Base precios OBYRA. Cat: {fila.get('categoria') or '-'} · "
                 f"tipo: {fila.get('tipo_recurso') or '-'} · "
                 f"Excel: {fila.get('fuente_excel') or '-'}")[:1000]

        row = existentes.get(key)
        if row is not None:
            row.precio_unitario = precio
            row.moneda = moneda
            row.fecha_actualizacion = hoy
            row.vigencia_hasta = vigencia
            row.fuente = perfil[:30]
            row.import_batch_id = batch.id
            row.modalidad = 'compra'
            row.notas = notas
            if key not in vistas:
                updated += 1
        else:
            row = ProviderPriceList(
                organizacion_id=org_precios, proveedor_id=None,
                descripcion=desc[:300], descripcion_normalizada=desc_norm, unidad=unidad,
                precio_unitario=precio, moneda=moneda, fecha_actualizacion=hoy,
                vigencia_hasta=vigencia, fuente=perfil[:30], zona=zona, modalidad='compra',
                import_batch_id=batch.id, created_by_user_id=user_id, notas=notas,
            )
            db.session.add(row)
            existentes[key] = row
            inserted += 1
        vistas.add(key)

    batch.total_input = len(filas)
    batch.total_inserted = inserted
    batch.total_updated = updated
    batch.total_invalid = invalid
    batch.estado = 'completado'
    batch.completed_at = datetime.utcnow()

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return {'ok': False, 'error': f'{type(e).__name__}: {e}'}

    return {
        'ok': True, 'batch_id': batch.id, 'total_input': len(filas),
        'inserted': inserted, 'updated': updated, 'invalid': invalid,
        'unicos_en_run': len(vistas),
    }
