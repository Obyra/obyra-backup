# -*- coding: utf-8 -*-
"""Exportación de un presupuesto ARMADO a Excel (.xlsx).

NO es el pliego original (eso ya se descarga aparte). Es el presupuesto calculado
con los MISMOS números que ve el cliente en el PDF: precio de VENTA (costo directo +
margen comercial), nunca el costo interno ni el % de margen. Se mantiene coherente
con blueprint_presupuestos/pdf_email.pdf_cliente:
  - Fuente de ítems: pipeline_ia_cache si existe (presupuestos IA); si no, los ítems
    en BD (presupuestos manuales).
  - Se omiten los descartados/incluidos/rojos (cache) y los solo-internos/excluidos (BD).
  - IVA sobre el subtotal de venta.
"""
from io import BytesIO
from decimal import Decimal
from datetime import datetime, timedelta
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

_MONEY_FMT = '$#,##0.00'
_QTY_FMT = '#,##0.00'
_AZUL = 'FF366092'
_AZUL_CLARO = 'FFD9E1F2'
_BORDE = Border(*(Side(style='thin') for _ in range(4)))


def _nombre_cliente(presupuesto):
    c = presupuesto.cliente
    if not c:
        return 'Sin especificar'
    return getattr(c, 'nombre_completo', None) or getattr(c, 'nombre', None) or 'Sin especificar'


def _filas_venta(presupuesto):
    """(filas, subtotal, omitidos) con precios de VENTA, mismas reglas que el PDF cliente.

    filas: lista de (descripcion, cantidad_float, unidad, pu_venta_float, total_venta_float).
    """
    from services.margen_comercial import precio_venta

    filas, subtotal, omitidos = [], Decimal('0'), 0
    cache = presupuesto.pipeline_ia_cache if isinstance(presupuesto.pipeline_ia_cache, dict) else {}
    items_cache = cache.get('items') or []

    if items_cache:
        for it in items_cache:
            if it.get('estado') in ('descartado', 'incluido'):
                continue
            costo_total = Decimal(str(it.get('costo_total') or 0))
            if it.get('color') == 'rojo' or costo_total <= 0:
                omitidos += 1
                continue
            pu_costo = it.get('precio_unitario')
            if pu_costo is None:
                pu_costo = it.get('costo_unitario') or 0
            total = precio_venta(costo_total, presupuesto)
            pu = precio_venta(pu_costo, presupuesto)
            filas.append((it.get('descripcion') or '', float(it.get('cantidad') or 0),
                          it.get('unidad') or '', float(pu), float(total)))
            subtotal += Decimal(str(total))
    else:
        # Fallback: ítems en BD (manuales, sin cache IA). El cliente no ve los
        # solo-internos ni los excluidos.
        for it in presupuesto.items.all():   # relación lazy='dynamic'
            if getattr(it, 'solo_interno', False) or getattr(it, 'excluido', False):
                continue
            costo_total = Decimal(str(it.total or 0))
            if costo_total <= 0:
                omitidos += 1
                continue
            total = precio_venta(costo_total, presupuesto)
            pu = precio_venta(Decimal(str(it.precio_unitario or 0)), presupuesto)
            filas.append((it.descripcion or '', float(it.cantidad or 0),
                          it.unidad or '', float(pu), float(total)))
            subtotal += Decimal(str(total))

    return filas, subtotal, omitidos


def export_presupuesto_excel(presupuesto):
    """Genera el .xlsx del presupuesto (precios de venta). Devuelve un BytesIO."""
    filas, subtotal, _omitidos = _filas_venta(presupuesto)

    iva_pct = Decimal(str(presupuesto.iva_porcentaje if presupuesto.iva_porcentaje is not None else 21))
    iva_monto = (subtotal * iva_pct / Decimal('100')).quantize(Decimal('1'))
    total_con_iva = subtotal + iva_monto

    wb = Workbook()
    ws = wb.active
    ws.title = 'Presupuesto'
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18

    hdr_fill = PatternFill(start_color=_AZUL, end_color=_AZUL, fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFFFF', size=11)
    tot_fill = PatternFill(start_color=_AZUL_CLARO, end_color=_AZUL_CLARO, fill_type='solid')

    r = 1
    ws.cell(r, 1, 'PRESUPUESTO').font = Font(bold=True, size=16); r += 1
    fecha = presupuesto.fecha or (presupuesto.fecha_creacion.date()
                                  if getattr(presupuesto, 'fecha_creacion', None) else datetime.now().date())
    ws.cell(r, 1, f'Número: {presupuesto.numero}')
    ws.cell(r, 4, f'Fecha: {fecha.strftime("%d/%m/%Y")}'); r += 1
    ws.cell(r, 1, f'Cliente: {_nombre_cliente(presupuesto)}')
    ws.cell(r, 4, f'Obra: {presupuesto.obra.nombre if presupuesto.obra else "Sin especificar"}'); r += 1
    vig_dias = presupuesto.vigencia_dias or 30
    fecha_vig = presupuesto.fecha_vigencia
    if not fecha_vig:
        try:
            fecha_vig = fecha + timedelta(days=int(vig_dias))
        except Exception:
            fecha_vig = None
    if fecha_vig:
        ws.cell(r, 1, f'Válido hasta: {fecha_vig.strftime("%d/%m/%Y")} ({vig_dias} días)'); r += 1
    r += 1

    # Encabezados de la tabla.
    for col, texto in enumerate(('Descripción', 'Cantidad', 'Unidad', 'Precio Unit.', 'Total'), 1):
        c = ws.cell(r, col, texto)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = _BORDE
        c.alignment = Alignment(horizontal='center')
    r += 1

    for desc, cant, unidad, pu, tot in filas:
        ws.cell(r, 1, desc).border = _BORDE
        cc = ws.cell(r, 2, cant); cc.number_format = _QTY_FMT; cc.border = _BORDE; cc.alignment = Alignment(horizontal='right')
        cu = ws.cell(r, 3, unidad); cu.border = _BORDE; cu.alignment = Alignment(horizontal='center')
        cp = ws.cell(r, 4, pu); cp.number_format = _MONEY_FMT; cp.border = _BORDE
        ct = ws.cell(r, 5, tot); ct.number_format = _MONEY_FMT; ct.border = _BORDE
        r += 1

    r += 1
    def _fila_total(label, valor, negrita=False, fill=False):
        nonlocal r
        cl = ws.cell(r, 4, label)
        cv = ws.cell(r, 5, float(valor)); cv.number_format = _MONEY_FMT
        if negrita:
            cl.font = Font(bold=True, size=12); cv.font = Font(bold=True, size=12)
        if fill:
            cv.fill = tot_fill
        cl.alignment = Alignment(horizontal='right')
        r += 1

    _fila_total('Subtotal:', subtotal)
    _fila_total(f'IVA ({int(iva_pct) if iva_pct == iva_pct.to_integral() else float(iva_pct)}%):', iva_monto)
    _fila_total('TOTAL:', total_con_iva, negrita=True, fill=True)

    r += 1
    nota = ws.cell(r, 1, 'Precios de venta finales. Los importes no incluyen IVA salvo la línea TOTAL.')
    nota.font = Font(italic=True, size=9, color='FF808080')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info('Excel exportado para presupuesto %s (%s ítems)', presupuesto.numero, len(filas))
    return output
