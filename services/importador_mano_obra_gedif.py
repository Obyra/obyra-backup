"""Importador especifico de la planilla Gedif (Fase 5.A).

Estructura esperada del Excel:
  Hoja "Costos por Categoria" (acepta con o sin acento):
    col 1: Categoria
    col 2: Base hora convenio
    col 3: Presentismo
    col 4: HH 50 estimada
    col 5: HH 100
    col 6: Adicional fijo
    col 7: Bruto recibo
    col 8: Neto bolsillo
    col 9: F931
    col 10: Fondo desempleo 12%
    col 11: UOCRA
    col 12: IERIC
    col 13: Comida
    col 14: SAC
    col 15: Vacaciones prop.
    col 16: EPP
    col 17: Total cargas + indirectos
    col 18: Bono anual prorrateado
    col 19: Costo empresa hora
    col 20: Costo jornal 8h
    col 21: Costo mes 176h
    col 22: Fuente base

  Hoja "Supuestos":
    Zona, Mes referencia, Horas mensuales, Jornal,
    Presentismo, Factor HH 50, HH 100, Adicional fijo,
    F931, Fondo desempleo, etc.

Estrategia:
  - Los % proporcionales (presentismo, hh_50, f931, fondo_desempleo) se leen
    desde Supuestos.
  - Los montos absolutos por hora (UOCRA, IERIC, SAC, vacaciones, comida, EPP,
    bono) se reciben como $/hora y se convierten a % o a monto mensual segun
    corresponda al modelo.
  - Despues de cargar, llamamos recalcular() del modelo para que la formula
    OBYRA reconstruya costo_empresa_hora desde los parametros guardados.
  - El valor de costo_empresa_hora del Excel se guarda en
    valores_excel_originales_json para auditoria + comparacion.
"""
from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional, Tuple


HOJA_COSTOS_CANDIDATAS = ('Costos por Categoria', 'Costos por Categoría')
HOJA_SUPUESTOS_CANDIDATAS = ('Supuestos',)


def _norm(s):
    if s is None:
        return ''
    s = str(s).strip().lower()
    s = unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'\s+', ' ', s)


def _to_float(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _detectar_hoja(wb, candidatas):
    """Devuelve el sheet matcheando alguna candidata case-insensitive sin acentos."""
    candidatas_norm = [_norm(c) for c in candidatas]
    for sn in wb.sheetnames:
        if _norm(sn) in candidatas_norm:
            return wb[sn]
    return None


def _parsear_supuestos(ws_sup) -> Dict[str, Any]:
    """Lee la hoja 'Supuestos' clave-valor. Devuelve dict con keys normalizadas."""
    sup = {}
    if not ws_sup:
        return sup
    for row in ws_sup.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        key = _norm(row[0])
        val = row[1] if len(row) > 1 else None
        sup[key] = val
    return sup


def _normalizar_categoria(texto: str) -> str:
    """Mapea 'Oficial especializado' -> 'oficial_especializado', etc."""
    n = _norm(texto)
    if 'oficial especializado' in n:
        return 'oficial_especializado'
    if 'medio oficial' in n:
        return 'medio_oficial'
    if 'sereno' in n:
        return 'sereno'
    if 'ayudante' in n:
        return 'ayudante'
    if 'oficial' in n:
        return 'oficial'
    # Catalogo abierto: si no matchea, se guarda con slug
    return re.sub(r'[^a-z0-9_]', '_', n).strip('_') or 'otro'


def _parsear_periodo(supuestos: Dict[str, Any]) -> str:
    """Convierte 'abr-26' / 'abril 2026' / datetime -> 'YYYY-MM'."""
    raw = supuestos.get('mes referencia')
    if isinstance(raw, datetime):
        return raw.strftime('%Y-%m')
    if not raw:
        return date.today().strftime('%Y-%m')
    s = str(raw).strip().lower()
    # Mapeo de meses ES (acentos ya normalizados por _norm)
    meses = {
        'ene': 1, 'enero': 1, 'feb': 2, 'febrero': 2, 'mar': 3, 'marzo': 3,
        'abr': 4, 'abril': 4, 'may': 5, 'mayo': 5, 'jun': 6, 'junio': 6,
        'jul': 7, 'julio': 7, 'ago': 8, 'agosto': 8, 'sep': 9, 'set': 9, 'septiembre': 9,
        'oct': 10, 'octubre': 10, 'nov': 11, 'noviembre': 11, 'dic': 12, 'diciembre': 12,
    }
    s_norm = _norm(s)
    # 'abr-26' -> ['abr', '26']
    parts = re.split(r'[-/\s]+', s_norm)
    mes = anio = None
    for p in parts:
        if p in meses:
            mes = meses[p]
        elif p.isdigit():
            n = int(p)
            anio = (2000 + n) if n < 100 else n
    if mes and anio:
        return f'{anio:04d}-{mes:02d}'
    return date.today().strftime('%Y-%m')


def _pct_seguro(supuestos: Dict[str, Any], key: str, default: float = 0.0) -> float:
    """Lee un valor que puede venir como decimal (0.20) o pct (20). Devuelve %."""
    v = supuestos.get(key)
    if v is None:
        return default
    try:
        f = float(v)
    except (TypeError, ValueError):
        return default
    # Si es < 1 lo interpretamos como decimal -> *100
    if 0 < f < 1:
        return f * 100
    return f


def parsear_planilla(file_like, *, organizacion_id=None, zona_default='CABA') -> Dict[str, Any]:
    """Parsea el Excel y devuelve estructura lista para crear ManoObraCostoReferencia.

    Args:
      file_like: file-like object (request.files['archivo']).
      organizacion_id: si None, las filas se crean como globales.
      zona_default: zona si la hoja Supuestos no la trae.

    Returns:
      {
        'periodo': 'YYYY-MM',
        'zona': str,
        'parametros_globales': {...},  # supuestos parseados
        'filas': [
          {
            'organizacion_id', 'categoria', 'descripcion', 'zona', 'periodo',
            'valor_hora_convenio', 'horas_mensuales',
            'presentismo_pct', 'hh_50_pct', 'hh_100_pct', 'adicional_fijo_hora',
            'f931_pct', 'fondo_desempleo_pct', 'uocra_pct', 'ieric_pct',
            'sac_pct', 'vacaciones_pct',
            'comida_monto_mes', 'epp_monto_mes', 'bono_anual_monto',
            'parametros_supuestos_json', 'valores_excel_originales_json',
            'fuente', 'observaciones',
          },
          ...
        ],
        'errores': [str],
        'advertencias': [str],
      }
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_like, data_only=True)
    ws_costos = _detectar_hoja(wb, HOJA_COSTOS_CANDIDATAS)
    ws_sup = _detectar_hoja(wb, HOJA_SUPUESTOS_CANDIDATAS)

    if not ws_costos:
        return {
            'periodo': None, 'zona': None, 'parametros_globales': {},
            'filas': [], 'errores': [
                'No se encontró la hoja "Costos por Categoría". Verificá el nombre.'
            ], 'advertencias': [],
        }

    supuestos = _parsear_supuestos(ws_sup)
    periodo = _parsear_periodo(supuestos)
    zona = str(supuestos.get('zona') or zona_default).strip()
    horas_mensuales = int(_to_float(supuestos.get('horas mensuales estimadas')) or 176)
    if horas_mensuales <= 0:
        horas_mensuales = 176

    presentismo_pct = _pct_seguro(supuestos, 'presentismo', 20.0)
    hh_50_pct = _pct_seguro(supuestos, 'factor hh 50', 2.27)
    hh_100_pct = _pct_seguro(supuestos, 'hh 100', 0.0)
    f931_pct = _pct_seguro(supuestos, 'f931', 58.74)
    fondo_desempleo_pct = _pct_seguro(supuestos, 'fondo desempleo', 12.0)

    filas = []
    errores = []
    advertencias = []

    # Header indices (asumir layout fijo Gedif)
    idx = {
        'categoria': 0,
        'base_hora': 1,
        'presentismo': 2,
        'hh_50': 3,
        'hh_100': 4,
        'adicional_fijo': 5,
        'bruto_recibo': 6,
        'neto_bolsillo': 7,
        'f931': 8,
        'fondo_desempleo': 9,
        'uocra': 10,
        'ieric': 11,
        'comida_hora': 12,
        'sac': 13,
        'vacaciones': 14,
        'epp_hora': 15,
        'total_cargas': 16,
        'bono_anual_hora': 17,
        'costo_empresa_hora': 18,
        'costo_jornal_8h': 19,
        'costo_mes_176h': 20,
        'fuente_base': 21,
    }

    for i, row in enumerate(ws_costos.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[idx['categoria']]:
            continue
        cat_raw = str(row[idx['categoria']]).strip()
        cat_norm = _normalizar_categoria(cat_raw)

        base_hora = _to_float(row[idx['base_hora']])
        if base_hora <= 0:
            advertencias.append(f'Fila {i}: categoría "{cat_raw}" sin base_hora_convenio. Saltada.')
            continue

        bruto_recibo = _to_float(row[idx['bruto_recibo']])
        adicional_fijo = _to_float(row[idx['adicional_fijo']])
        comida_hora = _to_float(row[idx['comida_hora']])
        epp_hora = _to_float(row[idx['epp_hora']])
        bono_hora = _to_float(row[idx['bono_anual_hora']])
        uocra_monto = _to_float(row[idx['uocra']])
        ieric_monto = _to_float(row[idx['ieric']])
        sac_monto = _to_float(row[idx['sac']])
        vac_monto = _to_float(row[idx['vacaciones']])
        costo_empresa_hora_excel = _to_float(row[idx['costo_empresa_hora']])

        # Convertir montos $/hora a % sobre bruto_recibo
        def _to_pct(monto):
            return (monto / bruto_recibo * 100.0) if bruto_recibo > 0 else 0.0

        uocra_pct = round(_to_pct(uocra_monto), 2)
        ieric_pct = round(_to_pct(ieric_monto), 2)
        sac_pct = round(_to_pct(sac_monto), 2)
        vacaciones_pct = round(_to_pct(vac_monto), 2)

        # Montos mensuales (hora * horas_mensuales)
        comida_mes = round(comida_hora * horas_mensuales, 2)
        epp_mes = round(epp_hora * horas_mensuales, 2)
        # Bono anual: bono_hora es un monto por hora del bono prorrateado mensual
        # entre 12 meses. Reconstruimos bono anual = bono_hora * horas_mes * 12
        bono_anual = round(bono_hora * horas_mensuales * 12, 2)

        valores_excel = {
            'base_hora_convenio': base_hora,
            'presentismo': _to_float(row[idx['presentismo']]),
            'hh_50_estimada': _to_float(row[idx['hh_50']]),
            'hh_100': _to_float(row[idx['hh_100']]),
            'adicional_fijo': adicional_fijo,
            'bruto_recibo': bruto_recibo,
            'neto_bolsillo': _to_float(row[idx['neto_bolsillo']]),
            'f931_monto': _to_float(row[idx['f931']]),
            'fondo_desempleo_monto': _to_float(row[idx['fondo_desempleo']]),
            'uocra_monto': uocra_monto,
            'ieric_monto': ieric_monto,
            'comida_hora': comida_hora,
            'sac_monto': sac_monto,
            'vacaciones_monto': vac_monto,
            'epp_hora': epp_hora,
            'total_cargas_indirectos': _to_float(row[idx['total_cargas']]),
            'bono_anual_hora': bono_hora,
            'costo_empresa_hora_excel': costo_empresa_hora_excel,
            'costo_jornal_8h_excel': _to_float(row[idx['costo_jornal_8h']]),
            'costo_mes_176h_excel': _to_float(row[idx['costo_mes_176h']]),
            'fuente_base': str(row[idx['fuente_base']] or '') if row[idx['fuente_base']] else None,
        }

        fila = {
            'organizacion_id': organizacion_id,
            'categoria': cat_norm,
            'descripcion': cat_raw[:120],
            'zona': zona[:40],
            'periodo': periodo,
            'fecha_vigencia_desde': date.today(),
            'fecha_vigencia_hasta': None,
            'valor_hora_convenio': base_hora,
            'horas_mensuales': horas_mensuales,
            'presentismo_pct': presentismo_pct,
            'hh_50_pct': hh_50_pct,
            'hh_100_pct': hh_100_pct,
            'adicional_fijo_hora': adicional_fijo,
            'f931_pct': f931_pct,
            'fondo_desempleo_pct': fondo_desempleo_pct,
            'uocra_pct': uocra_pct,
            'ieric_pct': ieric_pct,
            'sac_pct': sac_pct,
            'vacaciones_pct': vacaciones_pct,
            'comida_monto_mes': comida_mes,
            'epp_monto_mes': epp_mes,
            'bono_anual_monto': bono_anual,
            'parametros_supuestos_json': {
                'periodo': periodo,
                'zona': zona,
                'horas_mensuales': horas_mensuales,
                'presentismo_pct': presentismo_pct,
                'hh_50_pct': hh_50_pct,
                'f931_pct': f931_pct,
                'fondo_desempleo_pct': fondo_desempleo_pct,
            },
            'valores_excel_originales_json': valores_excel,
            'fuente': 'planilla_constructora_gedif',
            'observaciones': (
                f'Importado de planilla Gedif {periodo}. Cargas calculadas a partir '
                'del bruto recibo del Excel. Validar periódicamente con convenio vigente.'
            ),
        }
        filas.append(fila)

    return {
        'periodo': periodo,
        'zona': zona,
        'parametros_globales': {
            'horas_mensuales': horas_mensuales,
            'presentismo_pct': presentismo_pct,
            'hh_50_pct': hh_50_pct,
            'hh_100_pct': hh_100_pct,
            'f931_pct': f931_pct,
            'fondo_desempleo_pct': fondo_desempleo_pct,
        },
        'filas': filas,
        'errores': errores,
        'advertencias': advertencias,
    }


def importar_filas(parsed: Dict[str, Any], *, user_id=None) -> Dict[str, Any]:
    """Aplica las filas parseadas a BD. NO commitea — caller maneja la transaction.

    Reglas:
      - Por cada fila, busca por (organizacion_id, categoria, zona, periodo).
        Si existe -> update. Si no -> create.
      - Cuando se crea/actualiza una fila para periodo NUEVO de la misma
        (categoria, zona) y misma organizacion, las anteriores quedan activo=False.
      - Despues de crear/actualizar, llama recalcular() y compara con el
        costo_empresa_hora del Excel para alertar si difiere > 5%.

    Returns:
      {creadas, actualizadas, desactivadas, advertencias_calculo}
    """
    from extensions import db
    from models.mano_obra_costo_referencia import ManoObraCostoReferencia

    creadas = 0
    actualizadas = 0
    desactivadas_total = 0
    adv_calc: List[str] = []

    for fila in parsed.get('filas', []):
        org_id = fila['organizacion_id']
        cat = fila['categoria']
        zona = fila['zona']
        periodo = fila['periodo']

        # Desactivar versiones anteriores de la misma (org, cat, zona) y periodo distinto
        anteriores = ManoObraCostoReferencia.query.filter(
            ManoObraCostoReferencia.organizacion_id == org_id if org_id is not None
            else ManoObraCostoReferencia.organizacion_id.is_(None),
            ManoObraCostoReferencia.categoria == cat,
            ManoObraCostoReferencia.zona == zona,
            ManoObraCostoReferencia.periodo != periodo,
            ManoObraCostoReferencia.activo.is_(True),
        ).all()
        for a in anteriores:
            a.activo = False
            a.fecha_vigencia_hasta = date.today()
            desactivadas_total += 1

        # Buscar fila exacta del periodo
        existente = ManoObraCostoReferencia.query.filter_by(
            organizacion_id=org_id, categoria=cat, zona=zona, periodo=periodo,
        ).first()

        if existente:
            # Update
            for k, v in fila.items():
                if k == 'organizacion_id':
                    continue
                setattr(existente, k, v)
            existente.activo = True
            existente.recalcular()
            existente.updated_at = datetime.utcnow()
            registro = existente
            actualizadas += 1
        else:
            registro = ManoObraCostoReferencia(
                **{k: v for k, v in fila.items() if k != 'parametros_supuestos_json'
                   and k != 'valores_excel_originales_json'},
                created_by_user_id=user_id,
                activo=True,
                confianza='media',
            )
            registro.parametros_supuestos_json = fila.get('parametros_supuestos_json')
            registro.valores_excel_originales_json = fila.get('valores_excel_originales_json')
            registro.recalcular()
            db.session.add(registro)
            creadas += 1

        # Comparar con costo del Excel (auditoria de calidad)
        excel_costo = (fila.get('valores_excel_originales_json') or {}).get('costo_empresa_hora_excel') or 0
        if excel_costo > 0 and registro.costo_empresa_hora:
            diff = abs(float(registro.costo_empresa_hora) - excel_costo) / excel_costo
            if diff > 0.05:
                adv_calc.append(
                    f'{cat} ({zona} {periodo}): recalculado=${float(registro.costo_empresa_hora):.2f} '
                    f'vs Excel=${excel_costo:.2f} (diferencia {diff*100:.1f}%). '
                    'Revisar cargas/aportes.'
                )

    return {
        'creadas': creadas,
        'actualizadas': actualizadas,
        'desactivadas': desactivadas_total,
        'advertencias_calculo': adv_calc,
    }
