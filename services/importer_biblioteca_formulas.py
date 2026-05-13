"""
Importer de la biblioteca de formulas tecnicas - Fase 1 Plan 90%
=================================================================

Parsea `OBYRA_Biblioteca_Formulas_Computos_DETALLADA_SIN_PRECIOS.xlsx` y
popula las tablas formulas_tecnicas + coeficientes_tecnicos.

Estructura esperada del Excel:
  - 17 hojas. Las primeras 13 (01-13) + 14-16 son formulas.
  - Hoja 17_COEFICIENTES_EDITABLES es solo guia de coeficientes.
  - Headers en fila 2 (fila 1 es titulo de la hoja).

Cada formula tiene columnas:
  Codigo | Rubro | Item / concepto | Unidad salida | Formula tecnica
  Excel template | Inputs requeridos | Coeficiente tecnico editable
  Que calcula | Observaciones para Claude

NO conecta con presupuestos. Solo persiste la biblioteca para que el motor
de Fase 2 la consuma despues.

Idempotente: UPSERT por (codigo, organizacion_id=NULL).
"""
import hashlib
import json
import os
import re
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional


# Mapeo hoja -> categoria_calculo + rubro default
_HOJA_META = {
    '01_FORMULAS_GENERALES':     {'rubro': 'General',         'categoria': 'cantidad'},
    '02_MOVIMIENTO_SUELO':        {'rubro': 'Movimiento de suelos', 'categoria': 'cantidad'},
    '03_HORMIGON_DOSIFICACIONES': {'rubro': 'Hormigon',        'categoria': 'consumo_insumo'},
    '04_ACERO':                   {'rubro': 'Acero',           'categoria': 'consumo_insumo'},
    '05_ENCOFRADOS':              {'rubro': 'Encofrados',      'categoria': 'cantidad'},
    '06_MAMPOSTERIA':             {'rubro': 'Mamposteria',     'categoria': 'consumo_insumo'},
    '07_REVOQUES_MORTEROS':       {'rubro': 'Revoques',        'categoria': 'consumo_insumo'},
    '08_CONTRAPISOS_PISOS':       {'rubro': 'Contrapisos y pisos', 'categoria': 'consumo_insumo'},
    '09_CUBIERTAS_CIELORRASOS':   {'rubro': 'Cubiertas y cielorrasos', 'categoria': 'consumo_insumo'},
    '10_PINTURA':                 {'rubro': 'Pintura',         'categoria': 'consumo_insumo'},
    '11_INSTALACIONES':           {'rubro': 'Instalaciones',   'categoria': 'cantidad'},
    '12_MANO_OBRA_FORMULAS':      {'rubro': 'Mano de obra',    'categoria': 'apu'},
    '13_EQUIPOS_FLETES':          {'rubro': 'Equipos y fletes','categoria': 'apu'},
    '14_APU_SIN_PRECIOS':         {'rubro': 'APU',             'categoria': 'apu'},
    '15_PRESUPUESTO_INDICES':     {'rubro': 'Presupuesto',     'categoria': 'presupuesto'},
    '16_AVANCE_CERTIFICACION':    {'rubro': 'Avance',          'categoria': 'avance'},
}

_HOJA_COEFICIENTES = '17_COEFICIENTES_EDITABLES'

# Mapeo del campo "Tipo de coeficiente" del Excel -> codigo de tipo nuestro.
_MAPEO_TIPO_COEFICIENTE = {
    'merma por material':      'merma',
    'dosificacion hormigon':   'dosificacion',
    'dosificacion mortero':    'dosificacion',
    'rendimiento mo':          'rendimiento_mo',
    'rendimiento equipo':      'rendimiento_equipo',
    'capacidad flete':         'capacidad_flete',
    'peso especifico':         'peso_especifico',
    'rendimiento pintura':     'rendimiento_pintura',
    'rendimiento caja piso':   'rendimiento_pieza',
    'horas jornal':            'horas_jornal',
}


def _norm(s: Any) -> str:
    """Normaliza string: trim + lower + sin tildes (basico)."""
    if s is None:
        return ''
    s = str(s).strip().lower()
    # quita tildes basicas para matching
    rep = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ñ': 'n'}
    for k, v in rep.items():
        s = s.replace(k, v)
    return s


def _es_si(val: Any) -> bool:
    n = _norm(val)
    return n in ('si', 'sí', 's', 'yes', 'true', '1', 'si %', 'si  %')


def _checksum_archivo(path: str) -> str:
    h = hashlib.sha256()
    try:
        with open(path, 'rb') as f:
            for chunk in iter(lambda: f.read(65536), b''):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return ''


def _detectar_header_row(rows: List[tuple]) -> int:
    """Detecta la fila del header (la que tiene 'codigo' o 'Codigo' en col 0).
    Devuelve indice 0-based. Default 1 (segunda fila)."""
    for i, row in enumerate(rows[:5]):
        if row and row[0]:
            primer = _norm(row[0])
            if primer in ('codigo', 'código', 'cod', 'concepto', 'tipo de coeficiente'):
                return i
    return 1  # fallback


def _parsear_hoja_formulas(ws, hoja_nombre: str) -> List[Dict[str, Any]]:
    """Lee una hoja de formulas (01-16) y devuelve lista de dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header_idx = _detectar_header_row(rows)
    header = [_norm(c) for c in rows[header_idx]]

    def _idx(*candidates):
        for cand in candidates:
            cand_norm = _norm(cand)
            for i, h in enumerate(header):
                if cand_norm in h:
                    return i
        return None

    # Columnas conocidas (todas las hojas de formulas tienen formato similar)
    col_codigo = _idx('codigo')
    col_rubro = _idx('rubro')
    col_item = _idx('item', 'concepto')
    col_unidad = _idx('unidad salida', 'unidad')
    col_formula_texto = _idx('formula tecnica', 'formula')
    col_formula_expr = _idx('excel template', 'template')
    col_inputs = _idx('inputs requeridos', 'inputs')
    col_coef_edit = _idx('coeficiente tecnico editable', 'coeficiente editable', 'coef. editable')
    col_que_calcula = _idx('que calcula', 'qué calcula')
    col_observaciones = _idx('observaciones', 'observaciones para claude')

    # Hojas 14-16 tienen otro formato: 'Concepto | Formula | Excel template | Depende de | Regla para Claude'
    # o 'Codigo | Concepto | Formula | Excel template | Uso en OBYRA'.
    col_concepto = _idx('concepto')
    col_depende = _idx('depende', 'uso en obyra', 'regla para claude')

    salida = []
    orden = 0
    for row in rows[header_idx + 1:]:
        if not row or all(c is None for c in row):
            continue

        # Codigo: si la hoja no tiene columna codigo (14_APU), generamos uno
        # a partir del rubro + concepto.
        codigo = None
        if col_codigo is not None and col_codigo < len(row):
            codigo = row[col_codigo]
        if not codigo and col_concepto is not None:
            # 14_APU_SIN_PRECIOS no tiene codigo; usamos el concepto.
            concepto_txt = row[col_concepto] if col_concepto < len(row) else None
            if concepto_txt:
                codigo = f"APU-{_norm(concepto_txt).replace(' ', '_').upper()[:30]}"
        if not codigo:
            continue
        codigo = str(codigo).strip()
        if not codigo or codigo.startswith(('---', '01 ', '02 ', '03 ', '04 ',
                                              '05 ', '06 ', '07 ', '08 ',
                                              '09 ', '10 ', '11 ', '12 ',
                                              '13 ', '14 ', '15 ', '16 ')):
            # Fila de titulo de hoja, no es codigo real.
            continue
        if codigo.lower() == 'codigo' or codigo.lower() == 'código':
            continue

        orden += 1

        def _get(idx, default=None):
            if idx is None or idx >= len(row):
                return default
            v = row[idx]
            if v is None:
                return default
            return str(v).strip()

        # Resolver fields segun hoja
        rubro = _get(col_rubro) or _HOJA_META.get(hoja_nombre, {}).get('rubro', '')
        item_concepto = _get(col_item) or _get(col_concepto) or ''
        unidad = _get(col_unidad) or ''
        formula_texto = _get(col_formula_texto) or _get(_idx('formula')) or ''
        formula_expr = _get(col_formula_expr) or ''
        inputs_req = _get(col_inputs) or _get(col_depende) or ''
        coef_edit_raw = _get(col_coef_edit)
        que_calcula = _get(col_que_calcula) or ''
        observaciones = _get(col_observaciones) or _get(col_depende) or ''

        if not item_concepto:
            continue

        salida.append({
            'codigo': codigo[:40],
            'rubro': rubro[:80],
            'item_concepto': item_concepto[:300],
            'unidad_salida': unidad[:40] if unidad else None,
            'formula_texto': formula_texto or None,
            'formula_expr': formula_expr or None,
            'inputs_requeridos': inputs_req or None,
            'usa_coeficiente_editable': _es_si(coef_edit_raw),
            'categoria_calculo': _HOJA_META.get(hoja_nombre, {}).get('categoria'),
            'que_calcula': que_calcula or None,
            'observaciones': observaciones or None,
            'hoja_origen': hoja_nombre,
            'orden': orden,
        })

    return salida


def _parsear_hoja_coeficientes(ws) -> List[Dict[str, Any]]:
    """Lee la hoja 17_COEFICIENTES_EDITABLES.
    Estructura: Tipo de coeficiente | Ejemplo | Unidad | Donde se usa |
                Debe ser editable | Comentario"""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header_idx = _detectar_header_row(rows)
    header = [_norm(c) for c in rows[header_idx]]

    def _idx(*candidates):
        for cand in candidates:
            cand_norm = _norm(cand)
            for i, h in enumerate(header):
                if cand_norm in h:
                    return i
        return None

    col_tipo = _idx('tipo de coeficiente', 'tipo')
    col_ejemplo = _idx('ejemplo')
    col_unidad = _idx('unidad')
    col_donde = _idx('donde se usa', 'dónde se usa')
    col_editable = _idx('debe ser editable', 'editable')
    col_comentario = _idx('comentario')

    salida = []
    orden = 0
    for row in rows[header_idx + 1:]:
        if not row or all(c is None for c in row):
            continue

        def _get(idx, default=None):
            if idx is None or idx >= len(row):
                return default
            v = row[idx]
            if v is None:
                return default
            return str(v).strip()

        tipo_raw = _get(col_tipo) or ''
        if not tipo_raw or tipo_raw.startswith('17 '):
            continue

        tipo_norm = _norm(tipo_raw)
        tipo_codigo = _MAPEO_TIPO_COEFICIENTE.get(tipo_norm, tipo_norm.replace(' ', '_')[:40])

        ejemplo = _get(col_ejemplo) or ''
        unidad = _get(col_unidad) or ''
        donde = _get(col_donde) or ''
        comentario = _get(col_comentario) or ''

        orden += 1

        # Generamos un codigo unico de coeficiente: TIPO_EJEMPLOPRIMERAPALABRA.
        primera_palabra = re.split(r'[,;/\s]+', ejemplo.strip())[0] if ejemplo else f'item{orden}'
        codigo = f"{tipo_codigo.upper()}_{primera_palabra[:30]}".replace(' ', '_')

        salida.append({
            'codigo': codigo[:80],
            'tipo': tipo_codigo[:40],
            'descripcion': (tipo_raw + ' - ' + ejemplo)[:300] if ejemplo else tipo_raw[:300],
            'valor_default': 0,  # el Excel no trae valor, son guias. El super admin lo edita despues.
            'unidad': unidad[:40] if unidad else None,
            'rubro': None,
            'aplicable_a': ejemplo[:120] if ejemplo else None,
            'notas': (donde + '\n' + comentario).strip() or None,
        })

    return salida


def importar_excel_formulas(
    *,
    db,
    xlsx_path: str,
    user_id: Optional[int] = None,
) -> Dict[str, Any]:
    """Importa el Excel completo: formulas (hojas 01-16) + coeficientes (17).

    UPSERT por (codigo, organizacion_id IS NULL): todas las filas se cargan
    como GLOBAL (organizacion_id=NULL) porque son la biblioteca base curada
    por OBYRA. Los overrides por tenant son responsabilidad de Fase 5.

    Returns:
      dict con contadores + batch_id + errores.
    """
    import openpyxl
    from models.formulas import FormulaTecnica, Coeficiente, ImportBatchFormulas

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f'No existe el archivo: {xlsx_path}')

    batch_id = uuid.uuid4().hex[:12]
    batch = ImportBatchFormulas(
        batch_id=batch_id,
        filename=os.path.basename(xlsx_path)[:255],
        checksum=_checksum_archivo(xlsx_path),
        estado='en_curso',
        started_by_user_id=user_id,
    )
    db.session.add(batch)
    db.session.flush()

    contadores = {
        'formulas_creadas': 0,
        'formulas_actualizadas': 0,
        'coeficientes_creados': 0,
        'coeficientes_actualizados': 0,
        'invalidos': 0,
        'errores': [],
        'batch_id': batch_id,
    }

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        batch.estado = 'fallido'
        batch.errores_json = json.dumps({'error': str(e)})
        batch.finished_at = datetime.utcnow()
        db.session.commit()
        raise

    # 1) Formulas (hojas 01-16)
    for hoja_nombre in wb.sheetnames:
        if hoja_nombre not in _HOJA_META:
            continue
        ws = wb[hoja_nombre]
        filas = _parsear_hoja_formulas(ws, hoja_nombre)
        for f in filas:
            try:
                existente = FormulaTecnica.query.filter(
                    FormulaTecnica.codigo == f['codigo'],
                    FormulaTecnica.organizacion_id.is_(None),
                ).first()
                if existente:
                    for k, v in f.items():
                        setattr(existente, k, v)
                    existente.batch_id = batch_id
                    existente.updated_at = datetime.utcnow()
                    contadores['formulas_actualizadas'] += 1
                else:
                    nueva = FormulaTecnica(
                        organizacion_id=None,  # global
                        batch_id=batch_id,
                        **f,
                    )
                    db.session.add(nueva)
                    contadores['formulas_creadas'] += 1
            except Exception as e:
                contadores['invalidos'] += 1
                contadores['errores'].append({
                    'tipo': 'formula',
                    'codigo': f.get('codigo'),
                    'error': f'{type(e).__name__}: {str(e)[:200]}',
                })

    # 2) Coeficientes (hoja 17)
    if _HOJA_COEFICIENTES in wb.sheetnames:
        ws = wb[_HOJA_COEFICIENTES]
        filas_coef = _parsear_hoja_coeficientes(ws)
        for c in filas_coef:
            try:
                existente = Coeficiente.query.filter(
                    Coeficiente.codigo == c['codigo'],
                    Coeficiente.organizacion_id.is_(None),
                ).first()
                if existente:
                    for k, v in c.items():
                        setattr(existente, k, v)
                    existente.batch_id = batch_id
                    existente.updated_at = datetime.utcnow()
                    contadores['coeficientes_actualizados'] += 1
                else:
                    nuevo = Coeficiente(
                        organizacion_id=None,  # global
                        batch_id=batch_id,
                        **c,
                    )
                    db.session.add(nuevo)
                    contadores['coeficientes_creados'] += 1
            except Exception as e:
                contadores['invalidos'] += 1
                contadores['errores'].append({
                    'tipo': 'coeficiente',
                    'codigo': c.get('codigo'),
                    'error': f'{type(e).__name__}: {str(e)[:200]}',
                })

    # Cerrar batch
    batch.estado = 'completado' if contadores['invalidos'] == 0 else 'completado_con_errores'
    batch.formulas_creadas = contadores['formulas_creadas']
    batch.formulas_actualizadas = contadores['formulas_actualizadas']
    batch.coeficientes_creados = contadores['coeficientes_creados']
    batch.coeficientes_actualizados = contadores['coeficientes_actualizados']
    batch.invalidos = contadores['invalidos']
    if contadores['errores']:
        batch.errores_json = json.dumps(contadores['errores'][:50])
    batch.finished_at = datetime.utcnow()

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        contadores['errores'].append({'tipo': 'commit', 'error': str(e)[:300]})
        raise

    return contadores
