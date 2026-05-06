"""
Importador simple de Excel de licitacion.

Crea un Presupuesto + ItemPresupuesto a partir de un Excel con columnas
basicas (descripcion, unidad, cantidad, precio_unit opcional).

Flujo:
  GET  /presupuestos/importar-licitacion         -> formulario
  POST /presupuestos/importar-licitacion         -> procesa archivo y crea presupuesto
"""
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from flask import (render_template, request, flash, redirect, url_for,
                   current_app, abort, jsonify)
from flask_login import login_required, current_user

from blueprint_presupuestos import presupuestos_bp
from extensions import db
from models import Presupuesto, ItemPresupuesto, Cliente
from services.memberships import get_current_org_id


def _parse_decimal(val, default=0):
    if val is None or val == '':
        return Decimal(str(default))
    try:
        # Aceptar formato con coma decimal (AR)
        s = str(val).replace('.', '').replace(',', '.')
        # Si ya viene con punto decimal (EN), mejor intentar directo primero
        try:
            return Decimal(str(val))
        except (InvalidOperation, ValueError):
            return Decimal(s)
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(str(default))


def _norm_header(t):
    """Normaliza un valor de header: sin acentos, lower, trimmed."""
    if t is None:
        return ''
    import unicodedata
    return unicodedata.normalize('NFKD', str(t)).encode('ascii', 'ignore').decode().lower().strip()


def _match_desc(h_norm):
    """True si el texto normalizado parece columna de descripcion.
    NO incluye 'item' standalone — esa suele ser la columna de codigo/numero.
    """
    if not h_norm:
        return False
    if h_norm in ('descripcion', 'descripcion item', 'descripcion del item',
                  'detalle', 'articulo', 'concepto', 'tarea', 'rubro/tarea'):
        return True
    if h_norm.startswith('descrip') or 'descripcion' in h_norm:
        return True
    if 'concepto' in h_norm or 'detalle' in h_norm:
        return True
    return False


def _match_unidad(h_norm):
    if not h_norm:
        return False
    if h_norm in ('un', 'und', 'unid', 'unidad', 'um', 'medida', 'u/m', 'unid.', 'un.'):
        return True
    if h_norm.startswith('unidad'):
        return True
    # Fusionado: ej "MAT + M.O UNIDAD" en JMG. Aceptar si contiene 'unidad' como
    # palabra y NO es columna de precio unitario.
    if 'unidad' in h_norm and 'unitario' not in h_norm and 'precio' not in h_norm:
        return True
    return False


def _match_cantidad(h_norm):
    if not h_norm:
        return False
    if h_norm in ('cant', 'cantidad', 'cdad', 'q', 'cant.', 'cantidades'):
        return True
    if h_norm.startswith('cant'):
        return True
    return False


def _match_precio(h_norm):
    """Match para columna de precio unitario.
    Prioriza 'unitario' o 'P. Unit'. Acepta 'TOTAL' como sub-header de 'PRECIO UNITARIO ITEM'
    en plantillas tipo JMG (donde PRECIO UNITARIO ITEM tiene MO/MAT/TOTAL como sub-headers).
    """
    if not h_norm:
        return False
    if 'total' in h_norm and 'unitario' not in h_norm and 'precio unitario' not in h_norm:
        # 'precio total item' -> NO es precio unitario
        return False
    if h_norm in ('precio unitario', 'precio unit', 'p. unit', 'p.unit', 'pu',
                  'precio/un', 'precio unit.', 'pu.', 'precio',
                  'precio unitario item', 'precio unitario item total',
                  'precio unitario total'):
        return True
    if h_norm.startswith('$/un') or h_norm.startswith('$ un') or h_norm == '$/un':
        return True
    if 'precio unitario' in h_norm:
        return True
    if '/un' in h_norm and 'total' not in h_norm:
        return True
    return False


def _detectar_columnas(headers):
    """Busca indices de columnas en el header. Retorna dict desc/unidad/cantidad/precio."""
    idx = {'desc': None, 'unidad': None, 'cantidad': None, 'precio': None}
    norms = [_norm_header(h) for h in headers]

    for i, hn in enumerate(norms):
        if not hn:
            continue
        if idx['desc'] is None and _match_desc(hn):
            idx['desc'] = i
        if idx['unidad'] is None and _match_unidad(hn):
            idx['unidad'] = i
        if idx['cantidad'] is None and _match_cantidad(hn):
            idx['cantidad'] = i
        if idx['precio'] is None and _match_precio(hn):
            idx['precio'] = i
    return idx


def _construir_header_combinado(rows, header_idx):
    """Soporta plantillas con header en 2 filas (ej: JMG).
    Devuelve (headers_efectivos, data_start_offset).

    Si la fila header_idx por si sola permite detectar desc+cantidad, devuelve esa fila
    y data_start = header_idx+1. Si falta cantidad/unidad y existe header_idx+1, fusiona
    ambas filas (texto N+1 reemplaza None de N o se concatena), re-detecta y, si mejora,
    devuelve la fusionada con data_start = header_idx+2.
    """
    h1 = list(rows[header_idx]) if header_idx < len(rows) else []
    idx1 = _detectar_columnas(h1)
    if idx1['desc'] is not None and idx1['cantidad'] is not None:
        return h1, header_idx + 1, idx1

    if header_idx + 1 >= len(rows):
        return h1, header_idx + 1, idx1
    h2 = list(rows[header_idx + 1])
    n = max(len(h1), len(h2))
    merged = []
    for i in range(n):
        v1 = h1[i] if i < len(h1) else None
        v2 = h2[i] if i < len(h2) else None
        if v1 is not None and v2 is not None and str(v1).strip() and str(v2).strip():
            merged.append(f'{v1} {v2}')
        elif v2 is not None and str(v2).strip():
            merged.append(v2)
        else:
            merged.append(v1)
    idx2 = _detectar_columnas(merged)
    # Aceptar la fusion solo si gana cantidad (lo que el header de 1 fila no tenia)
    if idx2['cantidad'] is not None and idx2['desc'] is not None:
        return merged, header_idx + 2, idx2
    return h1, header_idx + 1, idx1


def _buscar_fila_header(rows, max_search=50):
    """Busca la primera fila que contenga keywords de header tipico.

    Retorna (idx, score). Score mas alto = mejor match.
    """
    import unicodedata

    def _norm(t):
        if t is None:
            return ''
        return unicodedata.normalize('NFKD', str(t)).encode('ascii', 'ignore').decode().lower()

    keywords = ['descripcion', 'detalle', 'cantidad', 'cant', 'unidad', 'precio', 'rubro', 'concepto', 'item']
    best_idx = None
    best_score = 0

    for i, row in enumerate(rows[:max_search]):
        score = 0
        for cell in row:
            if cell is None:
                continue
            cell_norm = _norm(cell).strip()
            for kw in keywords:
                if cell_norm == kw or kw in cell_norm:
                    score += 1
                    break
        if score > best_score:
            best_score = score
            best_idx = i

    if best_score >= 2:  # al menos 2 keywords coinciden
        return best_idx, best_score
    return None, 0


def _es_numero_validos(val):
    """True si val parsea a Decimal > 0."""
    try:
        d = _parse_decimal(val)
        return d > 0
    except Exception:
        return False


def _parece_nota_o_referencia(desc):
    """True si la descripcion parece una nota/aclaracion o referencia de
    sector — NO un rubro real. Heuristica conservadora basada en patrones
    observados en planillas reales (JMG y similares).

    Casos cubiertos:
      - "S/ 7P", "S/ 1ºSS", "S/ 1P A S/6P": referencia a piso/sector.
      - "POR LO DICHO, NO SE REPITEN ITEMS EN LA PRESENTE PLANILLA": nota.
      - Oraciones largas con coma: los rubros reales no usan coma.
    """
    if not desc:
        return False
    d = desc.strip()
    d_up = d.upper()
    # Sector/piso: "S/ 7P", "S/1ºSS", "S/1P A S/6P"
    if d_up.startswith('S/'):
        return True
    # Frases tipicas de nota
    notas = (
        'POR LO DICHO', 'PRESENTE PLANILLA', 'NO SE REPIT',
        'VER OBSERVACI', 'VER NOTA', 'SEGUN DETALLE',
    )
    if any(n in d_up for n in notas):
        return True
    if d_up.startswith('NOTA:') or d_up.startswith('ACLARACION'):
        return True
    # Oraciones con coma + varias palabras: los rubros reales no usan coma
    if ',' in d and len(d.split()) > 5:
        return True
    return False


def _parsear_xlsx(file_stream):
    """Lee el xlsx (todas las hojas) y retorna lista de items con etapa detectada.

    Detecta rubros/secciones (filas sin cantidad pero con descripcion) y los asigna
    como 'etapa_nombre' a los items que vienen abajo.
    """
    import openpyxl
    import re
    wb = openpyxl.load_workbook(file_stream, data_only=True)

    items_total = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx, score = _buscar_fila_header(rows)
        if header_idx is None:
            continue

        headers, start_data, idx = _construir_header_combinado(rows, header_idx)

        if idx['desc'] is None or idx['cantidad'] is None:
            continue

        # Saltar sub-header tipico solo si la fila siguiente parece extra ("$/un", "$/total")
        if start_data < len(rows):
            sub = rows[start_data]
            sub_txt = ' '.join(str(c).lower() for c in sub if c is not None)
            cells_no_empty = len([c for c in sub if c is not None and str(c).strip() != ''])
            if ('$/' in sub_txt or '$' in sub_txt) and cells_no_empty <= 8:
                start_data += 1

        # Detectar columna de "rubro" o numero (ej: "1", "1.01", "2.05")
        # En el header puede llamarse "Rubro", "Codigo", "Cod", "Item No", etc.
        idx_rubro = None
        for i, h in enumerate(headers):
            if h is None:
                continue
            h_low = str(h).lower().strip()
            if h_low in ('rubro', 'codigo', 'cod', 'cod.', 'item', 'no', 'nro', 'n°', '#',
                         ' item', 'item ', 'codigo item', 'numero', 'item no'):
                idx_rubro = i
                break

        # Patron para detectar codigos jerarquicos: "1", "1.01", "2.5.3"
        patron_codigo = re.compile(r'^\d+(\.\d+)*$')

        etapa_actual = None  # nombre de la seccion/rubro actual

        for row_i, row in enumerate(rows[start_data:], start=start_data + 1):
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue

            desc_raw = row[idx['desc']] if idx['desc'] < len(row) else None
            if not desc_raw or str(desc_raw).strip() == '':
                continue
            desc = str(desc_raw).strip()

            cant_raw = row[idx['cantidad']] if idx['cantidad'] is not None and idx['cantidad'] < len(row) else None
            tiene_cantidad = cant_raw is not None and str(cant_raw).strip() != '' and _es_numero_validos(cant_raw)

            # Detectar si es una fila de RUBRO/seccion (sin cantidad o con codigo de nivel 1)
            codigo = None
            if idx_rubro is not None and idx_rubro < len(row) and row[idx_rubro] is not None:
                codigo = str(row[idx_rubro]).strip()

            # Es rubro si: no tiene cantidad valida AND tiene una descripcion en mayusculas o codigo de 1 nivel
            # Palabras que NO son rubros sino totales/subtotales
            desc_lower = desc.lower()
            es_total = any(kw in desc_lower for kw in [
                'total', 'subtotal', 'resumen', 'gran total', 'suma',
                'iva', 'descuento', 'bonificacion', 'bonificación',
            ])

            es_rubro = False
            if (not tiene_cantidad and not es_total
                    and not _parece_nota_o_referencia(desc)):
                # Caso 1: codigo de un solo nivel (ej: "1", "2", "3")
                if codigo and patron_codigo.match(codigo) and '.' not in codigo:
                    es_rubro = True
                # Caso 2: descripcion sin codigo pero parece header (mayusculas o sin numeros)
                elif desc.upper() == desc and len(desc) > 3 and len(desc) < 80:
                    es_rubro = True

            if es_rubro:
                etapa_actual = desc[:100]
                continue

            # Si es un total/subtotal, saltar
            if es_total and not tiene_cantidad:
                continue

            # Si no tiene cantidad valida y no es rubro, saltar (puede ser subtotal, total, etc.)
            if not tiene_cantidad:
                continue

            try:
                cantidad_d = _parse_decimal(cant_raw)
            except Exception:
                continue
            if cantidad_d <= 0:
                continue

            unidad = row[idx['unidad']] if idx['unidad'] is not None and idx['unidad'] < len(row) else None
            precio = row[idx['precio']] if idx['precio'] is not None and idx['precio'] < len(row) else 0

            try:
                precio_d = _parse_decimal(precio)
            except Exception:
                precio_d = Decimal('0')

            items_total.append({
                'descripcion': desc[:300],
                'unidad': str(unidad).strip()[:20] if unidad else 'un',
                'cantidad': cantidad_d,
                'precio_unitario': precio_d,
                'total': cantidad_d * precio_d,
                'etapa_nombre': etapa_actual,
                'codigo': codigo,
                # Fase 6.A: trazabilidad de origen para multi-archivo
                'hoja_origen': sname,
                'fila_origen': row_i,
                'columna_descripcion_origen': _col_letter(idx.get('desc')) if idx.get('desc') is not None else None,
            })

    return items_total if items_total else None


def _diagnosticar_xlsx(file_stream):
    """Inspecciona el archivo y devuelve un diagnostico util cuando el parser
    no logra detectar items. Retorna dict serializable a JSON.
    """
    import openpyxl
    diag = {'hojas': [], 'columnas_esperadas': ['descripcion', 'unidad', 'cantidad', 'precio_unitario']}
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True, read_only=True)
    except Exception as e:
        diag['error_lectura'] = f'{type(e).__name__}: {str(e)[:200]}'
        return diag

    for sname in wb.sheetnames:
        try:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            diag['hojas'].append({'nombre': sname, 'error': f'{type(e).__name__}: {e}'})
            continue

        no_vacias_preview = []
        for i, r in enumerate(rows):
            if not r or all(c is None or str(c).strip() == '' for c in r):
                continue
            preview_cells = [(str(c)[:60] if c is not None else '') for c in r[:14]]
            no_vacias_preview.append({'fila': i + 1, 'celdas': preview_cells})
            if len(no_vacias_preview) >= 10:
                break

        # Intentar detectar header y columnas
        header_idx, score = _buscar_fila_header(rows)
        cols_detectadas = None
        header_efectivo = None
        if header_idx is not None:
            headers_eff, _start, idx_eff = _construir_header_combinado(rows, header_idx)
            header_efectivo = [(str(h)[:50] if h is not None else None) for h in headers_eff]
            cols_detectadas = {
                'desc': _col_letter(idx_eff.get('desc')),
                'unidad': _col_letter(idx_eff.get('unidad')),
                'cantidad': _col_letter(idx_eff.get('cantidad')),
                'precio': _col_letter(idx_eff.get('precio')),
            }

        # Motivo del fallo
        if header_idx is None:
            motivo = 'No se encontró fila de encabezado con keywords (descripcion/cantidad/unidad/precio).'
        elif cols_detectadas and cols_detectadas.get('desc') is None:
            motivo = 'No se identificó la columna de descripción.'
        elif cols_detectadas and cols_detectadas.get('cantidad') is None:
            motivo = 'No se identificó la columna de cantidad.'
        else:
            motivo = 'No se encontraron filas con cantidad numérica > 0.'

        diag['hojas'].append({
            'nombre': sname,
            'dimensiones': f'{ws.max_row}x{ws.max_column}',
            'header_fila_detectada': (header_idx + 1) if header_idx is not None else None,
            'header_efectivo': header_efectivo,
            'columnas_detectadas': cols_detectadas,
            'primeras_filas_no_vacias': no_vacias_preview,
            'motivo': motivo,
        })
    try:
        wb.close()
    except Exception:
        pass
    return diag


def _col_letter(idx):
    """Convierte indice 0-based a letra Excel: 0->'A', 1->'B', ..., 25->'Z', 26->'AA'."""
    if idx is None:
        return None
    s = ''
    n = idx
    while n >= 0:
        s = chr(ord('A') + (n % 26)) + s
        n = n // 26 - 1
        if n < 0:
            break
    return s


def _guardar_archivo_temp(file_storage):
    """Guarda el archivo subido en /tmp con un UUID y retorna el token."""
    import uuid
    import os
    token = str(uuid.uuid4())
    tmp_dir = '/tmp/obyra_excel_import'
    os.makedirs(tmp_dir, exist_ok=True)
    path = os.path.join(tmp_dir, f"{token}.xlsx")
    file_storage.stream.seek(0)
    file_storage.save(path)
    return token, path


def _path_archivo_temp(token):
    import os
    path = os.path.join('/tmp/obyra_excel_import', f"{token}.xlsx")
    return path if os.path.exists(path) else None


def _persistir_pliego(presu, temp_path, nombre_original=None):
    """Copia el Excel del pliego desde /tmp a static/uploads/pliegos/ para que
    quede como documento contractual del presupuesto (y luego de la obra).

    Guarda la ruta relativa (partiendo de 'static/') en presu.archivo_pliego_path
    y un nombre amigable en presu.archivo_pliego_nombre. No falla si no puede
    copiar — solo loggea y sigue.
    """
    import os
    import shutil
    try:
        base_dir = os.path.join('static', 'uploads', 'pliegos')
        os.makedirs(base_dir, exist_ok=True)
        dest_rel = os.path.join('uploads', 'pliegos', f'pres_{presu.id}.xlsx')
        dest_abs = os.path.join('static', dest_rel)
        shutil.copyfile(temp_path, dest_abs)
        presu.archivo_pliego_path = dest_rel.replace('\\', '/')
        presu.archivo_pliego_nombre = (
            nombre_original or f'Pliego_{presu.numero}.xlsx'
        )
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.warning(f'No se pudo persistir pliego: {e}')


def _leer_preview(path, max_rows=30):
    """Retorna las primeras N filas + columnas detectadas."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True, max_row=max_rows)):
            rows.append([('' if c is None else str(c)[:80]) for c in row])
        max_cols = max((len(r) for r in rows), default=0)
        # Padding para que todas las filas tengan misma cantidad de columnas
        for r in rows:
            while len(r) < max_cols:
                r.append('')
        sheets.append({'nombre': sname, 'rows': rows, 'num_cols': max_cols})
    return sheets


def _crear_perfil_tecnico_desde_form(presupuesto, form, user_id):
    """Crea el ProjectTechnicalProfile a partir del form del importador.

    Solo si el usuario tildo el checkbox 'perfil_tecnico_incluir'. Si no,
    no hace nada (perfil queda vacio y se completa despues desde el detalle).

    Es fail-safe: cualquier error de validacion se loguea y NO rompe la
    importacion (el presupuesto ya quedo creado).
    """
    if (form.get('perfil_tecnico_incluir') or '').strip() not in ('1', 'on', 'true'):
        return None

    payload = {
        'tipo_obra': form.get('ptp_tipo_obra'),
        'naturaleza_proyecto': form.get('naturaleza_proyecto'),  # heredamos del campo principal
        'tipo_estructura': form.get('ptp_tipo_estructura'),
        'tipo_fundacion': form.get('ptp_tipo_fundacion'),
        'sistema_constructivo': form.get('ptp_sistema_constructivo'),
        'cantidad_pisos': form.get('ptp_cantidad_pisos'),
        'cantidad_subsuelos': form.get('ptp_cantidad_subsuelos'),
        'tiene_planta_baja': form.get('ptp_tiene_planta_baja'),
        'tiene_terraza': form.get('ptp_tiene_terraza'),
        'superficie_total_m2': form.get('ptp_superficie_total_m2'),
        'superficie_por_planta_m2': form.get('ptp_superficie_por_planta_m2'),
        'altura_promedio_piso_m': form.get('ptp_altura_promedio_piso_m'),
        'espesor_losa_cm': form.get('ptp_espesor_losa_cm'),
        'cantidad_torres': form.get('ptp_cantidad_torres'),
        'cantidad_unidades_funcionales': form.get('ptp_cantidad_unidades_funcionales'),
        'cantidad_cocheras': form.get('ptp_cantidad_cocheras'),
        'criterio_distribucion': form.get('ptp_criterio_distribucion'),
        'cantidades_excel_son_totales': form.get('ptp_cantidades_excel_son_totales'),
    }

    try:
        from services.perfil_tecnico_service import upsert_perfil_tecnico
        result = upsert_perfil_tecnico(
            presupuesto=presupuesto,
            payload=payload,
            user_id=user_id,
            autogenerar_niveles=True,
        )
        try:
            from models.audit import registrar_audit
            registrar_audit(
                accion='crear_perfil_tecnico',
                entidad='presupuesto',
                entidad_id=presupuesto.id,
                detalle=(
                    f'Importacion Excel: tipo={result["profile"].tipo_obra} '
                    f'pisos={result["profile"].cantidad_pisos} '
                    f'criterio={result["profile"].criterio_distribucion} '
                    f'niveles={result["niveles_generados"]}'
                ),
            )
        except Exception:
            pass
        db.session.commit()
        return result
    except ValueError as ve:
        current_app.logger.warning(f'Perfil tecnico invalido en importacion: {ve}')
        db.session.rollback()
        return None
    except Exception:
        current_app.logger.exception('Error guardando perfil tecnico desde importacion')
        db.session.rollback()
        return None


def _crear_presupuesto_desde_items(org_id, cliente_id, numero, vigencia_dias, nombre_obra, items, modo_licitacion=True, ubicacion=None, ubicacion_lat=None, ubicacion_lng=None, ubicacion_normalizada=None, naturaleza_proyecto=None):
    """Helper compartido para crear el presupuesto + items.

    Si modo_licitacion=True (default), los precios del Excel se ignoran y
    los items se crean con precio_unitario=0 para que los carguen los proveedores.

    `naturaleza_proyecto` (obra_nueva | remodelacion | ampliacion) se persiste
    en `datos_proyecto` para que el modulo Ejecutivo (APU) filtre/sugiera las
    etapas internas que aplican al tipo de obra.
    """
    import json
    from calculadora_ia import normalizar_naturaleza_proyecto
    datos_proyecto = {
        'nombre_obra': nombre_obra,
        'modo_creacion': 'excel_licitacion',
        'modo_licitacion': modo_licitacion,
        'naturaleza_proyecto': normalizar_naturaleza_proyecto(naturaleza_proyecto),
    }
    if ubicacion:
        datos_proyecto['ubicacion'] = ubicacion
    presu = Presupuesto(
        numero=numero,
        organizacion_id=org_id,
        cliente_id=cliente_id,
        fecha=date.today(),
        vigencia_dias=vigencia_dias,
        estado='borrador',
        datos_proyecto=json.dumps(datos_proyecto),
        ubicacion_texto=(ubicacion or None),
        ubicacion_normalizada=(ubicacion_normalizada or None),
        geo_latitud=ubicacion_lat,
        geo_longitud=ubicacion_lng,
        currency='ARS',
    )
    db.session.add(presu)
    db.session.flush()

    from services.etapa_matcher import matchear_etapa_para_item

    subtotal = Decimal('0')
    for it in items:
        # Intentar matchear con etapa estandar del sistema
        etapa_excel = it.get('etapa_nombre')
        etapa_estandar = matchear_etapa_para_item(it['descripcion'], etapa_excel)
        # Usar la estandar si hay match, sino la del Excel
        etapa_final = etapa_estandar or etapa_excel

        precio = Decimal('0') if modo_licitacion else it['precio_unitario']
        total_item = it['cantidad'] * precio

        ip = ItemPresupuesto(
            presupuesto_id=presu.id,
            tipo='material',
            descripcion=it['descripcion'],
            unidad=it['unidad'],
            cantidad=it['cantidad'],
            precio_unitario=precio,
            total=total_item,
            origen='importado',
            currency='ARS',
            etapa_nombre=etapa_final,
        )
        db.session.add(ip)
        subtotal += total_item

    presu.subtotal_materiales = subtotal
    presu.total_sin_iva = subtotal
    iva_pct = Decimal('21')
    presu.iva_porcentaje = iva_pct
    presu.total_con_iva = subtotal + (subtotal * iva_pct / Decimal('100'))

    # Audit: presupuesto importado desde Excel
    try:
        from models.audit import registrar_audit
        registrar_audit(
            accion='importar_excel',
            entidad='presupuesto',
            entidad_id=presu.id,
            detalle=f'Presupuesto {numero} importado desde Excel con {len(items)} items',
        )
    except Exception:
        pass

    db.session.commit()
    return presu


@presupuestos_bp.route('/importar-licitacion', methods=['GET', 'POST'])
@login_required
def importar_licitacion():
    """Formulario + procesamiento de importacion de Excel de licitacion."""
    if not current_user.puede_gestionar():
        flash('No tienes permisos para crear presupuestos', 'danger')
        return redirect(url_for('presupuestos.lista'))

    org_id = get_current_org_id()
    if not org_id:
        flash('Seleccioná una organización.', 'warning')
        return redirect(url_for('auth.seleccionar_organizacion'))

    clientes = Cliente.query.filter_by(
        organizacion_id=org_id, activo=True
    ).order_by(Cliente.nombre).all()

    if request.method == 'GET':
        # Numero sugerido
        fecha_hoy = date.today().strftime('%Y%m%d')
        ultimo = (Presupuesto.query
                  .filter_by(organizacion_id=org_id)
                  .filter(Presupuesto.numero.like(f'PRES-{fecha_hoy}-%'))
                  .order_by(Presupuesto.id.desc()).first())
        if ultimo and ultimo.numero:
            try:
                n = int(ultimo.numero.split('-')[-1]) + 1
                numero_sug = f"PRES-{fecha_hoy}-{n:03d}"
            except (ValueError, IndexError):
                numero_sug = f"PRES-{fecha_hoy}-001"
        else:
            numero_sug = f"PRES-{fecha_hoy}-001"

        import os as _os
        return render_template('presupuestos/importar_licitacion.html',
                               clientes=clientes,
                               numero_sugerido=numero_sug,
                               google_maps_key=_os.environ.get('GOOGLE_MAPS_API_KEY', ''))

    # POST — Fase 6.A: aceptamos N archivos. El primero (legacy) sigue
    # llamandose 'archivo_excel'. getlist() devuelve la lista completa.
    archivos_excel = request.files.getlist('archivo_excel')
    # Filtrar vacios (input multiple puede mandar entries vacias)
    archivos_excel = [f for f in archivos_excel if f and f.filename]
    archivo = archivos_excel[0] if archivos_excel else None  # compat con codigo viejo
    nombre_obra = (request.form.get('nombre_obra') or '').strip()
    cliente_id = request.form.get('cliente_id', type=int)
    numero = (request.form.get('numero') or '').strip()
    vigencia_dias = request.form.get('vigencia_dias', 30, type=int)
    modo_licitacion = request.form.get('modo_licitacion') == '1'
    ubicacion = (request.form.get('ubicacion') or '').strip() or None
    ubicacion_lat = request.form.get('ubicacion_lat', type=float)
    ubicacion_lng = request.form.get('ubicacion_lng', type=float)
    ubicacion_normalizada = (request.form.get('ubicacion_normalizada') or '').strip() or None
    naturaleza_proyecto = (request.form.get('naturaleza_proyecto') or 'obra_nueva').strip()

    if not archivos_excel:
        flash('Subí al menos un archivo Excel', 'danger')
        return redirect(url_for('presupuestos.importar_licitacion'))

    # Validar formato de cada archivo
    for f in archivos_excel:
        if not f.filename.lower().endswith(('.xlsx', '.xls')):
            flash(f'El archivo "{f.filename}" no es .xlsx ni .xls', 'danger')
            return redirect(url_for('presupuestos.importar_licitacion'))

    if not nombre_obra:
        flash('Ingresá un nombre de obra', 'danger')
        return redirect(url_for('presupuestos.importar_licitacion'))

    # Detectar duplicado: presupuesto reciente con mismo nombre_obra
    forzar = request.form.get('forzar_duplicado') == '1'
    if not forzar:
        from datetime import timedelta
        umbral = datetime.utcnow() - timedelta(hours=24)
        # Buscar presupuestos del mismo dia con descripcion similar,
        # excluyendo los eliminados/perdidos (no deben bloquear nueva creacion).
        estados_excluidos = ('eliminado', 'perdido')
        if hasattr(Presupuesto, 'created_at'):
            recientes = Presupuesto.query.filter(
                Presupuesto.organizacion_id == org_id,
                Presupuesto.created_at >= umbral,
                ~Presupuesto.estado.in_(estados_excluidos),
            ).all()
        else:
            recientes = Presupuesto.query.filter(
                Presupuesto.organizacion_id == org_id,
                Presupuesto.fecha == date.today(),
                ~Presupuesto.estado.in_(estados_excluidos),
            ).all()
        for p in recientes:
            try:
                import json
                dp = json.loads(p.datos_proyecto) if p.datos_proyecto else {}
                if (dp.get('nombre_obra') or '').strip().lower() == nombre_obra.lower():
                    flash(
                        f'Ya existe un presupuesto reciente para "{nombre_obra}" ({p.numero}). '
                        f'Si querés crear uno nuevo igual, tildá "Crear de todas formas" y reintentá.',
                        'warning'
                    )
                    return redirect(url_for('presupuestos.importar_licitacion'))
            except Exception:
                continue

    # Numero: usar sugerido si no se provee
    if not numero:
        fecha_hoy = date.today().strftime('%Y%m%d')
        ultimo = (Presupuesto.query
                  .filter_by(organizacion_id=org_id)
                  .filter(Presupuesto.numero.like(f'PRES-{fecha_hoy}-%'))
                  .order_by(Presupuesto.id.desc()).first())
        if ultimo and ultimo.numero:
            try:
                n = int(ultimo.numero.split('-')[-1]) + 1
                numero = f"PRES-{fecha_hoy}-{n:03d}"
            except (ValueError, IndexError):
                numero = f"PRES-{fecha_hoy}-001"
        else:
            numero = f"PRES-{fecha_hoy}-001"

    # Validar cliente
    if cliente_id:
        c = Cliente.query.filter_by(id=cliente_id, organizacion_id=org_id).first()
        if not c:
            flash('Cliente inválido', 'danger')
            return redirect(url_for('presupuestos.importar_licitacion'))

    # Si el usuario tildó "revisar antes de importar", saltar al mapeo manual.
    # El mapeo manual no soporta multi-archivo; si hay varios, avisamos.
    revisar_antes = request.form.get('revisar_antes') == '1'
    if revisar_antes and len(archivos_excel) > 1:
        flash(
            'El "Revisar antes de importar" solo funciona con 1 archivo. '
            'Subí los demás después desde "Agregar archivo de licitación" en el detalle.',
            'warning',
        )
        archivos_excel = archivos_excel[:1]
        archivo = archivos_excel[0]

    if revisar_antes:
        # Flujo legacy: 1 solo archivo, mapeo manual
        try:
            token, path = _guardar_archivo_temp(archivo)
        except Exception as e:
            flash(f'Error al guardar archivo: {str(e)[:200]}', 'danger')
            return redirect(url_for('presupuestos.importar_licitacion'))
        return redirect(url_for(
            'presupuestos.importar_licitacion_mapear',
            token=token,
            nombre_obra=nombre_obra,
            cliente_id=cliente_id or '',
            numero=numero,
            vigencia_dias=vigencia_dias,
            modo_licitacion='1' if modo_licitacion else '0',
            ubicacion=ubicacion or '',
            ubicacion_lat=ubicacion_lat or '',
            ubicacion_lng=ubicacion_lng or '',
            ubicacion_normalizada=ubicacion_normalizada or '',
            naturaleza_proyecto=naturaleza_proyecto or 'obra_nueva',
        ))

    # Auto-detect multi-archivo (Fase 6.A): crear presupuesto vacio,
    # importar los N archivos, consolidar items.
    try:
        presu = _crear_presupuesto_desde_items(
            org_id, cliente_id, numero, vigencia_dias, nombre_obra, items=[],
            modo_licitacion=modo_licitacion, ubicacion=ubicacion,
            ubicacion_lat=ubicacion_lat, ubicacion_lng=ubicacion_lng,
            ubicacion_normalizada=ubicacion_normalizada,
            naturaleza_proyecto=naturaleza_proyecto,
        )
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error creando presupuesto vacio')
        flash(f'Error al crear presupuesto: {str(e)[:200]}', 'danger')
        return redirect(url_for('presupuestos.importar_licitacion'))

    # Importar todos los archivos
    try:
        from services.import_pliego_service import importar_pliego_multi
        resumen = importar_pliego_multi(
            presupuesto=presu,
            archivos_files=archivos_excel,
            user_id=current_user.id if current_user.is_authenticated else None,
            modo_licitacion=modo_licitacion,
        )
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error en import_pliego_multi')
        flash('Error inesperado al procesar los archivos.', 'danger')
        return redirect(url_for('presupuestos.importar_licitacion'))

    # Si TODOS los archivos fallaron, rollback total y eliminar el presupuesto.
    # Importante: NO hacer "from models.budgets import Presupuesto" aca dentro,
    # porque eso convierte Presupuesto en variable local en TODA la funcion y
    # rompe la rama del GET (UnboundLocalError). Presupuesto ya esta importado
    # al top del archivo via "from models import Presupuesto, ...".
    if resumen['total_archivos_ok'] == 0:
        try:
            Presupuesto.query.filter_by(id=presu.id).delete()
            db.session.commit()
        except Exception:
            db.session.rollback()
        # Construir mensaje de errores
        errores = '; '.join(
            f'{a["filename"]}: {a.get("error_message") or "error"}'
            for a in resumen['archivos']
        )[:500]
        flash(f'Ningún archivo se pudo importar. Errores: {errores}', 'danger')
        return redirect(url_for('presupuestos.importar_licitacion'))

    # Recalcular subtotal del presupuesto (los items se crearon dentro del servicio)
    try:
        presu.calcular_totales()
    except Exception:
        pass

    # Compat legacy: archivo_pliego_path apunta al primer archivo importado OK
    primer_ok = next(
        (a for a in resumen['archivos'] if a['estado'] == 'importado'), None
    )
    if primer_ok and not presu.archivo_pliego_path:
        from models.presupuesto_archivo import PresupuestoArchivo
        pa = PresupuestoArchivo.query.get(primer_ok['archivo_id'])
        if pa:
            presu.archivo_pliego_path = pa.file_path
            presu.archivo_pliego_nombre = pa.filename_original

    # Perfil tecnico (Fase 2). Fail-safe.
    perfil_result = _crear_perfil_tecnico_desde_form(
        presu, request.form, current_user.id if current_user.is_authenticated else None
    )

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error commit final importacion')

    # Mensaje de resumen al usuario
    msg_partes = []
    n_ok = resumen['total_archivos_ok']
    n_err = resumen['total_archivos_error']
    n_dup = resumen['duplicados_skipped']
    n_items = resumen['total_items_importados']
    if n_ok == 1 and n_err == 0 and n_dup == 0:
        msg = f'Presupuesto {numero} creado con {n_items} ítems'
        if modo_licitacion:
            msg += ' en $0 (modo licitación). Armá el Ejecutivo para desglosar en materiales.'
        else:
            msg += ' con precios del Excel.'
        msg_partes.append(msg)
    else:
        msg_partes.append(
            f'Presupuesto {numero} creado: {n_ok} archivo(s) importado(s), '
            f'{n_items} ítems consolidados.'
        )
        if n_err:
            errores_breve = '; '.join(
                f'{a["filename"]}' for a in resumen['archivos']
                if a['estado'] in ('error', 'rechazado')
            )[:300]
            msg_partes.append(f'{n_err} con error: {errores_breve}.')
        if n_dup:
            msg_partes.append(f'{n_dup} duplicados ignorados.')
    if perfil_result and perfil_result.get('niveles_generados'):
        msg_partes.append(
            f'Perfil técnico cargado y {perfil_result["niveles_generados"]} '
            'niveles generados.'
        )
    flash(' '.join(msg_partes), 'success' if n_err == 0 else 'warning')
    return redirect(url_for('presupuestos.ejecutivo_vista', id=presu.id))


@presupuestos_bp.route('/importar-licitacion/mapear/<token>', methods=['GET', 'POST'])
@login_required
def importar_licitacion_mapear(token):
    """Vista de mapeo manual de columnas del Excel."""
    if not current_user.puede_gestionar():
        abort(403)
    org_id = get_current_org_id()
    if not org_id:
        return redirect(url_for('auth.seleccionar_organizacion'))

    path = _path_archivo_temp(token)
    if not path:
        flash('Sesión expirada. Volvé a subir el archivo.', 'warning')
        return redirect(url_for('presupuestos.importar_licitacion'))

    nombre_obra = request.values.get('nombre_obra', '').strip()
    cliente_id_raw = request.values.get('cliente_id', '')
    cliente_id = int(cliente_id_raw) if cliente_id_raw and cliente_id_raw.isdigit() else None
    numero = request.values.get('numero', '').strip()
    vigencia_dias = int(request.values.get('vigencia_dias', 30) or 30)
    modo_licitacion = request.values.get('modo_licitacion', '1') == '1'
    ubicacion = (request.values.get('ubicacion') or '').strip() or None
    try:
        ubicacion_lat = float(request.values.get('ubicacion_lat') or '') or None
    except (ValueError, TypeError):
        ubicacion_lat = None
    try:
        ubicacion_lng = float(request.values.get('ubicacion_lng') or '') or None
    except (ValueError, TypeError):
        ubicacion_lng = None
    ubicacion_normalizada = (request.values.get('ubicacion_normalizada') or '').strip() or None
    naturaleza_proyecto = (request.values.get('naturaleza_proyecto') or 'obra_nueva').strip()

    if request.method == 'POST':
        # Procesar mapeo manual
        try:
            sheet_idx = int(request.form.get('sheet_idx', 0))
            header_row = int(request.form.get('header_row', 1)) - 1  # 1-based en UI
            col_desc = int(request.form.get('col_desc', -1))
            col_unidad_raw = request.form.get('col_unidad', '')
            col_cantidad = int(request.form.get('col_cantidad', -1))
            col_precio_raw = request.form.get('col_precio', '')

            col_unidad = int(col_unidad_raw) if col_unidad_raw not in ('', '-1') else None
            col_precio = int(col_precio_raw) if col_precio_raw not in ('', '-1') else None

            if col_desc < 0 or col_cantidad < 0:
                flash('Tenés que mapear al menos las columnas Descripción y Cantidad.', 'danger')
                return redirect(request.url)

            # Leer hoja seleccionada e importar
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            sheet = wb.worksheets[sheet_idx]
            rows = list(sheet.iter_rows(values_only=True))

            items = []
            for row in rows[header_row + 1:]:
                if not row or all(c is None or str(c).strip() == '' for c in row):
                    continue
                desc = row[col_desc] if col_desc < len(row) else None
                if not desc or str(desc).strip() == '':
                    continue
                cant_raw = row[col_cantidad] if col_cantidad < len(row) else 0
                if cant_raw is None or str(cant_raw).strip() == '':
                    continue
                try:
                    cant = _parse_decimal(cant_raw)
                except Exception:
                    continue
                if cant <= 0:
                    continue

                unidad = row[col_unidad] if col_unidad is not None and col_unidad < len(row) else None
                precio = row[col_precio] if col_precio is not None and col_precio < len(row) else 0
                try:
                    precio_d = _parse_decimal(precio)
                except Exception:
                    precio_d = Decimal('0')

                items.append({
                    'descripcion': str(desc).strip()[:300],
                    'unidad': str(unidad).strip()[:20] if unidad else 'un',
                    'cantidad': cant,
                    'precio_unitario': precio_d,
                    'total': cant * precio_d,
                })

            if not items:
                flash('No se detectaron items válidos con ese mapeo. Revisá la fila de encabezado y las columnas.', 'warning')
                return redirect(request.url)

            presu = _crear_presupuesto_desde_items(
                org_id, cliente_id, numero, vigencia_dias, nombre_obra, items,
                modo_licitacion=modo_licitacion, ubicacion=ubicacion,
                ubicacion_lat=ubicacion_lat, ubicacion_lng=ubicacion_lng,
                ubicacion_normalizada=ubicacion_normalizada,
                naturaleza_proyecto=naturaleza_proyecto,
            )
            _persistir_pliego(presu, path)
            import os
            try: os.remove(path)
            except Exception: pass
            suf = ' en $0 (modo licitación)' if modo_licitacion else ''
            flash(f'Presupuesto {numero} creado con {len(items)} ítems desde mapeo manual{suf}.', 'success')
            return redirect(url_for('presupuestos.ejecutivo_vista', id=presu.id))
        except Exception as e:
            db.session.rollback()
            current_app.logger.exception("Error en mapeo manual")
            flash(f'Error: {str(e)[:200]}', 'danger')
            return redirect(request.url)

    # GET: mostrar preview + selectores
    try:
        sheets = _leer_preview(path, max_rows=40)
    except Exception as e:
        flash(f'Error leyendo Excel: {str(e)[:200]}', 'danger')
        return redirect(url_for('presupuestos.importar_licitacion'))

    return render_template('presupuestos/importar_licitacion_mapear.html',
                           token=token,
                           sheets=sheets,
                           nombre_obra=nombre_obra,
                           cliente_id=cliente_id,
                           numero=numero,
                           vigencia_dias=vigencia_dias)


# ============================================================
# Fase 6.A: endpoints multi-archivo
# ============================================================

@presupuestos_bp.route('/<int:id>/archivos', methods=['GET'])
@login_required
def listar_archivos_presupuesto(id):
    """JSON con archivos del presupuesto. Para refrescar el panel sin reload."""
    org_id = get_current_org_id()
    presupuesto = Presupuesto.query.filter_by(id=id, organizacion_id=org_id).first_or_404()
    if not _puede_ver_presupuesto(presupuesto):
        return jsonify(ok=False, error='No autorizado'), 403

    from services.import_pliego_service import listar_archivos
    archivos = listar_archivos(presupuesto.id)
    return jsonify(ok=True, archivos=[a.to_dict() for a in archivos])


@presupuestos_bp.route('/<int:id>/archivos', methods=['POST'])
@login_required
def agregar_archivo_presupuesto(id):
    """Agrega 1 o varios archivos a un presupuesto existente.

    Solo permitido si el presupuesto esta en borrador o enviado.
    """
    if not current_user.puede_gestionar():
        return jsonify(ok=False, error='Sin permisos'), 403

    org_id = get_current_org_id()
    presupuesto = Presupuesto.query.filter_by(id=id, organizacion_id=org_id).first_or_404()

    if presupuesto.estado not in ('borrador', 'enviado'):
        return jsonify(ok=False, error=(
            f'No se pueden agregar archivos a un presupuesto en estado {presupuesto.estado}.'
        )), 400

    archivos_excel = request.files.getlist('archivo_excel')
    archivos_excel = [f for f in archivos_excel if f and f.filename]
    if not archivos_excel:
        return jsonify(ok=False, error='Subi al menos un archivo.'), 400
    for f in archivos_excel:
        if not f.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify(ok=False, error=f'Formato no soportado en "{f.filename}"'), 400

    try:
        from services.import_pliego_service import importar_pliego_multi
        resumen = importar_pliego_multi(
            presupuesto=presupuesto,
            archivos_files=archivos_excel,
            user_id=current_user.id if current_user.is_authenticated else None,
            modo_licitacion=True,  # respetamos default; cambio futuro: leer del presupuesto
        )
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Error agregando archivos')
        return jsonify(ok=False, error='Error inesperado al procesar archivos.'), 500

    # Recalcular subtotales
    try:
        presupuesto.calcular_totales()
        db.session.commit()
    except Exception:
        db.session.rollback()

    return jsonify(ok=True, **resumen)


@presupuestos_bp.route('/<int:pid>/archivos/<int:aid>/descargar', methods=['GET'])
@login_required
def descargar_archivo_pa(pid, aid):
    """Descarga protegida de un archivo del presupuesto.

    Valida acceso (org match o super admin). Crea audit log.
    """
    # IMPORTANTE: abort ya esta importado al top del archivo (linea 14-15).
    # NO hacer "from flask import abort" aca dentro: convertiria abort en
    # variable local en toda la funcion y rompe los usos previos al import.
    from flask import send_file

    presupuesto = Presupuesto.query.get_or_404(pid)
    if not _puede_ver_presupuesto(presupuesto):
        abort(403)

    org_id = get_current_org_id()
    es_super = bool(getattr(current_user, 'is_super_admin', False))

    try:
        from services.import_pliego_service import descargar_archivo_path
        archivo, path_abs = descargar_archivo_path(
            aid, organizacion_id_caller=org_id, es_super_admin=es_super,
        )
    except PermissionError:
        abort(403)
    except FileNotFoundError:
        abort(404)

    # Validar match presupuesto/archivo
    if archivo.presupuesto_id != pid:
        abort(404)

    # Audit
    try:
        from models.audit import registrar_audit
        registrar_audit(
            accion='archivo_descargado',
            entidad='presupuesto_archivo',
            entidad_id=archivo.id,
            detalle=f'pres={pid} file="{archivo.filename_original}"',
        )
        db.session.commit()
    except Exception:
        db.session.rollback()

    return send_file(
        path_abs,
        as_attachment=True,
        download_name=archivo.filename_original,
        mimetype=archivo.mime_type or 'application/octet-stream',
    )


@presupuestos_bp.route('/<int:pid>/archivos/<int:aid>/eliminar', methods=['POST'])
@login_required
def eliminar_archivo_pa(pid, aid):
    """Elimina un archivo de licitacion + sus items + sus observaciones.

    Es destructivo intencional: cuando el usuario equivoca un upload, necesita
    deshacer la importacion. NO se puede recuperar (salvo re-subir el Excel).

    Requiere:
      - rol admin/pm
      - presupuesto en estado 'borrador' o 'enviado'
      - el archivo pertenece al mismo presupuesto y org

    Hace:
      1. Soft-delete del PresupuestoArchivo (deleted_at = now)
      2. Hard-delete de items_presupuesto con archivo_origen_id = aid
      3. Hard-delete de precio_observado con origen_archivo_id = aid
      4. Borra archivo fisico de disco si existe
      5. Recalcula totales del presupuesto

    Devuelve resumen con conteos.
    """
    from datetime import datetime

    if not current_user.puede_gestionar():
        return jsonify(ok=False, error='Sin permisos'), 403

    org_id = get_current_org_id()
    presupuesto = Presupuesto.query.filter_by(id=pid, organizacion_id=org_id).first()
    if not presupuesto:
        return jsonify(ok=False, error='Presupuesto no encontrado'), 404
    if presupuesto.estado not in ('borrador', 'enviado'):
        return jsonify(ok=False, error=(
            f'No se puede eliminar archivos: presupuesto en estado {presupuesto.estado}'
        )), 400

    from models.presupuesto_archivo import PresupuestoArchivo
    from models.budgets import ItemPresupuesto

    archivo = PresupuestoArchivo.query.filter_by(id=aid, presupuesto_id=pid).first()
    if not archivo:
        return jsonify(ok=False, error='Archivo no encontrado en este presupuesto'), 404
    if archivo.deleted_at is not None:
        return jsonify(ok=False, error='El archivo ya fue eliminado'), 400

    items_a_borrar = (ItemPresupuesto.query
                      .filter_by(presupuesto_id=pid, archivo_origen_id=aid)
                      .all())
    items_count = len(items_a_borrar)

    obs_count = 0
    try:
        from models.precio_observado import PrecioObservado
        obs_count = (PrecioObservado.query
                     .filter_by(origen_archivo_id=aid)
                     .count())
    except Exception:
        pass

    # Borrar precio_observado primero (FK a archivo es SET NULL pero igual los limpiamos)
    try:
        if obs_count > 0:
            from models.precio_observado import PrecioObservado
            PrecioObservado.query.filter_by(origen_archivo_id=aid).delete(
                synchronize_session=False
            )
    except Exception:
        current_app.logger.exception('Error borrando precio_observado')

    # Borrar items
    try:
        ItemPresupuesto.query.filter_by(
            presupuesto_id=pid, archivo_origen_id=aid,
        ).delete(synchronize_session=False)
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error borrando items')
        return jsonify(ok=False, error=f'Error borrando items: {type(e).__name__}'), 500

    # Soft-delete del archivo
    archivo.deleted_at = datetime.utcnow()

    # Borrar archivo fisico (best-effort, no bloquea)
    try:
        import os
        if archivo.file_path and os.path.exists(archivo.file_path):
            os.remove(archivo.file_path)
    except Exception:
        current_app.logger.exception(f'No se pudo borrar archivo fisico {archivo.file_path}')

    # Audit
    try:
        from models.audit import registrar_audit
        registrar_audit(
            accion='archivo_eliminado',
            entidad='presupuesto_archivo',
            entidad_id=archivo.id,
            detalle=(f'pres={pid} file="{archivo.filename_original}" '
                     f'items_borrados={items_count} obs_borradas={obs_count}'),
        )
    except Exception:
        pass

    # Recalcular totales
    try:
        presupuesto.calcular_totales()
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error commiteando eliminar_archivo_pa')
        return jsonify(ok=False, error=f'Error al guardar: {type(e).__name__}'), 500

    return jsonify(
        ok=True,
        archivo_id=aid,
        items_eliminados=items_count,
        observaciones_eliminadas=obs_count,
        total_items_restantes=presupuesto.items.count(),
    )


@presupuestos_bp.route('/precios-observados/stats', methods=['GET'])
@login_required
def precios_observados_stats():
    """Etapa 1 base IA — JSON read-only con resumen de observaciones de precio
    capturadas durante imports. Sirve para validar que el hook esta funcionando.

    Solo admin/pm. Scoped al organizacion_id del usuario.
    """
    if current_user.role not in ('admin', 'administrador', 'pm', 'project_manager'):
        return jsonify(ok=False, error='Sin permisos'), 403

    from models.precio_observado import PrecioObservado
    org_id = get_current_org_id()

    base_q = PrecioObservado.query.filter_by(organizacion_id=org_id)

    total = base_q.count()

    por_origen = {}
    for row in (db.session.query(PrecioObservado.origen_tipo, db.func.count(PrecioObservado.id))
                .filter_by(organizacion_id=org_id)
                .group_by(PrecioObservado.origen_tipo).all()):
        por_origen[row[0]] = row[1]

    por_tipo = {}
    for row in (db.session.query(PrecioObservado.tipo_recurso, db.func.count(PrecioObservado.id))
                .filter_by(organizacion_id=org_id)
                .group_by(PrecioObservado.tipo_recurso).all()):
        por_tipo[row[0]] = row[1]

    por_archivo = []
    for row in (db.session.query(
                    PrecioObservado.origen_archivo_id,
                    db.func.count(PrecioObservado.id))
                .filter_by(organizacion_id=org_id)
                .filter(PrecioObservado.origen_archivo_id.isnot(None))
                .group_by(PrecioObservado.origen_archivo_id)
                .order_by(db.func.count(PrecioObservado.id).desc())
                .limit(20).all()):
        por_archivo.append({'archivo_id': row[0], 'observaciones': row[1]})

    ultimas = (base_q
               .order_by(PrecioObservado.created_at.desc())
               .limit(10)
               .all())

    return jsonify(
        ok=True,
        total=total,
        por_origen_tipo=por_origen,
        por_tipo_recurso=por_tipo,
        top_archivos=por_archivo,
        ultimas_10=[o.to_dict() for o in ultimas],
    )


def _puede_ver_presupuesto(presupuesto):
    """Helper: org match o super admin."""
    if not current_user.is_authenticated:
        return False
    if getattr(current_user, 'is_super_admin', False):
        return True
    org_id = get_current_org_id()
    return presupuesto.organizacion_id == org_id
