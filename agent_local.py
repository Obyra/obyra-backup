from flask import Blueprint, render_template, redirect, url_for, request, jsonify, make_response, send_file, flash
from flask_login import current_user, login_required
from models import Obra, Presupuesto, ItemInventario, Usuario, ConsultaAgente, db, Organizacion
from sqlalchemy import func, desc, or_, and_
from datetime import datetime, timedelta
import importlib.util
import time
import json
import io

# Optional openpyxl import (robusto si la librería no está instalada)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ModuleNotFoundError:  # pragma: no cover - executed when optional dep missing
    openpyxl = None  # type: ignore[assignment]
    Font = PatternFill = Alignment = None  # type: ignore[assignment]
    OPENPYXL_AVAILABLE = False

REPORTLAB_AVAILABLE = importlib.util.find_spec("reportlab") is not None
if REPORTLAB_AVAILABLE:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
else:  # pragma: no cover - executed only when optional deps missing
    colors = None  # type: ignore[assignment]
    letter = A4 = None  # type: ignore[assignment]
    inch = 72  # type: ignore[assignment]

    def _missing_reportlab(*args, **kwargs):
        raise RuntimeError(
            "La librería reportlab no está instalada. Ejecuta 'pip install reportlab' para habilitar la generación de PDFs."
        )

    SimpleDocTemplate = Table = TableStyle = Paragraph = Spacer = _missing_reportlab  # type: ignore[assignment]
    getSampleStyleSheet = ParagraphStyle = _missing_reportlab  # type: ignore[assignment]

agent_bp = Blueprint('agent_local', __name__)


@agent_bp.route('/diagnostico-agent')
@login_required
def diagnostico_agent():
    if not current_user.rol == 'administrador':
        return redirect(url_for('reportes.dashboard'))  # Redirige si no es admin

    return render_template('asistente/diagnostico_agent.html', user=current_user)


@agent_bp.route('/api/agente-consulta', methods=['POST'])
@login_required
def procesar_consulta():
    """Procesa consultas del agente IA con datos reales de la organización"""
    start_time = time.time()
    consulta_registro = None
    
    try:
        data = request.get_json()
        consulta_texto = data.get('consulta', '')
        consulta = consulta_texto.lower()
        
        # Crear registro de auditoría
        consulta_registro = ConsultaAgente(
            organizacion_id=current_user.organizacion_id,
            usuario_id=current_user.id,
            consulta_texto=consulta_texto,
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')
        )
        
        # Determinar tipo de consulta
        if 'obra' in consulta or 'construcción' in consulta or 'proyecto' in consulta:
            consulta_registro.tipo_consulta = 'obra'
        elif 'presupuesto' in consulta or 'costo' in consulta or 'precio' in consulta:
            consulta_registro.tipo_consulta = 'presupuesto'
        elif 'inventario' in consulta or 'stock' in consulta or 'material' in consulta:
            consulta_registro.tipo_consulta = 'inventario'
        elif 'usuario' in consulta or 'equipo' in consulta or 'persona' in consulta:
            consulta_registro.tipo_consulta = 'usuario'
        else:
            consulta_registro.tipo_consulta = 'general'
        
        # Obtener datos reales de la organización del usuario
        org_id = current_user.organizacion_id
        
        # Consultar obras activas
        # Usar eager loading para prevenir N+1 queries
        from sqlalchemy.orm import joinedload

        obras_activas = Obra.query.options(
            joinedload(Obra.cliente),
            joinedload(Obra.responsable)
        ).filter_by(
            organizacion_id=org_id,
            estado='en_progreso'
        ).all()

        obras_total = Obra.query.filter_by(organizacion_id=org_id).count()

        # Consultar presupuestos
        presupuestos_aprobados = Presupuesto.query.options(
            joinedload(Presupuesto.cliente),
            joinedload(Presupuesto.obra)
        ).filter_by(
            organizacion_id=org_id,
            estado='aprobado'
        ).all()

        # Consultar inventario con stock bajo
        items_stock_bajo = ItemInventario.query.options(
            joinedload(ItemInventario.organizacion),
            joinedload(ItemInventario.categoria)
        ).filter(
            ItemInventario.organizacion_id == org_id,
            ItemInventario.stock_actual <= ItemInventario.stock_minimo
        ).all()
        
        # Generar respuesta basada en la consulta
        if 'obra' in consulta or 'construcción' in consulta or 'proyecto' in consulta:
            if 'avanzada' in consulta or 'progreso' in consulta:
                # Consulta sobre obra más avanzada
                obras_todas = Obra.query.filter_by(organizacion_id=org_id).all()
                if not obras_todas:
                    respuesta = "No tienes obras registradas en el sistema."
                else:
                    obra_max_progreso = max(obras_todas, key=lambda x: x.progreso or 0)
                    respuesta = f"La obra más avanzada es <strong>{obra_max_progreso.nombre}</strong> con {obra_max_progreso.progreso or 0}% de progreso.<br>"
                    respuesta += f"Cliente: {obra_max_progreso.cliente}<br>"
                    if obra_max_progreso.fecha_fin_estimada:
                        respuesta += f"Fecha estimada: {obra_max_progreso.fecha_fin_estimada.strftime('%d/%m/%Y')}"
            else:
                # Consulta general sobre obras
                if not obras_activas:
                    if obras_total == 0:
                        respuesta = "Actualmente no tienes obras registradas en el sistema. Te recomiendo crear tu primera obra desde el menú 'Obras' → 'Nueva Obra'."
                    else:
                        obras_otras = Obra.query.filter_by(organizacion_id=org_id).all()
                        estados = [obra.estado for obra in obras_otras]
                        respuesta = f"No tienes obras activas en este momento. Tienes {obras_total} obra(s) registrada(s) con los siguientes estados: {', '.join(set(estados))}."
                else:
                    respuesta = f"Tienes <strong>{len(obras_activas)} obra(s) activa(s)</strong> en construcción:<br><br>"
                    for obra in obras_activas:
                        progreso = obra.progreso or 0
                        respuesta += f"• <strong>{obra.nombre}</strong> - {progreso}% de progreso<br>"
                        respuesta += f"  Cliente: {obra.cliente}<br>"
                        if obra.fecha_fin_estimada:
                            respuesta += f"  Fecha estimada de finalización: {obra.fecha_fin_estimada.strftime('%d/%m/%Y')}<br>"
                        respuesta += "<br>"
                    
                    respuesta += "¿Te gustaría ver detalles específicos de alguna obra?"
        
        elif 'presupuesto' in consulta or 'costo' in consulta or 'precio' in consulta:
            if 'mes' in consulta or 'mensual' in consulta:
                # Consulta sobre presupuestos del mes actual
                from datetime import datetime
                mes_actual = datetime.now().month
                año_actual = datetime.now().year
                
                presupuestos_mes = db.session.query(Presupuesto).join(Obra).filter(
                    Obra.organizacion_id == org_id,
                    Presupuesto.estado == 'aprobado',
                    func.extract('month', Presupuesto.fecha) == mes_actual,
                    func.extract('year', Presupuesto.fecha) == año_actual
                ).all()
                
                if not presupuestos_mes:
                    respuesta = f"No tienes presupuestos aprobados en el mes actual ({datetime.now().strftime('%B %Y')})."
                else:
                    total_mes = sum(float(p.total_con_iva or 0) for p in presupuestos_mes)
                    respuesta = f"<strong>Presupuestos aprobados en {datetime.now().strftime('%B %Y')}:</strong><br><br>"
                    respuesta += f"Total del mes: <strong>${total_mes:,.2f}</strong><br><br>"
                    for presupuesto in presupuestos_mes:
                        respuesta += f"• <strong>{presupuesto.numero}</strong> - ${float(presupuesto.total_con_iva):,.2f}<br>"
                        respuesta += f"  Obra: {presupuesto.obra.nombre}<br>"
                        respuesta += f"  Fecha: {presupuesto.fecha.strftime('%d/%m/%Y')}<br><br>"
                        
            elif 'excediendo' in consulta or 'exceso' in consulta or 'sobrepasando' in consulta:
                # Consulta sobre obras que exceden el presupuesto (simulada)
                obras_con_problemas = []
                obras_todas = Obra.query.filter_by(organizacion_id=org_id).all()
                
                for obra in obras_todas:
                    # Simulamos que una obra con progreso mayor al 80% pero estado no finalizado puede tener problemas
                    if obra.progreso and obra.progreso > 80 and obra.estado != 'finalizada':
                        obras_con_problemas.append(obra)
                
                if not obras_con_problemas:
                    respuesta = "✅ <strong>Buenas noticias:</strong> No hay obras que estén excediendo sus presupuestos actualmente."
                else:
                    respuesta = f"⚠️ <strong>Atención:</strong> Hay {len(obras_con_problemas)} obra(s) que podrían estar excediendo el presupuesto:<br><br>"
                    for obra in obras_con_problemas:
                        respuesta += f"• <strong>{obra.nombre}</strong> - {obra.progreso}% de progreso<br>"
                        respuesta += f"  Estado: {obra.estado}<br>"
                        respuesta += f"  Recomendación: Revisar costos adicionales<br><br>"
            else:
                # Consulta general sobre presupuestos
                presupuestos_aprobados = db.session.query(Presupuesto).join(Obra).filter(
                    Obra.organizacion_id == org_id,
                    Presupuesto.estado == 'aprobado'
                ).all()
                
                if not presupuestos_aprobados:
                    presupuestos_total = db.session.query(Presupuesto).join(Obra).filter(Obra.organizacion_id == org_id).count()
                    if presupuestos_total == 0:
                        respuesta = "No tienes presupuestos registrados en el sistema. Puedes crear presupuestos desde el menú 'Presupuestos' → 'Nuevo Presupuesto'."
                    else:
                        respuesta = f"Tienes {presupuestos_total} presupuesto(s) registrado(s) pero ninguno está aprobado. Puedes revisar y aprobar presupuestos desde el menú 'Presupuestos'."
                else:
                    total_valor = sum(float(p.total_con_iva or 0) for p in presupuestos_aprobados)
                    respuesta = f"Tienes <strong>{len(presupuestos_aprobados)} presupuesto(s) aprobado(s)</strong> por un valor total de <strong>${total_valor:,.2f}</strong>:<br><br>"
                    for presupuesto in presupuestos_aprobados:
                        respuesta += f"• <strong>{presupuesto.numero}</strong> - ${float(presupuesto.total_con_iva):,.2f}<br>"
                        if presupuesto.obra:
                            respuesta += f"  Obra: {presupuesto.obra.nombre}<br>"
                        respuesta += "<br>"
        
        elif 'inventario' in consulta or 'stock' in consulta or 'material' in consulta or 'herramienta' in consulta:
            if 'crítico' in consulta or 'bajo' in consulta or 'problema' in consulta:
                # Consulta específica sobre stock crítico
                if items_stock_bajo:
                    respuesta = f"⚠️ <strong>Alerta de inventario:</strong> Tienes {len(items_stock_bajo)} elemento(s) con stock crítico:<br><br>"
                    for item in items_stock_bajo:
                        respuesta += f"• <strong>{item.nombre}</strong> ({item.categoria})<br>"
                        respuesta += f"  Stock actual: {item.stock_actual} {item.unidad}<br>"
                        respuesta += f"  Stock mínimo: {item.stock_minimo} {item.unidad}<br><br>"
                    respuesta += "Te recomiendo reabastecer estos elementos urgentemente."
                else:
                    respuesta = "✅ <strong>Excelente:</strong> No tienes elementos con stock crítico. Todo tu inventario está en niveles seguros."
                    
            elif 'herramienta' in consulta and 'asignada' in consulta:
                # Consulta sobre herramientas asignadas
                herramientas_asignadas = ItemInventario.query.filter(
                    ItemInventario.organizacion_id == org_id,
                    ItemInventario.categoria == 'herramientas',
                    ItemInventario.stock_actual < ItemInventario.stock_minimo
                ).all()
                
                if not herramientas_asignadas:
                    total_herramientas = ItemInventario.query.filter_by(
                        organizacion_id=org_id,
                        categoria='herramientas'
                    ).count()
                    respuesta = f"Todas tus herramientas están disponibles. Tienes {total_herramientas} herramienta(s) registrada(s) en el inventario."
                else:
                    respuesta = f"<strong>Herramientas con stock bajo (posiblemente asignadas):</strong><br><br>"
                    for herr in herramientas_asignadas:
                        respuesta += f"• <strong>{herr.nombre}</strong> - {herr.stock_actual}/{herr.stock_minimo} disponible(s)<br>"
            else:
                # Consulta general sobre inventario
                if items_stock_bajo:
                    respuesta = f"⚠️ <strong>Resumen de inventario:</strong> Tienes {len(items_stock_bajo)} elemento(s) con stock bajo que requieren atención.<br><br>"
                    respuesta += "Para ver detalles, pregunta específicamente sobre 'stock crítico'."
                else:
                    total_items = ItemInventario.query.filter_by(organizacion_id=org_id).count()
                    if total_items == 0:
                        respuesta = "No tienes elementos en el inventario. Puedes agregar materiales, herramientas y equipos desde el menú 'Inventario'."
                    else:
                        respuesta = f"✅ <strong>Inventario en buen estado:</strong> Tienes {total_items} elemento(s) registrado(s) y todos están en niveles seguros."
        
        elif 'usuario' in consulta or 'equipo' in consulta or 'persona' in consulta or 'operario' in consulta or 'responsable' in consulta:
            if 'operario' in consulta or 'asignado' in consulta:
                # Consulta sobre operarios asignados a obras
                from models import AsignacionObra
                asignaciones = db.session.query(AsignacionObra).join(Obra).filter(
                    Obra.organizacion_id == org_id,
                    Obra.estado == 'en_progreso'
                ).all()
                
                if not asignaciones:
                    respuesta = "No hay operarios asignados a obras activas en este momento."
                else:
                    respuesta = "<strong>Operarios asignados a obras activas:</strong><br><br>"
                    obras_operarios = {}
                    for asignacion in asignaciones:
                        obra_nombre = asignacion.obra.nombre
                        if obra_nombre not in obras_operarios:
                            obras_operarios[obra_nombre] = []
                        obras_operarios[obra_nombre].append(f"{asignacion.usuario.nombre_completo} ({asignacion.rol_en_obra})")
                    
                    for obra, operarios in obras_operarios.items():
                        respuesta += f"• <strong>{obra}</strong>:<br>"
                        for operario in operarios:
                            respuesta += f"  - {operario}<br>"
                        respuesta += "<br>"
                        
            elif 'responsable' in consulta:
                # Consulta sobre responsables de obras
                from models import AsignacionObra
                responsables = db.session.query(AsignacionObra).join(Obra).filter(
                    Obra.organizacion_id == org_id,
                    AsignacionObra.rol_en_obra.in_(['jefe_obra', 'supervisor'])
                ).all()
                
                if not responsables:
                    respuesta = "No hay responsables asignados a las obras."
                else:
                    respuesta = "<strong>Responsables de obras:</strong><br><br>"
                    for resp in responsables:
                        respuesta += f"• <strong>{resp.obra.nombre}</strong>: {resp.usuario.nombre_completo} ({resp.rol_en_obra})<br>"
            else:
                # Consulta general sobre usuarios
                usuarios_activos = Usuario.query.filter_by(
                    organizacion_id=org_id,
                    activo=True
                ).count()
                respuesta = f"Tu organización tiene <strong>{usuarios_activos} usuario(s) activo(s)</strong>. Puedes gestionar el equipo desde el menú 'Equipos'."
        
        else:
            # Respuesta general con resumen
            respuesta = f"""<strong>Resumen de tu organización:</strong><br><br>
            • <strong>Obras activas:</strong> {len(obras_activas)}<br>
            • <strong>Obras totales:</strong> {obras_total}<br>
            • <strong>Presupuestos aprobados:</strong> {len(presupuestos_aprobados)}<br>
            • <strong>Elementos con stock bajo:</strong> {len(items_stock_bajo)}<br><br>
            Puedes hacer preguntas específicas sobre obras, presupuestos, inventario o usuarios."""
        
        # Calcular tiempo de respuesta
        tiempo_respuesta = int((time.time() - start_time) * 1000)
        
        # Completar registro de auditoría
        consulta_registro.respuesta_texto = respuesta
        consulta_registro.estado = 'exito'
        consulta_registro.tiempo_respuesta_ms = tiempo_respuesta
        
        # Metadata adicional
        metadata = {
            'obras_activas': len(obras_activas),
            'total_obras': obras_total,
            'items_stock_bajo': len(items_stock_bajo)
        }
        consulta_registro.set_metadata(metadata)
        
        db.session.add(consulta_registro)
        db.session.commit()
        
        return jsonify({
            'respuesta': respuesta,
            'tiempo_respuesta': tiempo_respuesta,
            'datos_reales': True
        })
    
    except Exception as e:
        # Registrar error
        tiempo_respuesta = int((time.time() - start_time) * 1000)
        error_msg = f'Error al procesar la consulta: {str(e)}'
        
        if consulta_registro:
            consulta_registro.respuesta_texto = error_msg
            consulta_registro.estado = 'error'
            consulta_registro.tiempo_respuesta_ms = tiempo_respuesta
            consulta_registro.error_detalle = str(e)
            db.session.add(consulta_registro)
            db.session.commit()
        
        return jsonify({
            'respuesta': error_msg,
            'tiempo_respuesta': tiempo_respuesta,
            'datos_reales': False
        }), 500


@agent_bp.route('/super-admin/auditoria')
@login_required
def auditoria_consultas():
    """Vista de superadministrador para auditoría de consultas del agente"""
    # Solo permitir acceso a emails específicos de superadministradoras
    emails_superadmin = ['brenda@gmail.com', 'admin@obyra.com']  # Agregar tu email aquí
    
    if current_user.email not in emails_superadmin:
        return redirect(url_for('reportes.dashboard'))
    
    # Filtros de la consulta
    filtro_org = request.args.get('organizacion', '')
    filtro_tipo = request.args.get('tipo', '')
    filtro_estado = request.args.get('estado', '')
    filtro_fecha_desde = request.args.get('fecha_desde', '')
    filtro_fecha_hasta = request.args.get('fecha_hasta', '')
    
    # Construir consulta base
    consultas_query = db.session.query(ConsultaAgente).join(Usuario).join(ConsultaAgente.organizacion)
    
    # Aplicar filtros
    if filtro_org:
        consultas_query = consultas_query.filter(ConsultaAgente.organizacion_id == filtro_org)
    if filtro_tipo:
        consultas_query = consultas_query.filter(ConsultaAgente.tipo_consulta == filtro_tipo)
    if filtro_estado:
        consultas_query = consultas_query.filter(ConsultaAgente.estado == filtro_estado)
    if filtro_fecha_desde:
        consultas_query = consultas_query.filter(ConsultaAgente.fecha_consulta >= datetime.strptime(filtro_fecha_desde, '%Y-%m-%d'))
    if filtro_fecha_hasta:
        consultas_query = consultas_query.filter(ConsultaAgente.fecha_consulta <= datetime.strptime(filtro_fecha_hasta + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    

    
    # Estadísticas generales
    total_consultas = db.session.query(ConsultaAgente).count()
    consultas_exitosas = db.session.query(ConsultaAgente).filter_by(estado='exito').count()
    consultas_error = db.session.query(ConsultaAgente).filter_by(estado='error').count()
    
    # Top consultas por tipo
    tipos_consulta = db.session.query(
        ConsultaAgente.tipo_consulta,
        func.count(ConsultaAgente.id).label('total')
    ).group_by(ConsultaAgente.tipo_consulta).all()
    
    # Top organizaciones que más consultan
    top_organizaciones = db.session.query(
        ConsultaAgente.organizacion_id,
        func.count(ConsultaAgente.id).label('total'),
        ConsultaAgente.organizacion
    ).join(ConsultaAgente.organizacion).group_by(
        ConsultaAgente.organizacion_id, Organizacion.id
    ).order_by(desc(func.count(ConsultaAgente.id))).limit(10).all()
    
    # Tiempo promedio de respuesta
    tiempo_promedio = db.session.query(
        func.avg(ConsultaAgente.tiempo_respuesta_ms)
    ).filter_by(estado='exito').scalar() or 0
    
    # Consultas por fecha (últimos 7 días)
    hace_7_dias = datetime.now() - timedelta(days=7)
    consultas_por_dia_query = db.session.query(
        func.date(ConsultaAgente.fecha_consulta).label('fecha'),
        func.count(ConsultaAgente.id).label('total')
    ).filter(ConsultaAgente.fecha_consulta >= hace_7_dias).group_by(
        func.date(ConsultaAgente.fecha_consulta)
    ).order_by(func.date(ConsultaAgente.fecha_consulta)).all()
    
    # Convertir a formato adecuado para el template
    consultas_por_dia = []
    for fecha_item, total in consultas_por_dia_query:
        # Asegurar que tenemos un objeto de fecha válido
        try:
            if isinstance(fecha_item, str):
                fecha_obj = datetime.strptime(fecha_item, '%Y-%m-%d').date()
            elif hasattr(fecha_item, 'strftime'):
                fecha_obj = fecha_item
            else:
                fecha_obj = datetime.now().date()
        except (ValueError, TypeError, AttributeError) as e:
            current_app.logger.warning(f"Error al convertir fecha en consultas_por_dia: {e}, usando fecha actual")
            fecha_obj = datetime.now().date()

        consultas_por_dia.append((fecha_obj, total))
    
    # Construir consulta base para auditoría
    query = db.session.query(ConsultaAgente).join(Usuario).join(ConsultaAgente.organizacion)
    
    # Aplicar filtros
    if filtro_org:
        query = query.filter(ConsultaAgente.organizacion_id == filtro_org)
    if filtro_tipo:
        query = query.filter(ConsultaAgente.tipo_consulta == filtro_tipo)
    if filtro_estado:
        query = query.filter(ConsultaAgente.estado == filtro_estado)
    if filtro_fecha_desde:
        query = query.filter(ConsultaAgente.fecha_consulta >= datetime.strptime(filtro_fecha_desde, '%Y-%m-%d'))
    if filtro_fecha_hasta:
        query = query.filter(ConsultaAgente.fecha_consulta <= datetime.strptime(filtro_fecha_hasta + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    
    # Si es exportación, obtener todos los datos
    if request.args.get('export'):
        export_format = request.args.get('export')
        consultas_export = query.order_by(desc(ConsultaAgente.fecha_consulta)).all()
        if export_format == 'excel':
            return exportar_excel(consultas_export)
        elif export_format == 'pdf':
            return exportar_pdf(consultas_export)
    
    # Paginación para vista normal
    page = request.args.get('page', 1, type=int)
    per_page = 20
    consultas = query.order_by(desc(ConsultaAgente.fecha_consulta)).paginate(
        page=page, per_page=per_page, error_out=False)
    
    # Obtener todas las organizaciones para el filtro
    organizaciones = Organizacion.query.all()
    
    return render_template('asistente/auditoria_consultas.html',
                         consultas=consultas,
                         organizaciones=organizaciones,
                         estadisticas={
                             'total_consultas': total_consultas,
                             'consultas_exitosas': consultas_exitosas,
                             'consultas_error': consultas_error,
                             'tasa_exito': round((consultas_exitosas / total_consultas * 100) if total_consultas > 0 else 0, 2),
                             'tiempo_promedio': round(tiempo_promedio, 2),
                             'tipos_consulta': tipos_consulta,
                             'top_organizaciones': top_organizaciones,
                             'consultas_por_dia': consultas_por_dia
                         },
                         filtros={
                             'organizacion': filtro_org,
                             'tipo': filtro_tipo,
                             'estado': filtro_estado,
                             'fecha_desde': filtro_fecha_desde,
                             'fecha_hasta': filtro_fecha_hasta
                         })


@agent_bp.route('/super-admin/consulta/<int:consulta_id>')
@login_required
def detalle_consulta(consulta_id):
    """Vista detallada de una consulta específica"""
    emails_superadmin = ['brenda@gmail.com', 'admin@obyra.com']
    
    if current_user.email not in emails_superadmin:
        return redirect(url_for('reportes.dashboard'))
    
    consulta = ConsultaAgente.query.get_or_404(consulta_id)
    
    return render_template('asistente/detalle_consulta.html', consulta=consulta)


def exportar_excel(consultas):
    """Exportar consultas a Excel"""
    if not OPENPYXL_AVAILABLE:
        flash(
            "La exportación a Excel requiere la librería openpyxl. Ejecuta 'pip install openpyxl' para habilitarla.",
            'warning'
        )
        destino = request.referrer or url_for('agent_local.auditoria_consultas')
        return redirect(destino)

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoría Consultas IA"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Encabezados
    headers = [
        'Fecha/Hora', 'Usuario', 'Email', 'Organización', 'Consulta', 
        'Respuesta', 'Tipo', 'Estado', 'Tiempo (ms)', 'IP'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    for row, consulta in enumerate(consultas, 2):
        ws.cell(row=row, column=1, value=consulta.fecha_consulta.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=2, value=consulta.usuario.nombre_completo)
        ws.cell(row=row, column=3, value=consulta.usuario.email)
        ws.cell(row=row, column=4, value=consulta.organizacion.nombre)
        ws.cell(row=row, column=5, value=consulta.consulta_texto)
        ws.cell(row=row, column=6, value=consulta.respuesta_texto)
        ws.cell(row=row, column=7, value=consulta.tipo_consulta)
        ws.cell(row=row, column=8, value=consulta.estado)
        ws.cell(row=row, column=9, value=consulta.tiempo_respuesta_ms)
        ws.cell(row=row, column=10, value=consulta.ip_address)
    
    # Ajustar anchos de columna
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (AttributeError, TypeError):
                # Celda sin valor o error al convertir a string
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=auditoria_consultas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    
    return response


def exportar_pdf(consultas):
    """Exportar consultas a PDF"""
    if not REPORTLAB_AVAILABLE:
        return jsonify({
            'error': "La exportación a PDF requiere la librería reportlab. Instálala con 'pip install reportlab'."
        }), 500

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centrado
    )
    
    story = []
    
    # Título
    story.append(Paragraph("Reporte de Auditoría - Agente IA", title_style))
    story.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # Estadísticas generales
    total_consultas = len(consultas)
    consultas_exitosas = len([c for c in consultas if c.estado == 'exito'])
    
    stats_data = [
        ['Estadística', 'Valor'],
        ['Total de Consultas', str(total_consultas)],
        ['Consultas Exitosas', str(consultas_exitosas)],
        ['Tasa de Éxito', f"{(consultas_exitosas/total_consultas*100):.1f}%" if total_consultas > 0 else "0%"]
    ]
    
    stats_table = Table(stats_data)
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(stats_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Tabla de consultas (últimas 50)
    story.append(Paragraph("Consultas Recientes (Últimas 50)", styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))
    
    data = [['Fecha', 'Usuario', 'Organización', 'Tipo', 'Estado']]
    
    for consulta in consultas[:50]:
        data.append([
            consulta.fecha_consulta.strftime('%d/%m %H:%M'),
            consulta.usuario.nombre_completo[:20],
            consulta.organizacion.nombre[:15],
            consulta.tipo_consulta.title(),
            consulta.estado.title()
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=auditoria_consultas_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
    
    return response
