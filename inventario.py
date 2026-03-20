import os

from flask import (
    Blueprint,
    render_template,
    request,
    flash,
    redirect,
    url_for,
    current_app,
    jsonify,
)
from flask_login import login_required, current_user
from datetime import date
from collections import defaultdict
from typing import Dict, List, Optional
from jinja2 import TemplateNotFound

from app import db
from extensions import csrf
from models import (
    ItemInventario,
    CategoriaInventario,
    MovimientoInventario,
    UsoInventario,
    Obra,
    # Nuevo sistema de ubicaciones
    Location,
    StockUbicacion,
    MovimientoStock,
)
from services.memberships import get_current_org_id

# from inventario_new import nuevo_item as nuevo_item_view  # Commented out - causes import error
from models import InventoryCategory, Organizacion
from seed_inventory_categories import seed_inventory_categories_for_company
from inventory_category_service import (
    ensure_categories_for_company,
    ensure_categories_for_company_id,
    serialize_category,
    render_category_catalog,
    user_can_manage_inventory_categories,
)


def _resolve_company_id() -> Optional[int]:
    org_id = get_current_org_id()
    if org_id:
        return org_id
    return getattr(current_user, 'organizacion_id', None)


def _resolve_company(company_id: int) -> Optional[Organizacion]:
    if getattr(current_user, 'organizacion', None) and current_user.organizacion.id == company_id:
        return current_user.organizacion
    return Organizacion.query.get(company_id)


def _build_category_tree(categorias: List[InventoryCategory]) -> List[Dict[str, object]]:
    children_map: Dict[Optional[int], List[InventoryCategory]] = defaultdict(list)
    for categoria in categorias:
        children_map[categoria.parent_id].append(categoria)

    for bucket in children_map.values():
        bucket.sort(key=lambda cat: ((cat.sort_order or 0), cat.nombre.lower()))

    def build(parent_id: Optional[int] = None) -> List[Dict[str, object]]:
        nodes: List[Dict[str, object]] = []
        for categoria in children_map.get(parent_id, []):
            nodes.append({
                'categoria': categoria,
                'children': build(categoria.id),
            })
        return nodes

    return build()

INVENTARIO_TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), 'templates')

inventario_bp = Blueprint('inventario', __name__, template_folder=INVENTARIO_TEMPLATE_DIR)

# ===== LISTA CANÓNICA DE CATEGORÍAS (orden lógico de obra) =====
CATEGORIAS_CANONICAS = [
    'Preliminares y Obrador',
    'Demoliciones',
    'Movimiento de Suelos',
    'Excavación',
    'Depresión de Napa / Bombeo',
    'Fundaciones',
    'Estructura',
    'Encofrados',
    'Mampostería',
    'Contrapisos y Carpetas',
    'Impermeabilizaciones y Aislaciones',
    'Revoque Grueso',
    'Revoque Fino',
    'Cielorrasos',
    'Yesería y Enlucidos',
    'Pisos y Revestimientos',
    'Carpintería y Aberturas',
    'Herrería de Obra',
    'Pintura',
    'Instalaciones Eléctricas',
    'Instalaciones Sanitarias y Provisiones',
    'Instalaciones de Gas',
    'Ventilaciones y Conductos',
    'Instalaciones Complementarias',
    'Construcción en Seco',
    'Techos y Cubiertas',
    'Limpieza Final y Puesta en Marcha',
    # Soporte
    'Mano de Obra',
    'Maquinarias y Equipos',
    'Materiales de Obra',
    'Seguridad e Higiene',
    'Consumibles e Insumos',
    'Logística y Depósito',
    'Otros',
]

# Mapeo de nombres variantes → categoría canónica destino
_FUSIONES = {
    'carpinteria + metalicas + aberturas': 'Carpintería y Aberturas',
    'carpintería y aberturas': 'Carpintería y Aberturas',
    'estructura': 'Estructura',
    'excavación': 'Excavación',
    'excavacion y movimiento suelo': 'Movimiento de Suelos',
    'movimiento de suelos': 'Movimiento de Suelos',
    'fundaciones': 'Fundaciones',
    'herreria de obra': 'Herrería de Obra',
    'herrería de obra': 'Herrería de Obra',
    'impermeabilizacion y aislacion': 'Impermeabilizaciones y Aislaciones',
    'impermeabilizaciones y aislaciones': 'Impermeabilizaciones y Aislaciones',
    'instalaciones electricas': 'Instalaciones Eléctricas',
    'instalaciones eléctricas': 'Instalaciones Eléctricas',
    'instalaciones sanitarias': 'Instalaciones Sanitarias y Provisiones',
    'instalaciones sanitarias y provisiones': 'Instalaciones Sanitarias y Provisiones',
    'instalaciones de gas': 'Instalaciones de Gas',
    'instalaciones': 'Instalaciones Complementarias',
    'instalaciones climatizacion': 'Instalaciones Complementarias',
    'instalaciones complementarias': 'Instalaciones Complementarias',
    'equipo contra incendios + maquinaria edificio': 'Instalaciones Complementarias',
    'limpieza final': 'Limpieza Final y Puesta en Marcha',
    'limpieza final y puesta en marcha': 'Limpieza Final y Puesta en Marcha',
    'mamposteria': 'Mampostería',
    'mampostería': 'Mampostería',
    'pinturas y revestimientos': 'Pintura',
    'pintura': 'Pintura',
    'pisos': 'Pisos y Revestimientos',
    'pisos y revestimientos': 'Pisos y Revestimientos',
    'revoque fino/yeseria': 'Revoque Fino',
    'revoque fino': 'Revoque Fino',
    'revoque grueso': 'Revoque Grueso',
    'yesería y enlucidos': 'Yesería y Enlucidos',
    'techos': 'Techos y Cubiertas',
    'techos y cubiertas': 'Techos y Cubiertas',
    'maquinarias': 'Maquinarias y Equipos',
    'maquinarias y equipos': 'Maquinarias y Equipos',
    'encofrados': 'Encofrados',
    'sistemas de encofrado y andamiaje': 'Encofrados',
    'apuntalamientos': 'Encofrados',
    'seguridad': 'Seguridad e Higiene',
    'seguridad e higiene': 'Seguridad e Higiene',
    'indumentaria': 'Seguridad e Higiene',
    'material de construcción': 'Materiales de Obra',
    'materiales de obra': 'Materiales de Obra',
    'consumibles e insumos': 'Consumibles e Insumos',
    'administrativo y oficina de obra': 'Otros',
    'contrapisos y carpetas': 'Contrapisos y Carpetas',
    'cielorrasos': 'Cielorrasos',
    'construcción en seco': 'Construcción en Seco',
    'construccion en seco': 'Construcción en Seco',
    'depresión de napa / bombeo': 'Depresión de Napa / Bombeo',
    'depresion de napa / bombeo': 'Depresión de Napa / Bombeo',
    'preliminares y obrador': 'Preliminares y Obrador',
    'demoliciones': 'Demoliciones',
    'ventilaciones y conductos': 'Ventilaciones y Conductos',
    'logística y depósito': 'Logística y Depósito',
    'logistica y deposito': 'Logística y Depósito',
    'mano de obra': 'Mano de Obra',
    'otros': 'Otros',
}


def _ensure_canonical_categories(org_id):
    """Auto-init: asegurar que existen las 34 categorías canónicas.
    Fusiona duplicados y elimina categorías vacías no canónicas.
    Solo hace cambios si es necesario."""
    if not org_id:
        return

    # Check rápido: si ya hay exactamente 34 categorías con los nombres correctos, no hacer nada
    existing = InventoryCategory.query.filter(
        InventoryCategory.company_id == org_id,
        InventoryCategory.is_active == True
    ).all()
    existing_names = {c.nombre for c in existing}
    canonical_set = set(CATEGORIAS_CANONICAS)

    if existing_names == canonical_set and len(existing) == len(CATEGORIAS_CANONICAS):
        # Solo actualizar sort_order si hace falta
        needs_sort = False
        for cat in existing:
            expected = CATEGORIAS_CANONICAS.index(cat.nombre) if cat.nombre in CATEGORIAS_CANONICAS else None
            if expected is not None and cat.sort_order != expected:
                cat.sort_order = expected
                needs_sort = True
        if needs_sort:
            db.session.commit()
        return

    # Paso 1: crear categorías canónicas que no existen
    canonicas_db = {}
    for idx, nombre in enumerate(CATEGORIAS_CANONICAS):
        cat = InventoryCategory.query.filter(
            InventoryCategory.company_id == org_id,
            InventoryCategory.nombre == nombre
        ).first()
        if not cat:
            cat = InventoryCategory(
                company_id=org_id,
                nombre=nombre,
                sort_order=idx,
                is_active=True
            )
            db.session.add(cat)
            db.session.flush()
        else:
            cat.sort_order = idx
        canonicas_db[nombre] = cat

    # Paso 2: fusionar categorías no canónicas
    todas = InventoryCategory.query.filter(
        InventoryCategory.company_id == org_id
    ).all()

    for cat in todas:
        if cat.nombre in canonical_set:
            continue

        key = cat.nombre.strip().lower()
        destino_nombre = _FUSIONES.get(key)

        if destino_nombre and destino_nombre in canonicas_db:
            destino = canonicas_db[destino_nombre]
            if destino.id != cat.id:
                # Mover items
                ItemInventario.query.filter_by(categoria_id=cat.id).update(
                    {'categoria_id': destino.id}, synchronize_session='fetch')
                # Mover hijos
                InventoryCategory.query.filter_by(parent_id=cat.id).update(
                    {'parent_id': destino.id}, synchronize_session='fetch')
                db.session.delete(cat)
        else:
            # Sin mapeo: si tiene 0 items, eliminar
            count = ItemInventario.query.filter_by(categoria_id=cat.id).count()
            if count == 0:
                db.session.delete(cat)
            # Si tiene items pero no tiene mapeo, mover a "Otros"
            elif 'Otros' in canonicas_db:
                ItemInventario.query.filter_by(categoria_id=cat.id).update(
                    {'categoria_id': canonicas_db['Otros'].id}, synchronize_session='fetch')
                db.session.delete(cat)

    db.session.commit()


@inventario_bp.route('/')
@login_required
def lista():
    current_app.logger.info(f"[INVENTARIO] Accediendo a lista. User: {current_user.email}")
    if not current_user.puede_acceder_modulo('inventario'):
        current_app.logger.info(f"[INVENTARIO] Usuario sin permisos: {current_user.email}")
        flash('No tienes permisos para acceder a este módulo.', 'danger')
        return redirect(url_for('reportes.dashboard'))

    categoria_id = request.args.get('categoria', '', type=int) or None
    buscar = request.args.get('buscar', '')
    stock_bajo = request.args.get('stock_bajo', '')
    con_stock = request.args.get('con_stock', '')

    # Obtener org_id del usuario actual
    org_id = get_current_org_id() or current_user.organizacion_id

    # Auto-init: asegurar categorías canónicas
    _ensure_canonical_categories(org_id)

    # Debug log
    current_app.logger.info(f"[INVENTARIO] org_id={org_id}, user.organizacion_id={current_user.organizacion_id}")

    # Query base - usar outerjoin para incluir items aunque la categoría no exista
    query = ItemInventario.query.outerjoin(ItemInventario.categoria)

    # Filtrar por organización del usuario
    if org_id:
        query = query.filter(ItemInventario.organizacion_id == org_id)
        current_app.logger.info(f"[INVENTARIO] Filtrando por org_id={org_id}")

    if categoria_id:
        from models.inventory import item_categorias_adicionales
        query = query.filter(
            db.or_(
                ItemInventario.categoria_id == categoria_id,
                ItemInventario.id.in_(
                    db.session.query(item_categorias_adicionales.c.item_id).filter(
                        item_categorias_adicionales.c.categoria_id == categoria_id
                    )
                )
            )
        )

    # Improved search with case-insensitive partial matching
    if buscar:
        buscar_pattern = f'%{buscar}%'
        query = query.filter(
            db.or_(
                ItemInventario.codigo.ilike(buscar_pattern),
                ItemInventario.nombre.ilike(buscar_pattern),
                ItemInventario.descripcion.ilike(buscar_pattern)
            )
        )

    if con_stock:
        # Filtrar solo artículos con stock mayor a 0
        query = query.filter(ItemInventario.stock_actual > 0)

    if stock_bajo:
        query = query.filter(ItemInventario.stock_actual <= ItemInventario.stock_minimo)

    # Paginación para mejor rendimiento en móviles
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)  # 25 items por página por defecto (optimizado)
    per_page = min(per_page, 100)  # Máximo 100 por página

    # OPTIMIZACIÓN: Detectar si estamos en vista de carpetas (sin filtros)
    vista_carpetas = not buscar and not categoria_id and not con_stock and not stock_bajo

    from models.projects import Obra
    from models.budgets import Presupuesto
    from models.inventory import StockObra

    items_con_obras = []
    pagination = None
    total_items = 0

    # Solo cargar items si hay filtros activos (no en vista de carpetas)
    if not vista_carpetas:
        base_query = query.filter(ItemInventario.activo == True).order_by(ItemInventario.nombre)
        total_items = base_query.count()

        # Paginación
        pagination = base_query.paginate(page=page, per_page=per_page, error_out=False)
        items = pagination.items

        current_app.logger.info(f"[INVENTARIO] Items en página {page}: {len(items)} de {total_items} total")

        for item in items:
            # Versión simplificada - no hacer queries pesadas para cada item
            items_con_obras.append({
                'item': item,
                'obras': [],
                'reservas': [],
                'stock_reservado': 0,
                'stock_fisico_inicial': float(item.stock_actual or 0),
                'traslados_obras': [],
                'total_trasladado': 0,
                'ubicaciones_stock': []
            })
    else:
        current_app.logger.info(f"[INVENTARIO] Vista de carpetas - no cargando items")

    # Load new inventory categories (propias de la org + globales)
    org_id = get_current_org_id() or current_user.organizacion_id
    categorias_nuevas = []
    if org_id:
        categorias_nuevas = InventoryCategory.query.filter(
            db.or_(
                InventoryCategory.company_id == org_id,
                InventoryCategory.is_global == True
            ),
            InventoryCategory.is_active == True
        ).order_by(InventoryCategory.sort_order, InventoryCategory.nombre).all()

    # CategoriaInventario está DEPRECATED - solo usar InventoryCategory
    categorias = []

    # Get all obras for dropdowns (solo si hay filtros activos)
    obras_disponibles = []
    if not vista_carpetas:
        obras_disponibles = Obra.query.join(Presupuesto).filter(
            Obra.organizacion_id == org_id,
            db.or_(
                Presupuesto.confirmado_como_obra == True,
                Presupuesto.estado.in_(['aprobado', 'convertido', 'confirmado'])
            )
        ).distinct().all()

    # Construir árbol de categorías con items agrupados (para vista de carpetas)
    # Conteo por categoría principal
    conteo_items = {}
    if org_id:
        conteo_query = db.session.query(
            ItemInventario.categoria_id,
            db.func.count(ItemInventario.id)
        ).filter(
            ItemInventario.organizacion_id == org_id,
            ItemInventario.activo == True
        ).group_by(ItemInventario.categoria_id).all()
        conteo_items = {cat_id: count for cat_id, count in conteo_query}

    # Conteo por categorías adicionales (many-to-many)
    conteo_adicionales = {}
    if org_id:
        from models.inventory import item_categorias_adicionales
        conteo_adic_query = db.session.query(
            item_categorias_adicionales.c.categoria_id,
            db.func.count(item_categorias_adicionales.c.item_id)
        ).join(
            ItemInventario, ItemInventario.id == item_categorias_adicionales.c.item_id
        ).filter(
            ItemInventario.organizacion_id == org_id,
            ItemInventario.activo == True
        ).group_by(item_categorias_adicionales.c.categoria_id).all()
        conteo_adicionales = {cat_id: count for cat_id, count in conteo_adic_query}

    arbol_categorias = []
    for cat_principal in categorias_nuevas:
        total_items_categoria = conteo_items.get(cat_principal.id, 0) + conteo_adicionales.get(cat_principal.id, 0)

        arbol_categorias.append({
            'id': cat_principal.id,
            'nombre': cat_principal.nombre,
            'subcategorias': [],
            'total_items': total_items_categoria,
            'tiene_items_directos': total_items_categoria > 0
        })

    return render_template('inventario/lista.html',
                         items_con_obras=items_con_obras,
                         categorias=categorias,
                         categorias_nuevas=categorias_nuevas,
                         arbol_categorias=arbol_categorias,
                         categoria_id=categoria_id,
                         buscar=buscar,
                         stock_bajo=stock_bajo,
                         con_stock=con_stock,
                         obras_disponibles=obras_disponibles,
                         pagination=pagination,
                         total_items=total_items,
                         page=page,
                         per_page=per_page)

@inventario_bp.route('/crear', methods=['GET', 'POST'])
@login_required
def crear():
    if not current_user.puede_acceder_modulo('inventario'):
        flash('No tienes permisos para crear items de inventario.', 'danger')
        return redirect(url_for('inventario.lista'))

    org_id = get_current_org_id() or current_user.organizacion_id

    if not org_id:
        flash('No tienes una organización activa', 'warning')
        return redirect(url_for('index'))

    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            categoria_id = request.form.get('categoria_id')
            codigo = request.form.get('codigo', '').strip().upper()
            nombre = request.form.get('nombre', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            unidad = request.form.get('unidad', '').strip()
            stock_actual = request.form.get('stock_actual', 0)
            stock_minimo = request.form.get('stock_minimo', 0)
            precio_promedio = request.form.get('precio_promedio', 0)
            precio_promedio_usd = request.form.get('precio_promedio_usd', 0)

            # Validaciones
            if not all([categoria_id, codigo, nombre, unidad]):
                flash('Categoría, código, nombre y unidad son campos obligatorios.', 'danger')
                categorias = InventoryCategory.query.filter_by(company_id=org_id, is_active=True).all()
                return render_template('inventario/crear.html', categorias=categorias)

            # Verificar si el código ya existe
            existing = ItemInventario.query.filter_by(codigo=codigo).first()
            if existing:
                flash(f'Ya existe un item con el código {codigo}', 'warning')
                categorias = InventoryCategory.query.filter_by(company_id=org_id, is_active=True).all()
                return render_template('inventario/crear.html', categorias=categorias)

            # Verificar si ya existe un item con el mismo nombre + descripción en la organización
            # Permite: Cemento (Avellaneda 50kg) y Cemento (Avellaneda 25kg)
            # Bloquea: Dos items con exactamente el mismo nombre y descripción
            descripcion_normalizada = (descripcion or '').strip().lower()

            query_duplicado = ItemInventario.query.filter(
                ItemInventario.organizacion_id == org_id,
                ItemInventario.nombre.ilike(nombre),
                ItemInventario.activo == True
            )

            # Si hay descripción, buscar coincidencia exacta de nombre + descripción
            # Si no hay descripción, buscar items sin descripción con el mismo nombre
            if descripcion_normalizada:
                query_duplicado = query_duplicado.filter(
                    db.func.lower(db.func.coalesce(ItemInventario.descripcion, '')).like(descripcion_normalizada)
                )
            else:
                query_duplicado = query_duplicado.filter(
                    db.or_(
                        ItemInventario.descripcion.is_(None),
                        ItemInventario.descripcion == ''
                    )
                )

            existing_duplicado = query_duplicado.first()
            if existing_duplicado:
                if descripcion:
                    flash(f'Ya existe "{nombre}" con descripción "{descripcion}" (código: {existing_duplicado.codigo}). '
                          f'Usa una descripción diferente para distinguirlo.', 'warning')
                else:
                    flash(f'Ya existe "{nombre}" sin descripción (código: {existing_duplicado.codigo}). '
                          f'Agrega una descripción (ej: marca, presentación) para crear un nuevo item.', 'warning')
                categorias = InventoryCategory.query.filter_by(company_id=org_id, is_active=True).all()
                return render_template('inventario/crear.html', categorias=categorias)

            # Convertir valores numéricos
            stock_actual = float(stock_actual)
            stock_minimo = float(stock_minimo)
            precio_promedio = float(precio_promedio)
            precio_promedio_usd = float(precio_promedio_usd)

            # Crear el item
            nuevo_item = ItemInventario(
                organizacion_id=org_id,
                categoria_id=categoria_id,
                codigo=codigo,
                nombre=nombre,
                descripcion=descripcion or None,
                unidad=unidad,
                stock_actual=stock_actual,
                stock_minimo=stock_minimo,
                precio_promedio=precio_promedio,
                precio_promedio_usd=precio_promedio_usd,
                activo=True
            )

            db.session.add(nuevo_item)
            db.session.flush()  # Para obtener el ID

            # Si hay stock inicial, crear movimiento de entrada
            if stock_actual > 0:
                movimiento = MovimientoInventario(
                    item_id=nuevo_item.id,
                    tipo='entrada',
                    cantidad=stock_actual,
                    precio_unitario=precio_promedio,
                    motivo='Inventario inicial',
                    observaciones='Stock inicial al crear el item',
                    usuario_id=current_user.id
                )
                db.session.add(movimiento)

            db.session.commit()

            flash(f'Item {nuevo_item.nombre} creado exitosamente', 'success')
            return redirect(url_for('inventario.lista'))

        except ValueError as e:
            db.session.rollback()
            flash('Error: Los valores numéricos no son válidos', 'danger')
            categorias = InventoryCategory.query.filter_by(company_id=org_id, is_active=True).all()
            return render_template('inventario/crear.html', categorias=categorias)
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error al crear item de inventario: {str(e)}")
            flash('Error al crear el item de inventario', 'danger')
            categorias = InventoryCategory.query.filter_by(company_id=org_id, is_active=True).all()
            return render_template('inventario/crear.html', categorias=categorias)

    # GET request
    categorias = InventoryCategory.query.filter_by(company_id=org_id, is_active=True).order_by(InventoryCategory.nombre).all()
    can_manage_categories = user_can_manage_inventory_categories(current_user)

    return render_template('inventario/crear.html',
                         categorias=categorias,
                         can_manage_categories=can_manage_categories)

@inventario_bp.route('/<int:id>')
@login_required
def detalle(id):
    if not current_user.puede_acceder_modulo('inventario'):
        flash('No tienes permisos para ver detalles de inventario.', 'danger')
        return redirect(url_for('reportes.dashboard'))

    item = ItemInventario.query.get_or_404(id)
    org_id = get_current_org_id() or current_user.organizacion_id

    # Verificar que el item pertenece a la organización (excepto superadmin)
    if not current_user.is_super_admin and item.organizacion_id != org_id:
        flash('No tienes permisos para ver este item.', 'danger')
        return redirect(url_for('inventario.lista'))

    # Obtener últimos movimientos
    movimientos = item.movimientos.order_by(MovimientoInventario.fecha.desc()).limit(10).all()
    
    # Obtener uso en obras
    usos_obra = item.usos.join(Obra).order_by(UsoInventario.fecha_uso.desc()).limit(10).all()
    
    # Categorías para el selector de edición (orden lógico de obra)
    categorias = InventoryCategory.query.filter(
        InventoryCategory.company_id == org_id,
        InventoryCategory.is_active == True
    ).order_by(InventoryCategory.sort_order, InventoryCategory.nombre).all()

    return render_template('inventario/detalle.html',
                         item=item,
                         movimientos=movimientos,
                         usos_obra=usos_obra,
                         categorias=categorias)


@inventario_bp.route('/<int:id>/editar', methods=['POST'])
@login_required
def editar_item(id):
    """Editar categoría y datos básicos de un item."""
    item = ItemInventario.query.get_or_404(id)
    org_id = get_current_org_id() or current_user.organizacion_id

    if not current_user.is_super_admin and item.organizacion_id != org_id:
        flash('No tienes permisos para editar este item.', 'danger')
        return redirect(url_for('inventario.lista'))

    categoria_id = request.form.get('categoria_id')
    nombre = request.form.get('nombre')
    unidad = request.form.get('unidad')
    stock_minimo = request.form.get('stock_minimo')

    if categoria_id:
        item.categoria_id = int(categoria_id) if categoria_id != '0' else None
    if nombre:
        item.nombre = nombre.strip()
    if unidad:
        item.unidad_medida = unidad.strip()
    if stock_minimo is not None and stock_minimo != '':
        item.stock_minimo = float(stock_minimo)

    db.session.commit()
    flash(f'Item {item.codigo} actualizado', 'success')
    return redirect(url_for('inventario.detalle', id=id))


@inventario_bp.route('/<int:id>/categorias-adicionales', methods=['POST'])
@csrf.exempt
@login_required
def categorias_adicionales(id):
    """Gestionar categorías adicionales de un item (many-to-many)."""
    item = ItemInventario.query.get_or_404(id)
    org_id = get_current_org_id() or current_user.organizacion_id

    if current_user.rol not in ('administrador', 'tecnico'):
        return jsonify(ok=False, error='Sin permisos'), 403

    data = request.get_json(silent=True)
    if not data:
        return jsonify(ok=False, error='Datos inválidos'), 400

    action = data.get('action')  # 'add' o 'remove' o 'set'
    categoria_ids = data.get('categoria_ids', [])

    if action == 'set':
        # Reemplazar todas las categorías adicionales
        cats = InventoryCategory.query.filter(
            InventoryCategory.id.in_(categoria_ids),
            InventoryCategory.company_id == org_id
        ).all()
        # Excluir la categoría principal
        item.categorias_adicionales = [c for c in cats if c.id != item.categoria_id]
    elif action == 'add':
        for cat_id in categoria_ids:
            if cat_id == item.categoria_id:
                continue
            cat = InventoryCategory.query.get(cat_id)
            if cat and cat not in item.categorias_adicionales:
                item.categorias_adicionales.append(cat)
    elif action == 'remove':
        for cat_id in categoria_ids:
            cat = InventoryCategory.query.get(cat_id)
            if cat and cat in item.categorias_adicionales:
                item.categorias_adicionales.remove(cat)
    else:
        return jsonify(ok=False, error='Acción inválida (add/remove/set)'), 400

    db.session.commit()
    return jsonify(
        ok=True,
        categorias=[{'id': c.id, 'nombre': c.nombre} for c in item.categorias_adicionales]
    )


@inventario_bp.route('/reclasificar-encofrados', methods=['POST'])
@login_required
def reclasificar_encofrados():
    """Reclasificar items de encofrado que están mal categorizados."""
    org_id = get_current_org_id() or current_user.organizacion_id

    # Buscar categoría ENCOFRADOS
    cat_encofrados = InventoryCategory.query.filter(
        InventoryCategory.company_id == org_id,
        InventoryCategory.nombre.ilike('%encofrado%')
    ).first()

    if not cat_encofrados:
        flash('No se encontró la categoría de Encofrados', 'warning')
        return redirect(url_for('inventario.lista'))

    # Keywords que identifican items de encofrado
    keywords_encofrado = ['viga h20', 'puntal', 'cabezal', 'tripode', 'trípode',
                          'fork', 'gato', 'mensula', 'ménsula', 'tensor',
                          'panel encofrado', 'tablero encofrado', 'placa encofrado']

    from sqlalchemy import or_
    filtros = [ItemInventario.nombre.ilike(f'%{kw}%') for kw in keywords_encofrado]

    items = ItemInventario.query.filter(
        ItemInventario.organizacion_id == org_id,
        or_(*filtros)
    ).all()

    count = 0
    for item in items:
        if item.categoria_id != cat_encofrados.id:
            item.categoria_id = cat_encofrados.id
            count += 1

    db.session.commit()
    flash(f'{count} items reclasificados a Encofrados', 'success')
    return redirect(url_for('inventario.lista'))


@inventario_bp.route('/mover-categoria-bulk', methods=['POST'])
@csrf.exempt
@login_required
def mover_categoria_bulk():
    """Mover múltiples items a otra categoría."""
    if current_user.rol != 'administrador':
        return jsonify(ok=False, error='Sin permisos'), 403

    data = request.get_json(silent=True)
    if not data:
        return jsonify(ok=False, error='Datos inválidos'), 400

    item_ids = data.get('item_ids', [])
    categoria_id = data.get('categoria_id')
    if not item_ids or not categoria_id:
        return jsonify(ok=False, error='Faltan datos'), 400

    org_id = get_current_org_id() or current_user.organizacion_id
    cat_id = int(categoria_id) if str(categoria_id) != '0' else None

    movidos = 0
    for item_id in item_ids:
        item = ItemInventario.query.filter_by(id=int(item_id), organizacion_id=org_id).first()
        if item:
            item.categoria_id = cat_id
            movidos += 1

    db.session.commit()
    return jsonify(ok=True, movidos=movidos)


@inventario_bp.route('/eliminar-bulk', methods=['POST'])
@csrf.exempt
@login_required
def eliminar_bulk():
    """Eliminar múltiples items."""
    if current_user.rol != 'administrador':
        return jsonify(ok=False, error='Sin permisos'), 403

    data = request.get_json(silent=True)
    if not data:
        return jsonify(ok=False, error='Datos inválidos'), 400

    item_ids = data.get('item_ids', [])
    if not item_ids:
        return jsonify(ok=False, error='No hay items seleccionados'), 400

    org_id = get_current_org_id() or current_user.organizacion_id
    eliminados = 0
    for item_id in item_ids:
        item = ItemInventario.query.filter_by(id=int(item_id), organizacion_id=org_id).first()
        if item:
            # Eliminar movimientos asociados
            MovimientoInventario.query.filter_by(item_id=item.id).delete()
            UsoInventario.query.filter_by(item_id=item.id).delete()
            db.session.delete(item)
            eliminados += 1

    db.session.commit()
    return jsonify(ok=True, eliminados=eliminados)


@inventario_bp.route('/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar(id):
    """Elimina (desactiva) un item de inventario."""
    if current_user.rol not in ['administrador']:
        flash('No tienes permisos para eliminar items de inventario.', 'danger')
        return redirect(url_for('inventario.lista'))

    item = ItemInventario.query.get_or_404(id)
    org_id = get_current_org_id() or current_user.organizacion_id

    # Verificar que el item pertenece a la organización (excepto superadmin)
    if not current_user.is_super_admin and item.organizacion_id != org_id:
        flash('No tienes permisos para eliminar este item.', 'danger')
        return redirect(url_for('inventario.lista'))

    # Si tiene stock, el admin puede forzar eliminación (vacía stock automáticamente)
    if item.stock_actual and float(item.stock_actual) > 0:
        forzar = request.form.get('forzar_eliminar') == '1'
        if not forzar:
            flash(f'No se puede eliminar "{item.nombre}" porque tiene stock ({item.stock_actual} {item.unidad}). '
                  f'Primero da de baja todo el stock.', 'warning')
            return redirect(url_for('inventario.lista'))
        # Admin forzó: registrar baja del stock y continuar
        from datetime import datetime as dt_now
        baja = MovimientoInventario(
            item_id=item.id,
            tipo='salida',
            cantidad=item.stock_actual,
            motivo=f'Eliminación forzada por admin - stock vaciado',
            fecha=dt_now.utcnow(),
            usuario_id=current_user.id,
        )
        db.session.add(baja)
        item.stock_actual = 0
        db.session.flush()

    # Verificar si tiene movimientos o usos recientes (últimos 30 días)
    from datetime import datetime as dt, timedelta
    fecha_limite = dt.utcnow() - timedelta(days=30)
    movimientos_recientes = item.movimientos.filter(MovimientoInventario.fecha >= fecha_limite).count()
    usos_recientes = item.usos.filter(UsoInventario.fecha_uso >= fecha_limite.date()).count()

    if movimientos_recientes > 0 or usos_recientes > 0:
        # Soft delete - solo desactivar
        item.activo = False
        db.session.commit()
        flash(f'Item "{item.nombre}" desactivado (tiene historial reciente). El código {item.codigo} queda libre.', 'success')
    else:
        # Hard delete si no tiene historial reciente
        codigo = item.codigo
        nombre = item.nombre
        try:
            # Eliminar movimientos antiguos
            item.movimientos.delete()
            item.usos.delete()
            db.session.delete(item)
            db.session.commit()
            flash(f'Item "{nombre}" eliminado completamente. El código {codigo} está disponible.', 'success')
        except Exception as e:
            db.session.rollback()
            # Si falla el hard delete, hacer soft delete
            item.activo = False
            db.session.commit()
            flash(f'Item "{nombre}" desactivado. El código {codigo} queda libre.', 'success')

    return redirect(url_for('inventario.lista'))


@inventario_bp.route('/<int:id>/movimiento', methods=['POST'])
@login_required
def registrar_movimiento(id):
    if current_user.role not in ['admin', 'pm', 'tecnico']:
        flash('No tienes permisos para registrar movimientos.', 'danger')
        return redirect(url_for('inventario.detalle', id=id))

    item = ItemInventario.query.get_or_404(id)
    org_id = get_current_org_id() or current_user.organizacion_id

    # Verificar que el item pertenece a la organización (excepto superadmin)
    if not current_user.is_super_admin and item.organizacion_id != org_id:
        flash('No tienes permisos para modificar este item.', 'danger')
        return redirect(url_for('inventario.lista'))

    tipo = request.form.get('tipo')
    cantidad = request.form.get('cantidad')
    precio_unitario = request.form.get('precio_unitario', 0)
    motivo = request.form.get('motivo')
    observaciones = request.form.get('observaciones')
    
    if not all([tipo, cantidad]):
        flash('Tipo y cantidad son obligatorios.', 'danger')
        return redirect(url_for('inventario.detalle', id=id))
    
    try:
        cantidad = float(cantidad)
        precio_unitario = float(precio_unitario)
        
        if cantidad <= 0:
            flash('La cantidad debe ser mayor a cero.', 'danger')
            return redirect(url_for('inventario.detalle', id=id))
        
        # Verificar stock para salidas
        if tipo == 'salida' and cantidad > item.stock_actual:
            flash('Stock insuficiente para la salida solicitada.', 'danger')
            return redirect(url_for('inventario.detalle', id=id))
        
        # Crear movimiento
        movimiento = MovimientoInventario(
            item_id=id,
            tipo=tipo,
            cantidad=cantidad,
            precio_unitario=precio_unitario,
            motivo=motivo,
            observaciones=observaciones,
            usuario_id=current_user.id
        )
        
        # Actualizar stock
        if tipo == 'entrada':
            item.stock_actual += cantidad
            # Actualizar precio promedio
            if precio_unitario > 0:
                total_valor = (item.stock_actual - cantidad) * item.precio_promedio + cantidad * precio_unitario
                item.precio_promedio = total_valor / item.stock_actual
        elif tipo == 'salida':
            item.stock_actual -= cantidad
        elif tipo == 'ajuste':
            # Validaciones mejoradas para ajustes
            if not motivo or motivo.strip() == '':
                flash('⚠️ El motivo es OBLIGATORIO para ajustes de inventario. Especificá la razón (ej: conteo físico, merma, error sistema).', 'danger')
                return redirect(url_for('inventario.detalle', id=id))

            # Calcular diferencia porcentual
            stock_anterior = float(item.stock_actual) if item.stock_actual else 0
            diferencia = cantidad - stock_anterior
            if stock_anterior > 0:
                diferencia_porcentual = abs(diferencia / stock_anterior * 100)
            else:
                diferencia_porcentual = 100 if cantidad > 0 else 0

            # Alerta si el ajuste es mayor al 20%
            if diferencia_porcentual > 20:
                current_app.logger.warning(
                    f"⚠️ AJUSTE SIGNIFICATIVO: {item.nombre} - "
                    f"Stock anterior: {stock_anterior}, Nuevo: {cantidad}, "
                    f"Diferencia: {diferencia_porcentual:.1f}% - "
                    f"Usuario: {current_user.email}, Motivo: {motivo}"
                )
                flash(
                    f'⚠️ ALERTA: Ajuste significativo de {diferencia_porcentual:.1f}%. '
                    f'Stock anterior: {stock_anterior}, Nuevo stock: {cantidad}. '
                    f'Este ajuste ha sido registrado para auditoría.',
                    'warning'
                )

            # Aplicar ajuste
            item.stock_actual = cantidad
        
        db.session.add(movimiento)
        db.session.commit()

        # Verificar stock bajo y generar alerta si es necesario
        try:
            from services.stock_alerts_service import verificar_y_notificar_stock_bajo
            if verificar_y_notificar_stock_bajo(item, current_user.id, org_id):
                flash(f'Alerta: El stock de "{item.nombre}" esta por debajo del minimo.', 'warning')
        except Exception as alert_err:
            current_app.logger.warning(f"Error al verificar alerta de stock: {alert_err}")

        flash('Movimiento registrado exitosamente.', 'success')

    except ValueError:
        flash('Cantidad y precio deben ser números válidos.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash('Error al registrar el movimiento.', 'danger')
    
    return redirect(url_for('inventario.detalle', id=id))

@inventario_bp.route('/uso-obra', methods=['GET', 'POST'])
@login_required
def uso_obra():
    if not current_user.puede_acceder_modulo('inventario'):
        flash('No tienes permisos para registrar uso en obra.', 'danger')
        return redirect(url_for('reportes.dashboard'))

    org_id = get_current_org_id() or current_user.organizacion_id

    # Obtener todas las ubicaciones disponibles para el selector
    ubicaciones = Location.query.filter_by(
        organizacion_id=org_id,
        activo=True
    ).order_by(Location.tipo, Location.nombre).all()

    # Si no hay depósito general, crearlo
    deposito_general = Location.get_or_create_deposito_general(org_id)

    # Obtener obras que tienen presupuesto aprobado y crear sus ubicaciones si no existen
    from models.budgets import Presupuesto
    obras = Obra.query.join(Presupuesto).filter(
        db.or_(
            Presupuesto.confirmado_como_obra == True,
            Presupuesto.estado.in_(['aprobado', 'convertido', 'confirmado'])
        ),
        Obra.organizacion_id == org_id
    ).distinct().order_by(Obra.nombre).all()

    # Crear ubicaciones para obras que no las tengan
    for obra in obras:
        Location.get_or_create_for_obra(obra)

    # Refrescar lista de ubicaciones
    ubicaciones = Location.query.filter_by(
        organizacion_id=org_id,
        activo=True
    ).order_by(Location.tipo.desc(), Location.nombre).all()  # WAREHOUSE primero, luego WORKSITE

    # Ya no cargamos todos los items aquí - ahora se buscan vía AJAX
    # para mejor rendimiento en móviles (11k+ items)
    items = []

    # Obtener categorías para el modal de crear nuevo item
    categorias = InventoryCategory.query.filter_by(company_id=org_id, is_active=True).all() if org_id else []

    # Obtener proveedores de la organización
    from models import Proveedor
    proveedores = Proveedor.query.filter_by(organizacion_id=org_id, activo=True).order_by(Proveedor.nombre).all() if org_id else []

    context = {
        'ubicaciones': ubicaciones,
        'obras': obras,  # Mantenido para compatibilidad
        'items': items,
        'today': date.today(),
        'categorias': categorias,
        'proveedores': proveedores
    }

    if request.method == 'POST':
        location_id = request.form.get('obra_id')  # Ahora es location_id pero mantenemos el nombre del campo
        item_id = request.form.get('item_id')
        cantidad = request.form.get('cantidad')
        unidad = request.form.get('unidad', '').strip()
        fecha_compra = request.form.get('fecha_compra')
        marca = request.form.get('marca', '').strip()
        modelo = request.form.get('modelo', '').strip()
        proveedor_id = request.form.get('proveedor_id', '').strip()
        remito = request.form.get('remito', '').strip()

        if not all([location_id, item_id, cantidad, unidad]):
            flash('Destino, artículo, cantidad y unidad de medida son obligatorios.', 'danger')
            return render_template('inventario/uso_obra.html', **context)

        try:
            cantidad = float(cantidad)
            item_base = ItemInventario.query.get(item_id)

            if item_base is None:
                flash('El artículo seleccionado no existe.', 'danger')
                return render_template('inventario/uso_obra.html', **context)

            if cantidad <= 0:
                flash('La cantidad debe ser mayor a cero.', 'danger')
                return render_template('inventario/uso_obra.html', **context)

            # Lógica de variantes: si se especificó marca/modelo/proveedor diferente,
            # buscar o crear una variante del artículo
            item = item_base
            proveedor_id_int = int(proveedor_id) if proveedor_id else None

            # Verificar si necesitamos crear una variante
            necesita_variante = False
            if marca and marca != (item_base.marca or ''):
                necesita_variante = True
            if modelo and modelo != (item_base.modelo or ''):
                necesita_variante = True
            if proveedor_id_int and proveedor_id_int != item_base.proveedor_id:
                necesita_variante = True

            if necesita_variante:
                # Buscar si ya existe una variante con estos datos
                variante_existente = ItemInventario.query.filter_by(
                    organizacion_id=org_id,
                    nombre=item_base.nombre,
                    marca=marca or None,
                    modelo=modelo or None,
                    proveedor_id=proveedor_id_int,
                    activo=True
                ).first()

                if variante_existente:
                    item = variante_existente
                else:
                    # Generar código único para la variante
                    import random
                    import string
                    sufijo = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
                    nuevo_codigo = f"{item_base.codigo}-{sufijo}"

                    # Asegurar que el código sea único
                    while ItemInventario.query.filter_by(codigo=nuevo_codigo).first():
                        sufijo = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
                        nuevo_codigo = f"{item_base.codigo}-{sufijo}"

                    # Crear la variante
                    item = ItemInventario(
                        organizacion_id=org_id,
                        categoria_id=item_base.categoria_id,
                        codigo=nuevo_codigo,
                        nombre=item_base.nombre,
                        descripcion=item_base.descripcion,
                        unidad=unidad,
                        marca=marca or None,
                        modelo=modelo or None,
                        proveedor_id=proveedor_id_int,
                        stock_actual=0,
                        stock_minimo=item_base.stock_minimo,
                        precio_promedio=item_base.precio_promedio,
                        activo=True
                    )
                    db.session.add(item)
                    db.session.flush()
                    current_app.logger.info(f"[INVENTARIO] Variante creada: {nuevo_codigo} - {item_base.nombre} ({marca} {modelo})")

            # Actualizar unidad del item si cambió
            if item.unidad != unidad:
                item.unidad = unidad

            # Convertir fecha
            from datetime import datetime
            fecha_compra_dt = datetime.now()
            if fecha_compra:
                fecha_compra_dt = datetime.strptime(fecha_compra, '%Y-%m-%d')

            # Manejar compatibilidad: 'general' = depósito general
            if location_id == 'general':
                location = deposito_general
            else:
                location = Location.query.get(int(location_id))

            if not location:
                flash('La ubicación seleccionada no existe.', 'danger')
                return render_template('inventario/uso_obra.html', **context)

            # ============================================================
            # NUEVO SISTEMA: Stock por Ubicación
            # ============================================================

            # Buscar o crear el registro de stock en esta ubicación
            stock_ubicacion = StockUbicacion.query.filter_by(
                location_id=location.id,
                item_inventario_id=item.id
            ).first()

            if not stock_ubicacion:
                stock_ubicacion = StockUbicacion(
                    location_id=location.id,
                    item_inventario_id=item.id,
                    cantidad_disponible=0,
                    cantidad_reservada=0,
                    cantidad_consumida=0
                )
                db.session.add(stock_ubicacion)
                db.session.flush()

            # Actualizar cantidad disponible
            stock_ubicacion.cantidad_disponible = float(stock_ubicacion.cantidad_disponible or 0) + cantidad
            stock_ubicacion.fecha_ultima_entrada = fecha_compra_dt

            # Obtener nombre del proveedor si existe
            nombre_proveedor = None
            if proveedor_id_int:
                from models import Proveedor as ProveedorModel
                prov = ProveedorModel.query.get(proveedor_id_int)
                if prov:
                    nombre_proveedor = prov.nombre

            # Crear movimiento de stock
            movimiento = MovimientoStock(
                stock_ubicacion_id=stock_ubicacion.id,
                tipo='entrada',
                cantidad=cantidad,
                fecha=fecha_compra_dt,
                usuario_id=current_user.id,
                motivo='Compra',
                proveedor=nombre_proveedor,
                remito=remito if remito else None,
                precio_unitario=item.precio_promedio,
                moneda='ARS'
            )
            db.session.add(movimiento)

            # ============================================================
            # LEGACY: Mantener compatibilidad con sistema anterior
            # ============================================================

            # Actualizar stock_actual del item (suma total)
            item.stock_actual = float(item.stock_actual or 0) + cantidad

            # Crear movimiento legacy
            obs_parts = []
            if nombre_proveedor:
                obs_parts.append(f'Proveedor: {nombre_proveedor}')
            if marca:
                obs_parts.append(f'Marca: {marca}')
            if modelo:
                obs_parts.append(f'Modelo: {modelo}')
            if remito:
                obs_parts.append(f'Remito: {remito}')

            mov_legacy = MovimientoInventario(
                item_id=item.id,
                tipo='entrada',
                cantidad=cantidad,
                motivo=f'Compra - {location.nombre}',
                observaciones=' | '.join(obs_parts) if obs_parts else None,
                usuario_id=current_user.id
            )
            db.session.add(mov_legacy)

            # Si es una obra, también actualizar StockObra (legacy)
            if location.tipo == 'WORKSITE' and location.obra_id:
                from models.inventory import StockObra, MovimientoStockObra

                stock_obra = StockObra.query.filter_by(
                    obra_id=location.obra_id,
                    item_inventario_id=item.id
                ).first()

                if not stock_obra:
                    stock_obra = StockObra(
                        obra_id=location.obra_id,
                        item_inventario_id=item.id,
                        cantidad_disponible=0,
                        cantidad_consumida=0
                    )
                    db.session.add(stock_obra)
                    db.session.flush()

                stock_obra.cantidad_disponible = float(stock_obra.cantidad_disponible or 0) + cantidad
                stock_obra.fecha_ultimo_traslado = fecha_compra_dt

                mov_stock_obra = MovimientoStockObra(
                    stock_obra_id=stock_obra.id,
                    tipo='entrada',
                    cantidad=cantidad,
                    observaciones=f'Compra directa - {nombre_proveedor}' if nombre_proveedor else 'Compra directa',
                    usuario_id=current_user.id
                )
                db.session.add(mov_stock_obra)

            db.session.commit()

            # Construir mensaje de éxito con detalles de marca/modelo si aplica
            item_desc = item.nombre
            if marca or modelo:
                item_desc += f' ({marca or ""} {modelo or ""})'.strip()

            flash(f'✅ Compra registrada en {location.icono} {location.nombre}: {cantidad:.2f} {item.unidad} de {item_desc}', 'success')
            return redirect(url_for('inventario.lista'))

        except ValueError:
            flash('La cantidad debe ser un número válido.', 'danger')
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error al registrar compra: {str(e)}")
            import traceback
            current_app.logger.error(traceback.format_exc())
            flash(f'Error al registrar la compra: {str(e)}', 'danger')

    return render_template('inventario/uso_obra.html', **context)

@inventario_bp.route('/categorias')
@login_required
def categorias():
    if not current_user.puede_acceder_modulo('inventario'):
        flash('No tienes permisos para acceder al catálogo de categorías.', 'danger')
        return redirect(url_for('reportes.dashboard'))

    if not user_can_manage_inventory_categories(current_user):
        flash('No tienes permisos para gestionar el catálogo global de categorías.', 'danger')
        return redirect(url_for('inventario.lista'))

    company_id = _resolve_company_id()
    if not company_id:
        flash('No pudimos determinar la organización actual.', 'warning')
        return redirect(url_for('reportes.dashboard'))

    company = _resolve_company(company_id)
    if not company:
        flash('No pudimos cargar la organización seleccionada.', 'danger')
        return redirect(url_for('reportes.dashboard'))

    categorias, seed_stats, auto_seeded = ensure_categories_for_company(company)

    category_tree = _build_category_tree(categorias)

    context = {
        'categorias': categorias,
        'category_tree': category_tree,
        'auto_seeded': auto_seeded,
        'seed_stats': seed_stats,
        'company': company,
    }

    try:
        return render_template('inventario/categorias.html', **context)
    except TemplateNotFound:
        return render_category_catalog(context)


@inventario_bp.route('/api/categorias', methods=['GET'])
@inventario_bp.route('/api/categorias/', methods=['GET'])
@login_required
def api_categorias():
    company_id = _resolve_company_id()
    if not company_id:
        return jsonify({'error': 'Organización no seleccionada'}), 400

    categorias, seed_stats, auto_seeded, _ = ensure_categories_for_company_id(company_id)

    if auto_seeded or seed_stats.get('created') or seed_stats.get('reactivated'):
        current_app.logger.info(
            "[inventario] categorías auto-sembradas para org=%s (creadas=%s existentes=%s reactivadas=%s)",
            company_id,
            seed_stats.get('created', 0),
            seed_stats.get('existing', 0),
            seed_stats.get('reactivated', 0),
        )

    payload = [serialize_category(categoria) for categoria in categorias]
    payload.sort(key=lambda categoria: (categoria.get('full_path') or '').casefold())

    if not payload:
        current_app.logger.warning(
            "Inventory catalogue empty for org %s despite seeding attempts", company_id
        )

    return jsonify(payload)


@inventario_bp.route('/api/generar-codigo', methods=['POST'])
@login_required
def api_generar_codigo():
    """
    API para generar código automático correlativo único basado en categoría.
    Formato: PREFIJO + NÚMERO CORRELATIVO (ej: MAT001, HER001, MAQ001)
    """
    import re

    data = request.get_json() or {}
    categoria_id = data.get('categoria_id')

    org_id = get_current_org_id() or current_user.organizacion_id

    if not categoria_id:
        return jsonify({'error': 'La categoría es requerida'}), 400

    try:
        # Obtener categoría
        categoria = InventoryCategory.query.get(categoria_id)
        if not categoria:
            return jsonify({'error': 'Categoría no encontrada'}), 404

        # Generar prefijo de 3 letras basado en el nombre de la categoría
        categoria_nombre = categoria.nombre.upper()
        # Remover caracteres especiales y quedarse con letras
        prefijo = re.sub(r'[^A-Z]', '', categoria_nombre)[:3]

        # Si el prefijo es muy corto, completar
        if len(prefijo) < 3:
            prefijo = prefijo.ljust(3, 'X')

        # Buscar el último código con este prefijo en la organización
        ultimo_item = ItemInventario.query.filter(
            ItemInventario.organizacion_id == org_id,
            ItemInventario.codigo.like(f'{prefijo}%')
        ).order_by(ItemInventario.codigo.desc()).first()

        # Determinar siguiente número
        siguiente_numero = 1
        if ultimo_item and ultimo_item.codigo:
            # Extraer número del código existente
            match = re.search(r'(\d+)$', ultimo_item.codigo)
            if match:
                siguiente_numero = int(match.group(1)) + 1

        # Generar código con formato PREFIJO + 3 dígitos
        codigo = f"{prefijo}{siguiente_numero:03d}"

        # Verificar que no exista (por si acaso)
        while ItemInventario.query.filter_by(codigo=codigo, organizacion_id=org_id).first():
            siguiente_numero += 1
            codigo = f"{prefijo}{siguiente_numero:03d}"

        return jsonify({
            'codigo': codigo,
            'prefijo': prefijo,
            'numero': siguiente_numero,
            'categoria_nombre': categoria.nombre,
            'mensaje': f'Código generado: {codigo}'
        })

    except Exception as e:
        current_app.logger.error(f"Error generando código: {str(e)}")
        return jsonify({'error': 'Error al generar código automático'}), 500


@inventario_bp.route('/api/crear-item', methods=['POST'])
@login_required
def api_crear_item():
    """
    API para crear un nuevo item de inventario (usado desde modal de Compras).
    """
    if not current_user.puede_acceder_modulo('inventario'):
        return jsonify({'ok': False, 'error': 'Sin permisos para crear items'}), 403

    org_id = get_current_org_id() or current_user.organizacion_id
    if not org_id:
        return jsonify({'ok': False, 'error': 'No tienes una organización activa'}), 400

    try:
        data = request.get_json() or {}

        codigo = (data.get('codigo') or '').strip().upper()
        nombre = (data.get('nombre') or '').strip()
        unidad = (data.get('unidad') or '').strip()
        categoria_id = data.get('categoria_id')
        stock_minimo = float(data.get('stock_minimo') or 0)
        descripcion = (data.get('descripcion') or '').strip()

        # Validaciones
        if not codigo or not nombre or not unidad:
            return jsonify({'ok': False, 'error': 'Código, nombre y unidad son obligatorios'}), 400

        # Verificar código duplicado
        existing = ItemInventario.query.filter_by(codigo=codigo).first()
        if existing:
            return jsonify({'ok': False, 'error': f'Ya existe un item con el código {codigo}'}), 400

        # Verificar nombre duplicado en la organización
        existing_nombre = ItemInventario.query.filter(
            ItemInventario.organizacion_id == org_id,
            ItemInventario.nombre.ilike(nombre),
            ItemInventario.activo == True
        ).first()
        if existing_nombre:
            return jsonify({
                'ok': False,
                'error': f'Ya existe "{nombre}" (código: {existing_nombre.codigo})'
            }), 400

        # Crear el item
        nuevo_item = ItemInventario(
            organizacion_id=org_id,
            categoria_id=int(categoria_id) if categoria_id else None,
            codigo=codigo,
            nombre=nombre,
            descripcion=descripcion,
            unidad=unidad,
            stock_actual=0,
            stock_minimo=stock_minimo,
            precio_promedio=0,
            activo=True
        )

        db.session.add(nuevo_item)
        db.session.commit()

        current_app.logger.info(f"Item creado via API: {codigo} - {nombre} por usuario {current_user.id}")

        return jsonify({
            'ok': True,
            'item': {
                'id': nuevo_item.id,
                'codigo': nuevo_item.codigo,
                'nombre': nuevo_item.nombre,
                'unidad': nuevo_item.unidad,
                'categoria': nuevo_item.categoria.nombre if nuevo_item.categoria else None
            }
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creando item via API: {str(e)}")
        return jsonify({'ok': False, 'error': 'Error al crear el artículo'}), 500


@inventario_bp.route('/api/tipo-cambio', methods=['GET'])
@login_required
def api_tipo_cambio():
    """
    API para obtener el tipo de cambio actual USD/ARS.
    Obtiene la cotización del Banco Nación Argentina (BNA).
    """
    from models import ExchangeRate
    from datetime import date, datetime, timedelta
    import requests
    from bs4 import BeautifulSoup

    def obtener_cotizacion_bna():
        """Obtiene la cotización del dólar del Banco Nación Argentina"""
        try:
            url = "https://www.bna.com.ar/Personas"
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()

            soup = BeautifulSoup(response.text, 'html.parser')

            # Buscar la tabla de cotizaciones
            tabla = soup.find('table', class_='table')
            if not tabla:
                return None

            # Buscar la fila del dólar estadounidense (Billetes)
            filas = tabla.find_all('tr')
            for fila in filas:
                celdas = fila.find_all('td')
                if len(celdas) >= 3:
                    moneda = celdas[0].get_text(strip=True).lower()
                    # BNA usa "Dolar U.S.A" para el dólar estadounidense
                    if 'dolar' in moneda and ('u.s.a' in moneda or 'estadounidense' in moneda):
                        # Obtener precio de venta (tercera columna)
                        venta_text = celdas[2].get_text(strip=True)
                        # Limpiar y convertir: "1.480,00" -> 1480.00
                        venta_text = venta_text.replace('.', '').replace(',', '.')
                        return float(venta_text)

            return None
        except Exception as e:
            current_app.logger.warning(f"Error obteniendo cotización BNA: {str(e)}")
            return None

    try:
        # Primero intentar obtener cotización en tiempo real del BNA
        cotizacion_bna = obtener_cotizacion_bna()

        if cotizacion_bna:
            # Guardar/actualizar en la base de datos
            try:
                rate = ExchangeRate.query.filter(
                    ExchangeRate.base_currency == 'USD',
                    ExchangeRate.quote_currency == 'ARS',
                    ExchangeRate.provider == 'BNA'
                ).first()

                if rate:
                    rate.value = cotizacion_bna
                    rate.as_of_date = date.today()
                    rate.fetched_at = datetime.utcnow()
                else:
                    rate = ExchangeRate(
                        base_currency='USD',
                        quote_currency='ARS',
                        value=cotizacion_bna,
                        provider='BNA',
                        as_of_date=date.today(),
                        fetched_at=datetime.utcnow()
                    )
                    db.session.add(rate)

                db.session.commit()
            except Exception as db_error:
                current_app.logger.warning(f"Error guardando cotización: {str(db_error)}")
                db.session.rollback()

            return jsonify({
                'success': True,
                'tipo_cambio': cotizacion_bna,
                'provider': 'BNA',
                'fecha': date.today().isoformat(),
                'actualizado': datetime.utcnow().isoformat()
            })

        # Si no se pudo obtener del BNA, buscar en BD
        rate = ExchangeRate.query.filter(
            ExchangeRate.base_currency == 'USD',
            ExchangeRate.quote_currency == 'ARS'
        ).order_by(
            ExchangeRate.as_of_date.desc(),
            ExchangeRate.fetched_at.desc()
        ).first()

        if rate:
            return jsonify({
                'success': True,
                'tipo_cambio': float(rate.value),
                'provider': rate.provider + ' (caché)',
                'fecha': rate.as_of_date.isoformat() if rate.as_of_date else None,
                'actualizado': rate.fetched_at.isoformat() if rate.fetched_at else None
            })
        else:
            # Valor por defecto si todo falla
            return jsonify({
                'success': True,
                'tipo_cambio': 1150.0,
                'provider': 'default',
                'fecha': date.today().isoformat(),
                'actualizado': None,
                'mensaje': 'Usando cotización por defecto - no se pudo conectar al BNA'
            })

    except Exception as e:
        current_app.logger.error(f"Error obteniendo tipo de cambio: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Error al obtener tipo de cambio',
            'tipo_cambio': 1150.0
        }), 500


@inventario_bp.route('/api/buscar-similares', methods=['POST'])
@login_required
def api_buscar_similares():
    """
    API para buscar materiales similares en el catálogo global.
    Retorna sugerencias de materiales existentes.
    """
    from models import GlobalMaterialCatalog, InventoryCategory

    data = request.get_json() or {}

    nombre = data.get('nombre', '').strip()
    categoria_id = data.get('categoria_id')
    marca = data.get('marca', '').strip() or None
    limit = int(data.get('limit', 10))

    if not nombre:
        return jsonify({'similares': []})

    # Obtener nombre de categoría si se especifica
    categoria_nombre = None
    if categoria_id:
        categoria = InventoryCategory.query.get(categoria_id)
        if categoria:
            categoria_nombre = categoria.nombre

    # Buscar materiales similares
    try:
        similares = GlobalMaterialCatalog.buscar_similares(
            nombre=nombre,
            categoria_nombre=categoria_nombre,
            marca=marca,
            limit=limit
        )

        resultados = []
        for material in similares:
            resultados.append({
                'id': material.id,
                'codigo': material.codigo,
                'nombre': material.nombre,
                'descripcion_completa': material.descripcion_completa,
                'categoria': material.categoria_nombre,
                'marca': material.marca,
                'unidad': material.unidad,
                'veces_usado': material.veces_usado,
                'precio_promedio_ars': float(material.precio_promedio_ars) if material.precio_promedio_ars else None,
                'precio_promedio_usd': float(material.precio_promedio_usd) if material.precio_promedio_usd else None,
            })

        return jsonify({
            'similares': resultados,
            'total': len(resultados),
            'mensaje': f'Se encontraron {len(resultados)} materiales similares en el catálogo global'
        })

    except Exception as e:
        current_app.logger.error(f"Error buscando similares: {str(e)}")
        return jsonify({'error': 'Error al buscar materiales similares', 'similares': []}), 500


@inventario_bp.route('/api/usar-material-global/<int:material_id>', methods=['POST'])
@login_required
def api_usar_material_global(material_id):
    """
    API para usar un material del catálogo global.
    Crea un item de inventario local y registra el uso.
    """
    from models import GlobalMaterialCatalog, GlobalMaterialUsage, ItemInventario, MovimientoInventario

    org_id = get_current_org_id() or current_user.organizacion_id
    if not org_id:
        return jsonify({'error': 'No tienes una organización activa'}), 400

    # Obtener material del catálogo global
    material = GlobalMaterialCatalog.query.get(material_id)
    if not material:
        return jsonify({'error': 'Material no encontrado en el catálogo global'}), 404

    data = request.get_json() or {}
    stock_inicial = float(data.get('stock_inicial', 0))
    precio_ars = float(data.get('precio_ars', 0))
    precio_usd = float(data.get('precio_usd', 0))

    try:
        # Verificar si ya existe un item con este código
        existing = ItemInventario.query.filter_by(
            codigo=material.codigo,
            organizacion_id=org_id
        ).first()

        if existing:
            return jsonify({
                'error': f'Ya tienes un material con código {material.codigo} en tu inventario',
                'item_id': existing.id
            }), 409

        # Verificar si ya existe un item con el mismo nombre + descripción
        descripcion_material = (material.descripcion_completa or '').strip().lower()

        query_duplicado = ItemInventario.query.filter(
            ItemInventario.organizacion_id == org_id,
            ItemInventario.nombre.ilike(material.nombre),
            ItemInventario.activo == True
        )

        if descripcion_material:
            query_duplicado = query_duplicado.filter(
                db.func.lower(db.func.coalesce(ItemInventario.descripcion, '')).like(descripcion_material)
            )
        else:
            query_duplicado = query_duplicado.filter(
                db.or_(
                    ItemInventario.descripcion.is_(None),
                    ItemInventario.descripcion == ''
                )
            )

        existing_duplicado = query_duplicado.first()
        if existing_duplicado:
            return jsonify({
                'error': f'Ya tienes "{material.nombre}" con la misma descripción (código: {existing_duplicado.codigo})',
                'item_id': existing_duplicado.id
            }), 409

        # Buscar categoría correspondiente
        categoria = InventoryCategory.query.filter_by(
            company_id=org_id,
            nombre=material.categoria_nombre,
            is_active=True
        ).first()

        if not categoria:
            # Crear categoría si no existe
            categoria = InventoryCategory(
                company_id=org_id,
                nombre=material.categoria_nombre,
                is_active=True
            )
            db.session.add(categoria)
            db.session.flush()

        # Crear item de inventario local
        nuevo_item = ItemInventario(
            organizacion_id=org_id,
            categoria_id=categoria.id,
            codigo=material.codigo,
            nombre=material.nombre,
            descripcion=material.descripcion_completa,
            unidad=material.unidad,
            stock_actual=stock_inicial,
            stock_minimo=0,
            precio_promedio=precio_ars or material.precio_promedio_ars or 0,
            precio_promedio_usd=precio_usd or material.precio_promedio_usd or 0,
            activo=True
        )

        db.session.add(nuevo_item)
        db.session.flush()

        # Registrar uso del material global
        uso = GlobalMaterialUsage(
            material_id=material.id,
            organizacion_id=org_id,
            item_inventario_id=nuevo_item.id
        )
        db.session.add(uso)

        # Actualizar contador de usos
        material.veces_usado = (material.veces_usado or 0) + 1

        # Si hay stock inicial, crear movimiento
        if stock_inicial > 0:
            movimiento = MovimientoInventario(
                item_id=nuevo_item.id,
                tipo='entrada',
                cantidad=stock_inicial,
                precio_unitario=precio_ars,
                motivo='Inventario inicial desde catálogo global',
                observaciones=f'Material importado del catálogo global: {material.codigo}',
                usuario_id=current_user.id
            )
            db.session.add(movimiento)

        db.session.commit()

        return jsonify({
            'success': True,
            'item_id': nuevo_item.id,
            'codigo': nuevo_item.codigo,
            'mensaje': f'Material {material.nombre} agregado a tu inventario exitosamente'
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error usando material global: {str(e)}")
        return jsonify({'error': 'Error al agregar material a tu inventario'}), 500


@inventario_bp.route('/categoria', methods=['POST'])
@login_required
def crear_categoria():
    if not current_user.puede_acceder_modulo('inventario'):
        flash('No tienes permisos para actualizar el catálogo.', 'danger')
        return redirect(url_for('inventario.categorias'))

    if not user_can_manage_inventory_categories(current_user):
        flash('No tienes permisos para modificar el catálogo global.', 'danger')
        return redirect(url_for('inventario.lista'))

    company_id = _resolve_company_id()
    if not company_id:
        flash('No pudimos determinar la organización actual.', 'warning')
        return redirect(url_for('inventario.categorias'))

    company = _resolve_company(company_id)
    if not company:
        flash('No encontramos la organización seleccionada.', 'danger')
        return redirect(url_for('inventario.categorias'))

    stats = seed_inventory_categories_for_company(company, mark_global=True)
    db.session.commit()

    created = stats.get('created', 0)
    existing = stats.get('existing', 0)
    reactivated = stats.get('reactivated', 0)
    message = (
        f"Catálogo listo: {created} nuevas, {existing} existentes, {reactivated} reactivadas."
    )

    flash(message, 'success' if created else 'info')
    return redirect(url_for('inventario.categorias'))

@inventario_bp.route('/api/buscar-items', methods=['GET'])
@login_required
def api_buscar_items():
    """
    API para búsqueda AJAX de items del inventario.
    Soporta Select2 con paginación para grandes volúmenes de datos.
    """
    try:
        org_id = get_current_org_id() or current_user.organizacion_id
        if not org_id:
            return jsonify({'results': [], 'pagination': {'more': False}})

        # Parámetros de búsqueda
        term = request.args.get('term', '').strip()
        page = request.args.get('page', 1, type=int)
        per_page = 30  # Items por página

        # Query base
        query = ItemInventario.query.filter_by(
            organizacion_id=org_id,
            activo=True
        )

        # Filtrar por término de búsqueda
        if term:
            search_term = f'%{term}%'
            query = query.filter(
                db.or_(
                    ItemInventario.codigo.ilike(search_term),
                    ItemInventario.nombre.ilike(search_term),
                    ItemInventario.descripcion.ilike(search_term)
                )
            )

        # Ordenar y paginar
        query = query.order_by(ItemInventario.nombre)
        total = query.count()
        items = query.offset((page - 1) * per_page).limit(per_page).all()

        # Formatear resultados para Select2
        results = []
        for item in items:
            text = f'[{item.codigo}] {item.nombre}'
            if item.marca:
                text += f' - {item.marca}'
            if item.modelo:
                text += f' {item.modelo}'
            if item.categoria:
                text += f' ({item.categoria.nombre})'

            results.append({
                'id': item.id,
                'text': text,
                'codigo': item.codigo,
                'nombre': item.nombre,
                'unidad': item.unidad or 'unidad',
                'marca': item.marca or '',
                'modelo': item.modelo or '',
                'categoria': item.categoria.nombre if item.categoria else ''
            })

        return jsonify({
            'results': results,
            'pagination': {
                'more': (page * per_page) < total
            }
        })

    except Exception as e:
        current_app.logger.error(f"Error en api_buscar_items: {str(e)}")
        return jsonify({'results': [], 'pagination': {'more': False}})


@inventario_bp.route('/items-disponibles', methods=['GET'])
@login_required
def items_disponibles():
    """
    Endpoint para obtener items del inventario disponibles para una obra.
    Retorna lista de materiales con stock actual para selector de consumo.
    """
    try:
        obra_id = request.args.get('obra_id', type=int)
        org_id = get_current_org_id() or current_user.organizacion_id

        if not org_id:
            return jsonify({'ok': False, 'error': 'No tienes una organización activa'}), 400

        # Obtener todos los items del inventario de la organización con stock > 0
        items = ItemInventario.query.filter_by(
            organizacion_id=org_id,
            activo=True
        ).filter(
            ItemInventario.stock_actual > 0
        ).order_by(ItemInventario.descripcion).all()

        items_data = []
        for item in items:
            items_data.append({
                'id': item.id,
                'descripcion': item.descripcion,
                'stock_actual': float(item.stock_actual or 0),
                'unidad': item.unidad or 'un',
                'categoria': item.categoria.nombre if item.categoria else 'Sin categoría'
            })

        return jsonify({
            'ok': True,
            'items': items_data
        })

    except Exception as e:
        current_app.logger.error(f"Error obteniendo items disponibles: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({'ok': False, 'error': str(e)}), 500


@inventario_bp.route('/analisis', methods=['GET'])
@login_required
def analisis():
    """
    Análisis de consumo de inventario:
    - Artículos más consumidos
    - Consumo por obra
    - Costos reales de materiales
    - Tendencias de consumo
    """
    if not current_user.puede_acceder_modulo('inventario'):
        flash('No tienes permisos para acceder a este módulo.', 'danger')
        return redirect(url_for('reportes.dashboard'))

    try:
        from datetime import datetime, timedelta
        from sqlalchemy import func, desc

        org_id = get_current_org_id() or current_user.organizacion_id

        if not org_id:
            flash('No tienes una organización activa', 'warning')
            return redirect(url_for('index'))

        # Filtros de fecha (por defecto últimos 30 días)
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        obra_id = request.args.get('obra_id', type=int)

        if not fecha_desde:
            fecha_desde = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        if not fecha_hasta:
            fecha_hasta = datetime.now().strftime('%Y-%m-%d')

        fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
        fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')

        # 1. ARTÍCULOS MÁS CONSUMIDOS (Top 10)
        query_top_items = db.session.query(
            ItemInventario.id,
            ItemInventario.descripcion,
            ItemInventario.unidad,
            func.sum(UsoInventario.cantidad).label('total_consumido'),
            func.count(UsoInventario.id).label('num_usos'),
            func.avg(ItemInventario.precio_promedio).label('precio_promedio')
        ).join(
            UsoInventario, ItemInventario.id == UsoInventario.item_id
        ).filter(
            ItemInventario.organizacion_id == org_id,
            UsoInventario.fecha >= fecha_desde_dt,
            UsoInventario.fecha <= fecha_hasta_dt
        )

        if obra_id:
            query_top_items = query_top_items.filter(UsoInventario.obra_id == obra_id)

        top_items = query_top_items.group_by(
            ItemInventario.id,
            ItemInventario.descripcion,
            ItemInventario.unidad
        ).order_by(desc('total_consumido')).limit(10).all()

        # Calcular costo total de cada item
        top_items_data = []
        for item in top_items:
            costo_total = float(item.total_consumido) * (float(item.precio_promedio) if item.precio_promedio else 0)
            top_items_data.append({
                'id': item.id,
                'descripcion': item.descripcion,
                'unidad': item.unidad,
                'total_consumido': float(item.total_consumido),
                'num_usos': item.num_usos,
                'precio_promedio': float(item.precio_promedio) if item.precio_promedio else 0,
                'costo_total': costo_total
            })

        # 2. CONSUMO POR OBRA
        query_obras = db.session.query(
            Obra.id,
            Obra.nombre,
            Obra.direccion,
            func.count(UsoInventario.id).label('num_consumos'),
            func.sum(UsoInventario.cantidad * ItemInventario.precio_promedio).label('costo_total')
        ).join(
            UsoInventario, Obra.id == UsoInventario.obra_id
        ).join(
            ItemInventario, UsoInventario.item_id == ItemInventario.id
        ).filter(
            Obra.organizacion_id == org_id,
            UsoInventario.fecha >= fecha_desde_dt,
            UsoInventario.fecha <= fecha_hasta_dt
        )

        if obra_id:
            query_obras = query_obras.filter(Obra.id == obra_id)

        consumo_obras = query_obras.group_by(
            Obra.id,
            Obra.nombre,
            Obra.direccion
        ).order_by(desc('costo_total')).all()

        consumo_obras_data = []
        for obra in consumo_obras:
            consumo_obras_data.append({
                'id': obra.id,
                'nombre': obra.nombre,
                'direccion': obra.direccion,
                'num_consumos': obra.num_consumos,
                'costo_total': float(obra.costo_total) if obra.costo_total else 0
            })

        # 3. ITEMS CON STOCK BAJO (alertas)
        items_stock_bajo = ItemInventario.query.filter(
            ItemInventario.organizacion_id == org_id,
            ItemInventario.activo == True,
            ItemInventario.stock_actual <= ItemInventario.stock_minimo
        ).order_by(ItemInventario.stock_actual).all()

        # 4. RESUMEN GENERAL
        total_items_activos = ItemInventario.query.filter_by(
            organizacion_id=org_id,
            activo=True
        ).count()

        valor_total_inventario = db.session.query(
            func.sum(ItemInventario.stock_actual * ItemInventario.precio_promedio)
        ).filter(
            ItemInventario.organizacion_id == org_id,
            ItemInventario.activo == True
        ).scalar() or 0

        total_consumos_periodo = UsoInventario.query.filter(
            UsoInventario.fecha >= fecha_desde_dt,
            UsoInventario.fecha <= fecha_hasta_dt
        ).join(ItemInventario).filter(
            ItemInventario.organizacion_id == org_id
        ).count()

        costo_total_periodo = db.session.query(
            func.sum(UsoInventario.cantidad * ItemInventario.precio_promedio)
        ).join(
            ItemInventario, UsoInventario.item_id == ItemInventario.id
        ).filter(
            ItemInventario.organizacion_id == org_id,
            UsoInventario.fecha >= fecha_desde_dt,
            UsoInventario.fecha <= fecha_hasta_dt
        ).scalar() or 0

        # Obtener todas las obras para el filtro
        todas_obras = Obra.query.filter_by(organizacion_id=org_id).order_by(Obra.nombre).all()

        return render_template('inventario/analisis.html',
                             top_items=top_items_data,
                             consumo_obras=consumo_obras_data,
                             items_stock_bajo=items_stock_bajo,
                             total_items_activos=total_items_activos,
                             valor_total_inventario=float(valor_total_inventario),
                             total_consumos_periodo=total_consumos_periodo,
                             costo_total_periodo=float(costo_total_periodo),
                             fecha_desde=fecha_desde,
                             fecha_hasta=fecha_hasta,
                             obra_id=obra_id,
                             todas_obras=todas_obras)

    except Exception as e:
        current_app.logger.error(f"Error en análisis de inventario: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        flash('Error al generar el análisis de inventario', 'danger')
        return redirect(url_for('inventario.lista'))


@inventario_bp.route('/dar_baja/<int:id>', methods=['POST'])
@login_required
def dar_baja(id):
    """Da de baja un item de inventario por uso o rotura"""
    if not current_user.puede_acceder_modulo('inventario'):
        return jsonify({'success': False, 'message': 'No tienes permisos'}), 403

    try:
        item = ItemInventario.query.get_or_404(id)

        from decimal import Decimal
        cantidad = Decimal(str(request.form.get('cantidad', 0)))
        motivo = request.form.get('motivo', '')
        obra_id = request.form.get('obra_id')
        observaciones = request.form.get('observaciones', '')

        if cantidad <= 0:
            return jsonify({'success': False, 'message': 'La cantidad debe ser mayor a 0'}), 400

        stock_actual = Decimal(str(item.stock_actual)) if item.stock_actual else Decimal('0')
        if cantidad > stock_actual:
            return jsonify({'success': False, 'message': 'No hay suficiente stock disponible'}), 400

        # Create movement record
        movimiento = MovimientoInventario(
            item_id=item.id,
            tipo='salida',
            cantidad=float(cantidad),
            motivo=motivo,
            observaciones=observaciones,
            usuario_id=current_user.id
        )
        db.session.add(movimiento)

        # Update stock
        item.stock_actual = float(stock_actual - cantidad)

        # If it's usage on an obra, record it
        if obra_id and motivo == 'Uso en obra':
            uso = UsoInventario(
                obra_id=int(obra_id),
                item_id=item.id,
                cantidad_usada=float(cantidad),
                fecha_uso=date.today(),
                observaciones=observaciones,
                usuario_id=current_user.id
            )
            db.session.add(uso)

        db.session.commit()
        flash(f'Se dio de baja {cantidad} {item.unidad} de {item.nombre}', 'success')
        return redirect(url_for('inventario.lista'))

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error en dar_baja: {str(e)}")
        flash(f'Error al dar de baja: {str(e)}', 'danger')
        return redirect(url_for('inventario.lista'))


@inventario_bp.route('/trasladar/<int:id>', methods=['POST'])
@login_required
def trasladar(id):
    """Traslada un item de inventario general al stock de una obra"""
    if not current_user.puede_acceder_modulo('inventario'):
        return jsonify({'success': False, 'message': 'No tienes permisos'}), 403

    try:
        from models.inventory import StockObra, MovimientoStockObra
        from decimal import Decimal
        from datetime import datetime

        item = ItemInventario.query.get_or_404(id)

        cantidad = Decimal(str(request.form.get('cantidad', 0)))
        obra_destino_id = request.form.get('obra_destino_id')
        observaciones = request.form.get('observaciones', '')

        if cantidad <= 0:
            return jsonify({'success': False, 'message': 'La cantidad debe ser mayor a 0'}), 400

        if not obra_destino_id:
            return jsonify({'success': False, 'message': 'Debe seleccionar una obra destino'}), 400

        # Validación estricta de stock - NO permite stock negativo
        stock_actual = Decimal(str(item.stock_actual)) if item.stock_actual else Decimal('0')
        if cantidad > stock_actual:
            current_app.logger.warning(
                f"❌ Traslado bloqueado: stock insuficiente para {item.nombre}. "
                f"Disponible: {stock_actual}, Trasladar: {cantidad}"
            )
            return jsonify({
                'success': False,
                'message': f'Stock insuficiente. Disponible: {float(stock_actual):.2f} {item.unidad}, solicitado: {float(cantidad):.2f} {item.unidad}'
            }), 400

        # Validar que la obra pertenezca a la organización del usuario
        org_id = get_current_org_id() or current_user.organizacion_id
        obra_destino = Obra.query.filter_by(
            id=int(obra_destino_id),
            organizacion_id=org_id
        ).first()

        if not obra_destino:
            return jsonify({'success': False, 'message': 'Obra no encontrada o no pertenece a tu organización'}), 404

        # 1. Registrar movimiento de SALIDA del inventario general
        movimiento_salida = MovimientoInventario(
            item_id=item.id,
            tipo='salida',
            cantidad=float(cantidad),
            motivo=f'Traslado a obra: {obra_destino.nombre}',
            observaciones=observaciones,
            usuario_id=current_user.id
        )
        db.session.add(movimiento_salida)

        # 2. Descontar del inventario general
        item.stock_actual = float(stock_actual - cantidad)

        # 3. Buscar o crear el registro de StockObra
        stock_obra = StockObra.query.filter_by(
            obra_id=int(obra_destino_id),
            item_inventario_id=item.id
        ).first()

        if not stock_obra:
            stock_obra = StockObra(
                obra_id=int(obra_destino_id),
                item_inventario_id=item.id,
                cantidad_disponible=0,
                cantidad_consumida=0
            )
            db.session.add(stock_obra)
            db.session.flush()  # Para obtener el ID

        # 4. Sumar al stock de la obra
        stock_obra.cantidad_disponible = float(
            Decimal(str(stock_obra.cantidad_disponible or 0)) + cantidad
        )
        stock_obra.fecha_ultimo_traslado = datetime.utcnow()

        # 5. Registrar movimiento de ENTRADA en el stock de obra
        movimiento_entrada = MovimientoStockObra(
            stock_obra_id=stock_obra.id,
            tipo='entrada',
            cantidad=float(cantidad),
            fecha=datetime.utcnow(),
            usuario_id=current_user.id,
            observaciones=f'Traslado desde inventario general. {observaciones}',
            precio_unitario=item.precio_promedio,
            moneda='ARS'
        )
        db.session.add(movimiento_entrada)

        db.session.commit()

        flash(f'Se trasladaron {cantidad} {item.unidad} de {item.nombre} a {obra_destino.nombre}', 'success')
        return redirect(url_for('inventario.lista'))

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error en trasladar: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@inventario_bp.route('/api/<int:item_id>/stock-obras', methods=['GET'])
@login_required
def api_stock_en_obras(item_id):
    """Obtiene el stock de un item en todas las obras"""
    try:
        item = ItemInventario.query.get_or_404(item_id)

        # Obtener stock en obras
        stocks = StockObra.query.filter_by(item_inventario_id=item_id).all()

        stock_obras = []
        for stock in stocks:
            if stock.cantidad_disponible and float(stock.cantidad_disponible) > 0:
                stock_obras.append({
                    'obra_id': stock.obra_id,
                    'obra_nombre': stock.obra.nombre if stock.obra else 'Obra desconocida',
                    'cantidad_disponible': float(stock.cantidad_disponible or 0),
                    'cantidad_consumida': float(stock.cantidad_consumida or 0),
                    'fecha_ultimo_traslado': stock.fecha_ultimo_traslado.strftime('%d/%m/%Y') if stock.fecha_ultimo_traslado else None
                })

        return jsonify({
            'success': True,
            'stock_obras': stock_obras
        })

    except Exception as e:
        current_app.logger.error(f"Error en api_stock_en_obras: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@inventario_bp.route('/api/<int:item_id>/traslados', methods=['GET'])
@login_required
def api_traslados_item(item_id):
    """Obtiene el historial de traslados de un item desde la tabla stock_obra"""
    try:
        item = ItemInventario.query.get_or_404(item_id)

        # Obtener traslados desde stock_obra (datos reales)
        stocks = StockObra.query.filter_by(item_inventario_id=item_id).all()

        traslados = []
        for stock in stocks:
            if stock.cantidad_disponible and float(stock.cantidad_disponible) > 0:
                traslados.append({
                    'fecha': stock.fecha_ultimo_traslado.strftime('%d/%m/%Y %H:%M') if stock.fecha_ultimo_traslado else '-',
                    'obra_id': stock.obra_id,
                    'obra_nombre': stock.obra.nombre if stock.obra else 'Desconocida',
                    'cantidad': float(stock.cantidad_disponible or 0),
                    'usuario': None  # No tenemos esta info en stock_obra
                })

        return jsonify({
            'success': True,
            'traslados': traslados
        })

    except Exception as e:
        current_app.logger.error(f"Error en api_traslados_item: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================
# ALERTAS DE STOCK BAJO
# ============================================================

@inventario_bp.route('/api/alertas-stock', methods=['GET'])
@login_required
def api_alertas_stock():
    """Obtiene las alertas de stock bajo para la organización actual"""
    from services.stock_alerts_service import obtener_resumen_alertas
    from services.memberships import get_current_org_id

    try:
        org_id = get_current_org_id()
        if not org_id:
            return jsonify({'success': False, 'message': 'No hay organización seleccionada'}), 400

        resumen = obtener_resumen_alertas(org_id)

        return jsonify({
            'success': True,
            'alertas': resumen
        })

    except Exception as e:
        current_app.logger.error(f"Error en api_alertas_stock: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@inventario_bp.route('/api/alertas-stock/count', methods=['GET'])
@login_required
def api_alertas_stock_count():
    """Obtiene solo el conteo de alertas de stock bajo (para badges)"""
    from services.stock_alerts_service import contar_alertas_stock
    from services.memberships import get_current_org_id

    try:
        org_id = get_current_org_id()
        if not org_id:
            return jsonify({'success': False, 'count': 0}), 200

        conteo = contar_alertas_stock(org_id)

        return jsonify({
            'success': True,
            'count': conteo['total'],
            'critico': conteo['critico'],
            'bajo': conteo['bajo'] + conteo['muy_bajo'] + conteo['alerta']
        })

    except Exception as e:
        current_app.logger.error(f"Error en api_alertas_stock_count: {str(e)}")
        return jsonify({'success': False, 'count': 0}), 200


@inventario_bp.route('/api/alertas-stock/generar', methods=['POST'])
@login_required
def api_generar_alertas():
    """Genera notificaciones para todos los items con stock bajo"""
    from services.stock_alerts_service import generar_alertas_masivas
    from services.memberships import get_current_org_id

    try:
        org_id = get_current_org_id()
        if not org_id:
            return jsonify({'success': False, 'message': 'No hay organización seleccionada'}), 400

        stats = generar_alertas_masivas(org_id, current_user.id)

        return jsonify({
            'success': True,
            'message': f"Se crearon {stats['notificaciones_creadas']} notificaciones nuevas",
            'stats': stats
        })

    except Exception as e:
        current_app.logger.error(f"Error en api_generar_alertas: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@inventario_bp.route('/alertas-stock')
@login_required
def alertas_stock():
    """Vista de alertas de stock bajo"""
    from services.stock_alerts_service import obtener_resumen_alertas
    from services.memberships import get_current_org_id

    org_id = get_current_org_id()
    if not org_id:
        flash('Debe seleccionar una organización primero', 'warning')
        return redirect(url_for('reportes.dashboard'))

    resumen = obtener_resumen_alertas(org_id)

    return render_template(
        'inventario/alertas_stock.html',
        resumen=resumen,
        conteo=resumen['conteo'],
        items=resumen['todos_los_items']
    )


@inventario_bp.route('/seed-items-ia', methods=['POST'])
@csrf.exempt
@login_required
def seed_items_ia():
    """Carga items de la calculadora IA al inventario de la organización."""
    if current_user.rol != 'administrador':
        return jsonify({'ok': False, 'error': 'Solo administradores'}), 403

    org_id = get_current_org_id() or current_user.organizacion_id
    if not org_id:
        return jsonify({'ok': False, 'error': 'Sin organización activa'}), 400

    try:
        from calculadora_ia import ETAPA_REGLAS_BASE, PRECIO_REFERENCIA

        # Recopilar items por etapa de construcción
        # codigo -> {descripcion, unidad, precio, etapa_nombre}
        items_data = {}

        for slug, regla in ETAPA_REGLAS_BASE.items():
            etapa_nombre = regla.get('nombre', slug.replace('-', ' ').title())

            for mat in regla.get('materiales', []):
                codigo = mat['codigo']
                if codigo not in items_data:
                    items_data[codigo] = {
                        'descripcion': mat['descripcion'],
                        'unidad': mat.get('unidad', 'unidades'),
                        'precio': PRECIO_REFERENCIA.get(codigo, 0),
                        'etapa': etapa_nombre,
                    }

            for mo in regla.get('mano_obra', []):
                codigo = mo['codigo']
                if codigo not in items_data:
                    items_data[codigo] = {
                        'descripcion': mo['descripcion'],
                        'unidad': mo.get('unidad', 'jornal'),
                        'precio': PRECIO_REFERENCIA.get(codigo, 0),
                        'etapa': etapa_nombre,
                    }

            for eq in regla.get('equipos', []):
                codigo = eq['codigo']
                if codigo not in items_data:
                    items_data[codigo] = {
                        'descripcion': eq['descripcion'],
                        'unidad': eq.get('unidad', 'día'),
                        'precio': PRECIO_REFERENCIA.get(codigo, 0),
                        'etapa': etapa_nombre,
                    }

        # Items de PRECIO_REFERENCIA que no están en ninguna etapa → "Otros"
        for codigo, precio in PRECIO_REFERENCIA.items():
            if codigo not in items_data:
                items_data[codigo] = {
                    'descripcion': codigo.replace('MAT-', '').replace('MO-', '').replace('EQ-', '').replace('-', ' ').title(),
                    'unidad': 'jornal' if codigo.startswith('MO-') else ('día' if codigo.startswith('EQ-') else 'unidades'),
                    'precio': precio,
                    'etapa': 'Otros',
                }

        # Crear categorías por etapa de construcción
        etapas_unicas = sorted(set(d['etapa'] for d in items_data.values()))
        cat_map = {}  # etapa_nombre -> category_id
        for idx, etapa_nombre in enumerate(etapas_unicas):
            cat = InventoryCategory.query.filter_by(
                company_id=org_id, nombre=etapa_nombre
            ).first()
            if not cat:
                cat = InventoryCategory(
                    company_id=org_id,
                    nombre=etapa_nombre,
                    is_active=True,
                    is_global=False,
                    sort_order=idx,
                )
                db.session.add(cat)
                db.session.flush()
            cat_map[etapa_nombre] = cat.id

        # Cargar todos los códigos existentes en UNA sola query
        codigos_existentes = set(
            r[0] for r in db.session.query(ItemInventario.codigo)
            .filter_by(organizacion_id=org_id)
            .filter(ItemInventario.codigo.in_(list(items_data.keys())))
            .all()
        )

        # Crear items nuevos en batch
        creados = 0
        omitidos = len(codigos_existentes)

        for codigo, data in sorted(items_data.items()):
            if codigo in codigos_existentes:
                continue

            item = ItemInventario(
                organizacion_id=org_id,
                categoria_id=cat_map.get(data['etapa']),
                codigo=codigo,
                nombre=data['descripcion'],
                descripcion=f"Etapa: {data['etapa']}",
                unidad=data['unidad'],
                precio_promedio=data['precio'],
                stock_actual=0,
                stock_minimo=0,
                activo=True,
            )
            db.session.add(item)
            creados += 1

            # Commit cada 50 items para evitar timeout
            if creados % 50 == 0:
                db.session.flush()

        db.session.commit()

        return jsonify({
            'ok': True,
            'creados': creados,
            'omitidos': omitidos,
            'total': len(items_data),
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error en seed_items_ia: {e}", exc_info=True)
        return jsonify({'ok': False, 'error': str(e)}), 500


@inventario_bp.route('/seed-constructoras', methods=['POST'])
@csrf.exempt
@login_required
def seed_constructoras():
    """Carga datos hardcodeados de constructoras de referencia (Cortes y Sistemas).
    No requiere upload de archivos - los datos ya están en el código."""
    if current_user.rol != 'administrador':
        return jsonify({'ok': False, 'error': 'Solo administradores'}), 403

    org_id = get_current_org_id() or current_user.organizacion_id
    if not org_id:
        return jsonify({'ok': False, 'error': 'Sin organizacion activa'}), 400

    try:
        from models.budgets import ItemReferenciaConstructora
        from seed_constructora_data import DATOS_CONSTRUCTORAS

        creados = 0
        actualizados = 0
        omitidos = 0

        for item_data in DATOS_CONSTRUCTORAS:
            existing = ItemReferenciaConstructora.query.filter_by(
                organizacion_id=org_id,
                constructora=item_data['constructora'],
                etapa_nombre=item_data['etapa_nombre'],
                codigo_excel=item_data['codigo'],
            ).first()

            if existing:
                if item_data['precio_unitario'] > 0:
                    existing.precio_unitario = item_data['precio_unitario']
                    existing.descripcion = item_data['descripcion']
                    existing.unidad = item_data['unidad']
                    actualizados += 1
                else:
                    omitidos += 1
            else:
                nuevo = ItemReferenciaConstructora(
                    organizacion_id=org_id,
                    constructora=item_data['constructora'],
                    etapa_nombre=item_data['etapa_nombre'],
                    codigo_excel=item_data['codigo'],
                    descripcion=item_data['descripcion'],
                    unidad=item_data['unidad'],
                    precio_unitario=item_data['precio_unitario'],
                    planilla=item_data.get('planilla', ''),
                )
                db.session.add(nuevo)
                creados += 1

        db.session.commit()

        return jsonify({
            'ok': True,
            'creados': creados,
            'actualizados': actualizados,
            'omitidos': omitidos,
            'total': len(DATOS_CONSTRUCTORAS),
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error en seed_constructoras: {e}", exc_info=True)
        return jsonify({'ok': False, 'error': str(e)}), 500


def limpiar_marca_item(nombre):
    """Elimina marcas comerciales conocidas del nombre de un item de inventario."""
    import re
    # Marcas comunes de construcción en Argentina
    marcas = [
        'Sinis-Kaufmann', 'Sinis Kaufmann', 'Kaufmann', 'Sinis',
        'Klaukol', 'Weber', 'Loma Negra', 'Acindar', 'Siderar', 'Ternium',
        'Isover', 'Durlock', 'Knauf', 'Tigre', 'Amanco', 'Rotoplas',
        'Cerro Negro', 'Ctibor', 'Fanelli', 'FV', 'Griferia FV',
        'Ferrum', 'Roca', 'DECA', 'Briggs', 'Piazza',
        'Aluar', 'Modena', 'De Angeli', 'Rehau', 'Veka',
        'Sherwin Williams', 'Alba', 'Tersuave', 'Sinteplast', 'Colorin',
        'Ceresita', 'Sika', 'Mapei', 'Plavicon',
        'Leiten', 'Peisa', 'Rheem', 'Orbis', 'Eskabe', 'Emege',
        'Daikin', 'Carrier', 'Samsung', 'LG', 'Midea', 'Surrey',
        'Schneider', 'Legrand', 'Siemens', 'ABB', 'Bticino',
        'Stanley', 'Black & Decker', 'DeWalt', 'Bosch', 'Makita',
        'BOMAG', 'Caterpillar', 'CAT', 'Volvo', 'Komatsu',
    ]
    resultado = nombre
    for marca in sorted(marcas, key=len, reverse=True):  # Más largas primero
        patron = re.compile(re.escape(marca), re.IGNORECASE)
        resultado = patron.sub('', resultado)
    # Limpiar espacios dobles, guiones sueltos, etc.
    resultado = re.sub(r'\s*-\s*-\s*', ' ', resultado)
    resultado = re.sub(r'\s{2,}', ' ', resultado)
    resultado = resultado.strip(' -,')
    return resultado


@inventario_bp.route('/importar-excel', methods=['POST'])
@login_required
def importar_excel():
    """Importa items de inventario desde un archivo Excel/CSV. Solo super_admin."""
    if not current_user.is_super_admin:
        return jsonify({'ok': False, 'error': 'Solo super admin'}), 403

    import openpyxl
    from io import BytesIO

    archivo = request.files.get('archivo')
    if not archivo:
        return jsonify({'ok': False, 'error': 'No se recibió archivo'}), 400

    org_id = get_current_org_id() or current_user.organizacion_id
    creados = 0
    actualizados = 0
    errores = 0

    try:
        contenido = archivo.read()
        nombre_archivo = archivo.filename.lower()

        filas = []
        if nombre_archivo.endswith('.csv'):
            import csv
            from io import StringIO
            texto = contenido.decode('utf-8', errors='replace')
            reader = csv.DictReader(StringIO(texto))
            filas = list(reader)
        else:
            wb = openpyxl.load_workbook(BytesIO(contenido), read_only=True, data_only=True)
            ws = wb.active

            # Leer headers de la primera fila
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                val = str(cell.value or '').strip().lower()
                # Normalizar nombres de columna
                if val in ('nombre', 'descripcion', 'descripción', 'material', 'item', 'articulo'):
                    headers.append('nombre')
                elif val in ('unidad', 'unid', 'u.m.', 'um', 'medida'):
                    headers.append('unidad')
                elif val in ('precio', 'precio_unitario', 'precio unitario', 'p.u.', 'pu', 'costo'):
                    headers.append('precio')
                elif val in ('stock', 'cantidad', 'cant', 'stock_actual', 'existencia'):
                    headers.append('stock')
                elif val in ('categoria', 'categoría', 'rubro', 'tipo', 'cat'):
                    headers.append('categoria')
                elif val in ('codigo', 'código', 'cod', 'sku', 'code'):
                    headers.append('codigo')
                else:
                    headers.append(val)

            for row in ws.iter_rows(min_row=2, values_only=True):
                fila = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        fila[headers[i]] = val
                if fila.get('nombre'):
                    filas.append(fila)
            wb.close()

        # Procesar filas
        for fila in filas:
            nombre_raw = str(fila.get('nombre', '')).strip()
            if not nombre_raw:
                continue
            nombre = limpiar_marca_item(nombre_raw)

            try:
                unidad = str(fila.get('unidad', 'u')).strip() or 'u'
                precio = float(fila.get('precio', 0) or 0)
                stock = float(fila.get('stock', 0) or 0)
                categoria_nombre = str(fila.get('categoria', '')).strip()
                codigo = str(fila.get('codigo', '')).strip()

                # Buscar si ya existe por nombre en esta org
                existente = ItemInventario.query.filter_by(
                    organizacion_id=org_id,
                    nombre=nombre
                ).first()

                if existente:
                    # Actualizar precio y stock si vienen con datos
                    if precio > 0:
                        existente.precio_promedio = precio
                    if stock > 0:
                        existente.stock_actual = stock
                    actualizados += 1
                else:
                    # Buscar o crear categoría
                    categoria_id = None
                    if categoria_nombre:
                        from models import InventoryCategory
                        cat = InventoryCategory.query.filter_by(
                            company_id=org_id,
                            nombre=categoria_nombre
                        ).first()
                        if not cat:
                            cat = InventoryCategory(
                                company_id=org_id,
                                nombre=categoria_nombre,
                                is_active=True
                            )
                            db.session.add(cat)
                            db.session.flush()
                        categoria_id = cat.id

                    item = ItemInventario(
                        organizacion_id=org_id,
                        nombre=nombre,
                        codigo=codigo or None,
                        unidad=unidad,
                        precio_promedio=precio,
                        stock_actual=stock,
                        stock_minimo=0,
                        categoria_id=categoria_id,
                        activo=True,
                    )
                    db.session.add(item)
                    creados += 1

            except Exception as e:
                current_app.logger.warning(f"Error importando fila '{nombre}': {e}")
                errores += 1

        db.session.commit()
        return jsonify({
            'ok': True,
            'creados': creados,
            'actualizados': actualizados,
            'errores': errores,
            'total': len(filas)
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception(f"Error importando Excel: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@inventario_bp.route('/limpiar-marcas', methods=['POST'])
@login_required
def limpiar_marcas():
    """Limpia marcas comerciales de todos los items del inventario. Solo super_admin."""
    if not current_user.is_super_admin:
        return jsonify({'ok': False, 'error': 'Solo super admin'}), 403

    org_id = get_current_org_id() or current_user.organizacion_id
    items = ItemInventario.query.filter_by(organizacion_id=org_id, activo=True).all()

    limpiados = 0
    for item in items:
        nombre_limpio = limpiar_marca_item(item.nombre)
        if nombre_limpio != item.nombre:
            item.nombre = nombre_limpio
            limpiados += 1

    db.session.commit()
    return jsonify({'ok': True, 'limpiados': limpiados, 'total': len(items)})


@inventario_bp.route('/deposito')
@login_required
def deposito():
    """Vista del depósito general — stock disponible para trasladar a obras."""
    if not current_user.puede_acceder_modulo('inventario'):
        flash('No tienes acceso al módulo de inventario.', 'danger')
        return redirect(url_for('home'))

    from models.inventory import StockObra
    from models.projects import Obra
    from sqlalchemy import func

    org_id = get_current_org_id() or current_user.organizacion_id

    # Items con stock > 0 en inventario general
    items = ItemInventario.query.filter(
        ItemInventario.organizacion_id == org_id,
        ItemInventario.activo == True,
        ItemInventario.stock_actual > 0
    ).order_by(ItemInventario.nombre).all()

    # Stock distribuido en obras (para info)
    stock_en_obras = db.session.query(
        StockObra.item_inventario_id,
        func.sum(StockObra.cantidad_disponible).label('total_en_obras')
    ).join(Obra, StockObra.obra_id == Obra.id)\
     .filter(Obra.organizacion_id == org_id)\
     .group_by(StockObra.item_inventario_id).all()
    stock_obras_map = {s.item_inventario_id: float(s.total_en_obras or 0) for s in stock_en_obras}

    # Obras disponibles para traslado
    obras = Obra.query.filter_by(
        organizacion_id=org_id
    ).filter(Obra.estado.in_(['en_curso', 'pendiente', 'activa'])
    ).order_by(Obra.nombre).all()

    # Movimientos recientes
    movimientos = MovimientoInventario.query.join(
        ItemInventario
    ).filter(
        ItemInventario.organizacion_id == org_id
    ).order_by(MovimientoInventario.fecha.desc()).limit(20).all()

    # Totales
    total_items = len(items)
    valor_total = sum(float(i.stock_actual or 0) * float(i.precio_promedio or 0) for i in items)

    return render_template('inventario/deposito.html',
                         items=items,
                         stock_obras_map=stock_obras_map,
                         obras=obras,
                         movimientos=movimientos,
                         total_items=total_items,
                         valor_total=valor_total)
